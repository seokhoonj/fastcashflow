"""Render the cookbook 9.5 sample files as spreadsheet-style PNGs.

These are the figures embedded in ``docs/cookbook/workflow/sample-walkthrough.md``.
They are NOT placeholders -- each one is a faithful, spreadsheet-styled render
of the packaged sample workbook / csv files and the close pack, drawn with
spreadsheet chrome (column letters A/B/C..., row numbers, grid, header band)
so the figures read like the Excel/CSV a user would actually open.

The script is self-contained: it exports the sample set, runs the close-pack
flow exactly as chapter 9.5 documents it, then renders the seven figures --
so the committed images stay in step with the engine and the sample data.

Run from the repo root with the project venv::

    .venv/bin/python docs/generate_sheet_images.py

Re-run it whenever the sample data, the engine output, or the close-pack
layout changes.
"""
from __future__ import annotations

import os
import string

import matplotlib

matplotlib.use("Agg")  # headless: render straight to file, no display needed

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import polars as pl
from matplotlib.patches import Rectangle

import fastcashflow as fcf

_HERE = os.path.dirname(os.path.abspath(__file__))
_IMAGES = os.path.join(_HERE, "images")
_SAMPLES = os.path.join(os.path.dirname(_HERE), "samples")

# Excel-ish palette
HDR_BAND = "#f2f2f2"   # column-letter / row-number band
HDR_ROW = "#d9e6f2"    # the data header row (row 1)
GRID = "#c0c0c0"
TEXT = "#202020"
FONT = "DejaVu Sans Mono"


def _col_letters(n):
    out = []
    for i in range(n):
        s, j = "", i
        while True:
            s = string.ascii_uppercase[j % 26] + s
            j = j // 26 - 1
            if j < 0:
                break
        out.append(s)
    return out


def render(matrix, path, title=None, max_rows=None):
    """matrix: list of rows (row 0 = data header), all cells str."""
    if max_rows is not None and len(matrix) > max_rows:
        matrix = matrix[:max_rows]
    n_rows = len(matrix)
    n_cols = max(len(r) for r in matrix)
    matrix = [list(r) + [""] * (n_cols - len(r)) for r in matrix]

    # column widths in characters (cap to keep figure sane)
    widths = []
    for c in range(n_cols):
        w = max(len(str(matrix[r][c])) for r in range(n_rows))
        widths.append(min(max(w, len(_col_letters(n_cols)[c])), 32))

    char_w = 0.105       # inch per character
    pad = 0.18           # inch cell horizontal padding
    row_h = 0.30         # inch per row
    rownum_w = 0.42      # left row-number gutter (inch)

    col_w = [w * char_w + pad for w in widths]
    fig_w = rownum_w + sum(col_w) + 0.2
    fig_h = (n_rows + 1) * row_h + (0.45 if title else 0.12) + 0.12

    fig = plt.figure(figsize=(fig_w, fig_h), dpi=200)
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.axis("off")

    top = fig_h - (0.40 if title else 0.06)
    if title:
        ax.text(rownum_w, fig_h - 0.22, title, fontsize=10, fontweight="bold",
                family="DejaVu Sans", color=TEXT, ha="left", va="center")

    letters = _col_letters(n_cols)

    def cell_x(c):
        return rownum_w + sum(col_w[:c])

    # column-letter band
    for c in range(n_cols):
        x = cell_x(c)
        ax.add_patch(Rectangle((x, top - row_h), col_w[c], row_h,
                               facecolor=HDR_BAND, edgecolor=GRID, lw=0.6))
        ax.text(x + col_w[c] / 2, top - row_h / 2, letters[c], fontsize=8,
                family="DejaVu Sans", color="#606060", ha="center", va="center")
    # corner
    ax.add_patch(Rectangle((0, top - row_h), rownum_w, row_h,
                           facecolor=HDR_BAND, edgecolor=GRID, lw=0.6))

    # data rows (incl. row-number gutter)
    for r in range(n_rows):
        y = top - row_h * (r + 2)
        # row number gutter
        ax.add_patch(Rectangle((0, y), rownum_w, row_h,
                               facecolor=HDR_BAND, edgecolor=GRID, lw=0.6))
        ax.text(rownum_w - 0.06, y + row_h / 2, str(r + 1), fontsize=7.5,
                family="DejaVu Sans", color="#606060", ha="right", va="center")
        for c in range(n_cols):
            x = cell_x(c)
            face = HDR_ROW if r == 0 else "white"
            ax.add_patch(Rectangle((x, y), col_w[c], row_h,
                                   facecolor=face, edgecolor=GRID, lw=0.6))
            val = str(matrix[r][c])
            if len(val) > 32:
                val = val[:29] + "..."
            # numbers right-align, text left-align; header always bold-ish
            try:
                float(val.replace(",", ""))
                ha, tx = "right", x + col_w[c] - 0.06
            except ValueError:
                ha, tx = "left", x + 0.06
            ax.text(tx, y + row_h / 2, val, fontsize=8, family=FONT,
                    color=TEXT, ha=ha, va="center",
                    fontweight="bold" if r == 0 else "normal")

    fig.savefig(path, dpi=200, facecolor="white")
    plt.close(fig)
    print("wrote", path, f"({n_rows}x{n_cols})")


def sheet_matrix(wb, name, max_rows=None):
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else _fmt(v) for v in row])
        if max_rows and len(rows) >= max_rows:
            break
    return rows


def csv_matrix(path):
    df = pl.read_csv(path)
    rows = [df.columns]
    for r in df.iter_rows():
        rows.append([_fmt(v) for v in r])
    return rows


def _fmt(v):
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return f"{v:,}"
    if isinstance(v, float):
        if abs(v) >= 1000 or v == int(v):
            return f"{round(v):,}"        # money -> 1,575,567 (no scientific)
        return f"{v:g}"                    # small floats (rates) -> 0.03
    return str(v)


def _build_samples():
    """Export the sample set and run the chapter 9.5 close-pack flow, so the
    inputs the figures render from exist on disk (samples/ is gitignored)."""
    fcf.samples.export(_SAMPLES, template="gmm", quiet=True)

    def s(name):
        return os.path.join(_SAMPLES, name)

    basis = fcf.read_basis(s("basis.xlsx"))
    model_points, state = fcf.read_inforce_policies(
        s("inforce_policies.csv"),
        coverages=s("coverages.csv"),
        calculation_methods=s("calculation_methods.csv"),
    )
    profitability = np.where(
        fcf.gmm.measure(model_points, basis).loss_component > 0.0,
        "onerous", "remaining")
    goc = fcf.settle_group_of_contracts(
        model_points, state, basis,
        period_months=12,
        coverage_units="count",
        profitability=profitability,
    )
    pack = fcf.close([fcf.reconcile(goc)])
    fcf.write_close_pack(pack, s("close_pack_2026Q1.xlsx"), movements=[goc])


def main():
    _build_samples()

    def out(name):
        return os.path.join(_IMAGES, name)

    def s(name):
        return os.path.join(_SAMPLES, name)

    wb = openpyxl.load_workbook(s("basis.xlsx"), read_only=True, data_only=True)
    render(sheet_matrix(wb, "segments"), out("sample-basis-segments.png"),
           title="basis.xlsx  --  segments")
    render(sheet_matrix(wb, "mortality_tables", max_rows=16),
           out("sample-basis-mortality.png"),
           title="basis.xlsx  --  mortality_tables  (top rows)")
    render(sheet_matrix(wb, "coverages"), out("sample-basis-coverages.png"),
           title="basis.xlsx  --  coverages")

    render(csv_matrix(s("policies.csv")), out("sample-policies.png"),
           title="policies.csv")
    render(csv_matrix(s("coverages.csv")), out("sample-coverages.png"),
           title="coverages.csv", max_rows=16)
    render(csv_matrix(s("inforce_state.csv")), out("sample-inforce-state.png"),
           title="inforce_state.csv")

    cp = openpyxl.load_workbook(s("close_pack_2026Q1.xlsx"),
                                read_only=True, data_only=True)
    render(sheet_matrix(cp, "01_SoFP"), out("close-pack-output.png"),
           title="close_pack_2026Q1.xlsx  --  01_SoFP")


if __name__ == "__main__":
    main()
