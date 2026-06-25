"""IFRS 17 disclosure assembly -- tidy frames over the settlement reconciliations.

The reporting layer that turns the footing settlement reconciliations into a
presentable close pack. This module is pure assembly + serialization over numbers
the settlement layer already produces and foots -- no measurement, no kernels.

Two ideas anchor it:

* a CANONICAL tidy frame -- one row per disclosure line -- so a reconciliation can
  leave Python as data (a sheet / PDF template / audit join consumes a tidy frame,
  not a __str__). The frame is LEAN: ``(model, group_id, statement, period_start,
  period_end, block, line, amount)``. The richer audit columns (line_code, the
  IFRS 17 paragraph anchor, the memo flag, the sort order) are reference data keyed
  on the line, materialised by the emitter at the contract boundary -- not
  denormalised onto every in-process row.
* the block SPEC -- ``_*_RECON_BLOCKS`` -- is the single source for the line spine:
  the ordered (block -> lines) structure, each line carrying its display name,
  the reconciliation field it reads, its paragraph anchor and whether it is a P&L
  memo (outside the balance recursion). The Phase-0 oracle pins that the spec
  covers exactly the reconciliation's fields, so a serialized line cannot drift
  from the reconciliation it explains.
"""
from __future__ import annotations

from functools import singledispatch

import polars as pl

from fastcashflow._measurement.model import model_tag
from fastcashflow.io import _write_frame, write_measurement
import fastcashflow.gmm._results as _gmm
import fastcashflow._paa as _paa
import fastcashflow._reinsurance as _reinsurance
import fastcashflow.vfa._results as _vfa
from fastcashflow.gmm._results import _GMM_RECON_BLOCKS
from fastcashflow._paa import _PAA_RECON_BLOCKS
from fastcashflow._reinsurance import _REINSURANCE_RECON_BLOCKS
from fastcashflow.vfa._results import _VFA_RECON_BLOCKS

# The lean canonical schema returned by reconciliation_to_frame / to_frame.
_LEAN_COLUMNS = (
    "model", "group_id", "statement", "period_start", "period_end",
    "block", "line", "amount",
)
_LEAN_SCHEMA = {
    "model": pl.Utf8, "group_id": pl.Utf8, "statement": pl.Utf8,
    "period_start": pl.Int64, "period_end": pl.Int64,
    "block": pl.Utf8, "line": pl.Utf8, "amount": pl.Float64,
}
# The emitted artifact (file / sheet / audit join) is self-contained: the rich
# reference columns -- the machine line_code, the IFRS 17 paragraph anchor, the
# P&L-memo flag and the deterministic order -- are MATERIALISED at the write
# boundary, not denormalised onto the lean in-process frame.
_RICH_SCHEMA = {
    "model": pl.Utf8, "group_id": pl.Utf8, "statement": pl.Utf8,
    "period_start": pl.Int64, "period_end": pl.Int64,
    "block": pl.Utf8, "line": pl.Utf8, "line_code": pl.Utf8,
    "ifrs17_paragraph": pl.Utf8, "is_memo": pl.Boolean, "sort_order": pl.Int64,
    "amount": pl.Float64,
}

# The settlement reconciliation block specs (_*_RECON_BLOCKS) are the single
# source for each family's line spine -- the ordered (block -> lines) structure
# with each line's display name, reconciliation field, IFRS 17 paragraph and
# P&L-memo flag. They live in the owning model modules (next to each settlement reconciliation
# class, whose __str__ renders from them) and are imported here so the printed
# table and this disclosure frame never drift.


def _recon_frame(recon, blocks, *, rich: bool = False) -> pl.DataFrame:
    """The canonical tidy frame for one reconciliation: one row per disclosure
    line, read from the block spec so the spine has a single source. period_start
    / period_end are the relative period (the close assembler re-stamps absolute
    reporting positions when stacking a schedule). ``rich`` adds the audit
    reference columns for an emitted, self-contained artifact."""
    rows = []
    order = 0
    for block, lines in blocks:
        for line, field, para, memo in lines:
            row = {
                "model": recon.model, "group_id": None, "statement": "settlement",
                "period_start": 0, "period_end": int(recon.period_months),
                "block": block, "line": line, "amount": float(getattr(recon, field)),
            }
            if rich:
                row.update({"line_code": field, "ifrs17_paragraph": para,
                            "is_memo": memo, "sort_order": order})
            rows.append(row)
            order += 1
    return pl.DataFrame(rows, schema=_RICH_SCHEMA if rich else _LEAN_SCHEMA)


@singledispatch
def reconciliation_to_frame(recon) -> pl.DataFrame:
    """Return the lean canonical tidy frame for a settlement reconciliation."""
    raise TypeError(
        f"reconciliation_to_frame: no disclosure spec for {model_tag(recon)}")


@reconciliation_to_frame.register
def _(recon: _gmm.SettlementReconciliation) -> pl.DataFrame:
    return _recon_frame(recon, _GMM_RECON_BLOCKS)


@reconciliation_to_frame.register
def _(recon: _vfa.SettlementReconciliation) -> pl.DataFrame:
    return _recon_frame(recon, _VFA_RECON_BLOCKS)


@reconciliation_to_frame.register
def _(recon: _reinsurance.SettlementReconciliation) -> pl.DataFrame:
    return _recon_frame(recon, _REINSURANCE_RECON_BLOCKS)


@reconciliation_to_frame.register
def _(recon: _paa.SettlementReconciliation) -> pl.DataFrame:
    return _recon_frame(recon, _PAA_RECON_BLOCKS)


# (model, block spec, reconciliation class) for the four settlement families --
# the single registry the spec-covers-fields oracle iterates.
_RECON_SPECS = (
    ("gmm", _GMM_RECON_BLOCKS, _gmm.SettlementReconciliation),
    ("vfa", _VFA_RECON_BLOCKS, _vfa.SettlementReconciliation),
    ("reinsurance", _REINSURANCE_RECON_BLOCKS, _reinsurance.SettlementReconciliation),
    ("paa", _PAA_RECON_BLOCKS, _paa.SettlementReconciliation),
)
_SPEC_BY_TYPE = {cls: (model, blocks) for model, blocks, cls in _RECON_SPECS}


def write_reconciliation(reconciliation, path) -> None:
    """Serialize a settlement reconciliation (or a list of them, one per
    reporting period) to a tidy file -- parquet / csv / xlsx -- with the rich
    audit columns materialised. A list is stacked with a 0-based ``period_index``
    so a multi-period close schedule round-trips as one long frame.

    Mirrors :func:`write_measurement` (which serializes the per-MP movements);
    this is the disclosure-shaped (reconciliation / close) serialization path.
    """
    recons = (list(reconciliation)
              if isinstance(reconciliation, (list, tuple)) else [reconciliation])
    frames = []
    for i, recon in enumerate(recons):
        _, blocks = _SPEC_BY_TYPE[type(recon)]
        frame = _recon_frame(recon, blocks, rich=True)
        frames.append(frame.with_columns(pl.lit(i, dtype=pl.Int64).alias("period_index")))
    _write_frame(pl.concat(frames), str(path))


def line_metadata() -> pl.DataFrame:
    """The disclosure line registry as a frame -- (model, block, line, line_code,
    ifrs17_paragraph, is_memo, sort_order) -- the single source the lean
    :func:`reconciliation_to_frame` and the emitter both read. Exposed so a user
    can join the reference columns onto a lean frame themselves."""
    rows = []
    for model, blocks, _cls in _RECON_SPECS:
        order = 0
        for block, lines in blocks:
            for line, field, para, memo in lines:
                rows.append({
                    "model": model, "block": block, "line": line,
                    "line_code": field, "ifrs17_paragraph": para,
                    "is_memo": memo, "sort_order": order})
                order += 1
    return pl.DataFrame(rows)


# The close pack's sheet order. The aggregate statements an entity reads; the
# per-model-point movement detail goes to a parquet sidecar (Excel's row limit).
_CLOSE_PACK_SHEETS = (
    ("00_Index", None),                      # cover metadata, built below
    ("01_SoFP", "sofp"),
    ("02_Service_Result", "service_result"),  # only if assembled
    ("03_Finance", "finance"),
    ("04_Reconciliation", "reconciliation"),  # the rich tidy detail
)


def _append_sheet(workbook, title: str, frame: pl.DataFrame) -> None:
    """Append a polars frame to a new sheet -- header row then data rows."""
    sheet = workbook.create_sheet(title=title)
    sheet.append(list(frame.columns))
    for row in frame.iter_rows():
        sheet.append(list(row))


def _index_frame(period_months: int, reconciliation: pl.DataFrame,
                 sheets, sidecar) -> pl.DataFrame:
    """The 00_Index cover sheet -- the reporting period, the models and groups
    in the pack, the sheet list, and the per-MP sidecar reference."""
    models = ", ".join(reconciliation["model"].unique().sort().to_list())
    groups = [g for g in reconciliation["group_id"].unique().sort().to_list()
              if g is not None]
    rows = [
        {"item": "Reporting period (months)", "value": str(period_months)},
        {"item": "Models", "value": models},
        {"item": "Groups", "value": ", ".join(groups) if groups else "(unlabelled)"},
        {"item": "Sheets", "value": ", ".join(sheets)},
        {"item": "Per-MP detail",
         "value": sidecar if sidecar else "(not included)"},
    ]
    return pl.DataFrame(rows, schema={"item": pl.Utf8, "value": pl.Utf8})


def write_close_pack(package, path, *, movements=None) -> None:
    """Write a close pack (a :class:`~fastcashflow.closing.ClosePackage`) to a
    multi-sheet ``.xlsx`` -- the aggregate IFRS 17 statements an entity reads --
    and, when ``movements`` is given, a per-model-point parquet sidecar.

    The workbook carries an index cover, the statement of financial position, the
    service result (if it was assembled), the finance statement, and the
    reconciliation detail with the rich audit columns (line_code, the IFRS 17
    paragraph anchor, the memo flag, the deterministic order) materialised by
    joining :func:`line_metadata` -- so the artifact is self-contained.

    The per-model-point settlement movement does NOT go in the workbook (a sheet
    caps at ~1,048,576 rows); ``movements`` -- one settlement movement or a list
    -- is written to ``<path>_per_mp[_i].parquet`` beside the workbook via
    :func:`write_measurement`, and the index sheet names the file(s).
    """
    p = str(path)
    if not p.endswith(".xlsx"):
        raise ValueError(
            f"write_close_pack: path must be a .xlsx workbook, got {path!r}")
    import openpyxl

    frames = package.to_frames()
    stem = p[:-len(".xlsx")]

    # Per-MP sidecar(s) first, so the index can name them. The naming keys off
    # the CALL SHAPE, not the count: a single movement -> one bare file; a list
    # (or tuple) -> one indexed file per entry, even if it holds just one.
    sidecar_label = None
    movement_list = None
    if movements is not None:
        is_single = not isinstance(movements, (list, tuple))
        movement_list = [movements] if is_single else list(movements)
        sidecar_paths = ([f"{stem}_per_mp.parquet"] if is_single
                         else [f"{stem}_per_mp_{i}.parquet"
                               for i in range(len(movement_list))])
        sidecar_label = ", ".join(sp.rsplit("/", 1)[-1] for sp in sidecar_paths)

    # The sheets actually present (service result only if assembled).
    present = [(title, key) for title, key in _CLOSE_PACK_SHEETS
               if key is None or key in frames]
    sheet_titles = [title for title, _key in present]

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)        # drop the default empty sheet
    for title, key in present:
        if key is None:
            frame = _index_frame(package.period_months, frames["reconciliation"],
                                 sheet_titles, sidecar_label)
        elif key == "reconciliation":
            # Materialise the rich audit columns at the emit boundary (DELTA 2):
            # the lean in-process frame stays lean; the file is self-contained.
            frame = frames[key].join(
                line_metadata(), on=["model", "block", "line"], how="left")
        else:
            frame = frames[key]
        _append_sheet(workbook, title, frame)
    workbook.save(p)

    if movement_list is not None:
        for movement, sidecar_path in zip(movement_list, sidecar_paths):
            write_measurement(movement, sidecar_path)
