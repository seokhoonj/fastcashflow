"""File I/O for model points, the actuarial basis and valuation results.

Model points and results go through polars; the actuarial basis -- read by
:func:`read_basis` -- comes from an Excel workbook via openpyxl.

Model points come in two shapes, both producing the same ``ModelPoints``:

* **wide** -- one row per policy, every benefit a column:
  ``<coverage>_benefit`` per rate-driven coverage, plus the survival
  benefits ``maturity_benefit`` and ``annuity_payment``. The convenient
  form for a single, homogeneous product.
* a policies frame (contract attributes) plus a coverages
  frame, one row per policy x coverage carrying ``amount`` and ``premium``.
  The form for a heterogeneous, multi-product portfolio.

:func:`read_model_points` reads either. The engine ships no ModelPoints ->
file exporter: both forms are lossy projections (they cannot carry per-coverage
waiting / reduction rules, ``issue_class``, ``elapsed_months`` or the VFA
account fields), so they are accepted only as external input, never produced.

The core engine stays identifier-free: the kernel never needs a policy id, so
none is carried through ``ModelPoints`` or ``Measurement``. Identifiers are a
file-boundary concern -- pass them to :func:`write_measurement` (or via
``measure_stream``'s ``id_column``) to join results back to policies.
"""
from __future__ import annotations

import importlib.resources as resources
import warnings
from functools import singledispatch
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np
import openpyxl
import polars as pl

from fastcashflow._measurement.model import model_tag, supported_model_tags
from fastcashflow._typing import FloatArray, IntArray
from fastcashflow.basis import (
    Basis, BasisRouter, CoverageRate, ExpenseItem,
)
from fastcashflow.state_model import STATE_MODELS
from fastcashflow.coverage import (
    CalculationMethod, RATE_DRIVEN_METHODS,
)
from fastcashflow.model_points import (
    STATE_ACTIVE, STATE_NAMES, NO_GUARANTEE_RATE, ModelPoints,
)

# ``engine`` is the largest module in the package (codegen + the numba CPU
# kernels) and importing it at module load pulls all of that into any
# downstream that needs the I/O layer. The two engine names used here --
# ``Measurement`` for write_measurement's type hint and ``measure`` for the
# ``measure_stream`` stream -- are imported under TYPE_CHECKING (for the hint)
# and lazily inside ``measure_stream`` (for the call), so a script that only
# reads model points or writes a results frame never imports engine.py.
if TYPE_CHECKING:  # pragma: no cover -- import only for type hints
    from fastcashflow._measurement.gmm import Measurement


def _read_frame(path) -> pl.DataFrame:
    p = str(path)
    if p.endswith(".parquet"):
        return pl.read_parquet(p)
    if p.endswith(".csv"):
        return pl.read_csv(p)
    if p.endswith(".xlsx"):
        return pl.read_excel(p, engine="openpyxl")
    if p.endswith((".feather", ".arrow")):
        return pl.read_ipc(p)
    raise ValueError(
        f"unsupported file type: {path!r} "
        "(expected .parquet, .csv, .xlsx or .feather)"
    )


def _write_frame(df: pl.DataFrame, path) -> None:
    p = str(path)
    if p.endswith(".parquet"):
        df.write_parquet(p)
    elif p.endswith(".csv"):
        df.write_csv(p)
    elif p.endswith((".feather", ".arrow")):
        df.write_ipc(p)
    elif p.endswith(".xlsx"):
        # Single-sheet xlsx via openpyxl (polars's own write_excel needs an
        # extra ``xlsxwriter`` dependency we do not require). The reader
        # side already accepts .xlsx via ``_read_frame``.
        # Caveat: openpyxl reads an all-None row back as a blank separator, so
        # a fully-null data row would not survive an xlsx round-trip. Use csv,
        # parquet or feather for round-tripping frames that may carry such rows.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(list(df.columns))
        for row in df.iter_rows():
            ws.append(list(row))
        wb.save(p)
    else:
        raise ValueError(
            f"unsupported file type: {path!r} "
            "(expected .parquet, .csv, .xlsx or .feather)"
        )


# ---------------------------------------------------------------------------
# Actuarial basis -- the basis workbook
# ---------------------------------------------------------------------------
#
# A single workbook (``basis.xlsx``) carries every assumption the engine
# needs. Ten sheets:
#
#   * ``segments``       -- (product, channel) -> which tables + scalar params
#                           (a ``_DEFAULTS`` row that blank cells inherit).
#   * ``coverages``      -- (product) -> coverage, type, optional rate_table.
#   * ``mortality_tables``, ``incidence_rate_tables``, ``waiver_tables``,
#     ``lapse_tables``, ``discount_tables``, ``surrender_value_tables``,
#     ``expense_tables``, ``inflation_tables`` -- the named rate tables the
#     segments reference.
#
# See docs/basis-format.md for the column-level schema and
# docs/naming-conventions.md for the value-case rules.
#
# v1 limitation (refined in a later round): the discount, inflation and
# maintenance tables are read but used flat (their first entry). The reader
# returns ``{(product, channel): Basis}`` -- splitting model points by
# segment and valuing each is left to the caller.


def _sheet_dicts(ws):
    """Yield each data row of a worksheet as a dict keyed by the header row."""
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    names = [str(h).strip() if h is not None else "" for h in header]
    for row in rows:
        if all(c is None for c in row):
            continue
        yield {n: v for n, v in zip(names, row) if n}


def _require_sheet(wb, sheet_name):
    """Return ``wb[sheet_name]`` or raise a friendly error.

    Without this wrap a missing required sheet surfaces as a raw openpyxl
    ``KeyError(sheet_name)`` -- non-obvious to a non-programmer actuary
    reading the traceback.
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"basis workbook is missing required sheet "
            f"{sheet_name!r}; known sheets: {sorted(wb.sheetnames)}"
        )
    return wb[sheet_name]


def _require_row_cols(row, required, *, sheet, table_id=None, what="row"):
    """Raise a friendly error if any required column is missing from ``row``.

    ``row`` is a dict yielded by :func:`_sheet_dicts`. Without this wrap a
    missing column surfaces as a bare ``KeyError(col)`` deep in the
    reader -- the user gets a column name but no context.
    """
    missing = [c for c in required if c not in row]
    if not missing:
        return
    ctx = f" (table_id={table_id!r})" if table_id is not None else ""
    raise ValueError(
        f"sheet {sheet!r}{ctx}: {what} is missing required column(s) "
        f"{missing}; row has columns {sorted(row)}"
    )


def _truncate_list(items, cap=10):
    """Format a list capped at ``cap`` items with a ``(... and N more)`` suffix.

    Used for not-found errors that enumerate the registered alternatives --
    a workbook with 100+ table_ids would otherwise produce an unreadable
    multi-line traceback.
    """
    items = list(items)
    if len(items) <= cap:
        return repr(items)
    extra = len(items) - cap
    return f"{items[:cap]!r} (... and {extra} more)"


# Axes a rate table may carry, in the order they index the internal grid.
# A sheet may include any subset; missing axes broadcast (the rate is held
# flat over that axis at lookup time). ``age`` (attained) is mutually
# exclusive with ``issue_age`` / ``duration`` (select-and-ultimate schema).
# ``issue_class`` is the at-issue classification axis (occupational / UW
# class) -- absent from most tables, broadcasts to a no-op when absent.
# ``elapsed`` is the semi-Markov sojourn axis (state-duration in years
# since entering the source state) -- carried by re-incidence /
# post-event mortality tables; broadcasts to a no-op when absent.
_RATE_AXES = ("sex", "issue_age", "duration", "age", "issue_class", "elapsed")


def _flex_rate_table(ws, *, value_col="rate"):
    """Schema-detecting rate-table reader -- returns ``{table_id: callable}``.

    The sheet may carry any subset of ``_RATE_AXES`` plus ``table_id`` and
    ``value_col`` (``rate`` or ``amount``). The reader detects which axes are
    present and returns a callable per table with the standard
    ``(sex, issue_age, duration)`` signature; axes not in the sheet broadcast
    (the rate is held flat over them), and lookups past the table's range
    clip to the edge.

    Supported schemas (any subset of ``{sex, age, issue_age, duration}``):

    * ``[rate]``                          -- flat scalar
    * ``[age, rate]``                     -- by attained age, sex broadcast
    * ``[sex, age, rate]``                -- by sex x age (the historical default)
    * ``[duration, rate]``                -- by duration, sex / age broadcast (lapse)
    * ``[sex, issue_age, duration, rate]`` -- full select-and-ultimate

    A sheet mixing ``age`` (attained) with ``issue_age`` / ``duration``
    (select schema) is rejected -- pick one parameterisation.
    """
    rows = list(_sheet_dicts(ws))
    if not rows:
        return {}
    header = set(rows[0].keys())
    if "table_id" not in header:
        raise ValueError(
            f"sheet {ws.title!r} is missing required column 'table_id'; "
            f"row has columns {sorted(header)}"
        )
    if value_col not in header:
        raise ValueError(
            f"sheet {ws.title!r} is missing required column {value_col!r} "
            f"(every rate-table row carries the rate in this column); "
            f"row has columns {sorted(header)}"
        )
    axes = tuple(a for a in _RATE_AXES if a in header)
    if "age" in axes and ("issue_age" in axes or "duration" in axes):
        raise ValueError(
            f"sheet {ws.title!r} mixes 'age' (attained) with "
            "'issue_age' / 'duration' (select schema) -- pick one"
        )

    by_id: dict[str, dict[tuple, float]] = {}
    for r in rows:
        tid = str(r["table_id"]).strip()
        try:
            key = tuple(int(r[a]) for a in axes)
        except KeyError as exc:
            raise ValueError(
                f"sheet {ws.title!r} table {tid!r}: row is missing axis "
                f"column {exc.args[0]!r} (header declares axes {axes!r}, "
                "so every row must populate them)"
            ) from None
        bucket = by_id.setdefault(tid, {})
        if key in bucket:
            raise ValueError(
                f"sheet {ws.title!r} table {tid!r}: duplicate row at "
                f"{dict(zip(axes, key))} -- a rate table must have one "
                "entry per axis combination (the last row would silently "
                "overwrite the first)"
            )
        bucket[key] = float(r[value_col])
    return {tid: _build_rate_callable(axes, list(entries.items()), ws.title, tid)
            for tid, entries in by_id.items()}


def _build_rate_callable(axes, entries, sheet_title, table_id):
    """Pack rows into a dense numpy grid and wrap in a lookup closure."""
    if not axes:
        # Flat scalar table -- one row, one rate.
        if len(entries) != 1:
            raise ValueError(
                f"sheet {sheet_title!r} table {table_id!r}: a flat (axis-less) "
                f"table must have exactly one row, got {len(entries)}"
            )
        val = entries[0][1]

        def rate(sex, issue_age, duration, issue_class, elapsed):
            shape = np.broadcast_shapes(
                np.asarray(sex).shape, np.asarray(issue_age).shape,
                np.asarray(duration).shape, np.asarray(issue_class).shape,
                np.asarray(elapsed).shape,
            )
            return np.full(shape, val, dtype=np.float64)
        rate._fcf_table_id = table_id
        rate._fcf_sheet = sheet_title
        rate._fcf_modifiers = ()
        return rate

    keys = np.array([k for k, _ in entries], dtype=np.int64)
    values = np.array([v for _, v in entries], dtype=np.float64)
    mins = keys.min(axis=0)
    maxs = keys.max(axis=0)
    shape = tuple(int(maxs[i] - mins[i] + 1) for i in range(len(axes)))
    grid = np.full(shape, np.nan, dtype=np.float64)
    for k, v in zip(keys, values):
        idx = tuple(int(k[i] - mins[i]) for i in range(len(axes)))
        grid[idx] = v
    if np.isnan(grid).any():
        raise ValueError(
            f"sheet {sheet_title!r} table {table_id!r} is not dense over its "
            f"axes {axes} -- some cells in the cartesian product are missing"
        )

    def rate(sex, issue_age, duration, issue_class, elapsed):
        sex = np.asarray(sex, dtype=np.int64)
        issue_age = np.asarray(issue_age, dtype=np.int64)
        duration = np.asarray(duration, dtype=np.int64)
        issue_class = np.asarray(issue_class, dtype=np.int64)
        elapsed = np.asarray(elapsed, dtype=np.int64)
        # One index array per axis present in the table.
        idxs = []
        for i, a in enumerate(axes):
            if a == "sex":
                raw = sex
            elif a == "age":
                raw = issue_age + duration                # attained age
            elif a == "issue_age":
                raw = issue_age
            elif a == "issue_class":
                raw = issue_class
            elif a == "elapsed":
                raw = elapsed
            else:                                          # duration
                raw = duration
            idxs.append(np.clip(raw - int(mins[i]), 0, shape[i] - 1))
        # Broadcast each index to the input's full broadcast shape so that
        # numpy fancy-indexing returns a result of that shape (axes absent
        # from the table contribute through broadcast, not indexing).
        target = np.broadcast_shapes(
            sex.shape, issue_age.shape, duration.shape,
            issue_class.shape, elapsed.shape,
        )
        return grid[tuple(np.broadcast_to(ix, target) for ix in idxs)]
    rate._fcf_table_id = table_id
    rate._fcf_sheet = sheet_title
    rate._fcf_modifiers = ()
    return rate


def _rate_fn_from_records(records, *, value_col="rate", where="inline rate table"):
    """Build one ``RateFn`` from row-dicts -- the in-memory (DataFrame) path.

    Mirrors ``_flex_rate_table`` but for a single table given as a list of
    ``{column: value}`` dicts (a polars / pandas DataFrame's rows), with no
    ``table_id`` grouping. Axes are auto-detected from the columns present
    (any subset of ``_RATE_AXES``); the rate is read from ``value_col``. Reuses
    ``_build_rate_callable`` so the resulting callable is byte-identical to the
    workbook path for the same numbers.
    """
    rows = list(records)
    if not rows:
        raise ValueError(f"{where}: no rows (an empty rate table)")
    header = set(rows[0].keys())
    if value_col not in header:
        raise ValueError(
            f"{where}: missing required column {value_col!r} (the rate); "
            f"columns are {sorted(header)}"
        )
    axes = tuple(a for a in _RATE_AXES if a in header)
    if "age" in axes and ("issue_age" in axes or "duration" in axes):
        raise ValueError(
            f"{where}: mixes 'age' (attained) with 'issue_age' / 'duration' "
            "(select schema) -- pick one parameterisation"
        )
    entries: dict[tuple, float] = {}
    for r in rows:
        key = tuple(int(r[a]) for a in axes)
        if key in entries:
            raise ValueError(
                f"{where}: duplicate row at {dict(zip(axes, key))} -- one "
                "entry per axis combination"
            )
        entries[key] = float(r[value_col])
    return _build_rate_callable(axes, list(entries.items()), where, "inline")


def _read_expense_tables(ws) -> dict[str, tuple[ExpenseItem, ...]]:
    """Read the optional ``expense_tables`` sheet.

    Each row is one ``ExpenseItem`` -- the item-form expense ledger the
    engine dispatches on. Columns: ``table_id``, ``category``, ``base``,
    ``value``. The same ``table_id`` may span multiple rows (an acquisition
    row plus a maintenance row, plus a claims row, ...).
    Returns ``{table_id: tuple[ExpenseItem, ...]}`` for the
    segments-side ``expense_table`` lookup to consume. Inflation is
    *not* a row attribute -- it lives on the segment as the global
    economic ``expense_inflation`` curve (see :data:`inflation_tables`
    sheet).
    """
    by_id: dict[str, list[ExpenseItem]] = {}
    first = True
    for r in _sheet_dicts(ws):
        if first:
            _require_row_cols(
                r, ("table_id", "category", "base", "value"),
                sheet=ws.title,
            )
            first = False
        tid = str(r["table_id"]).strip()
        by_id.setdefault(tid, []).append(ExpenseItem(
            category=str(r["category"]).strip(),
            base=str(r["base"]).strip(),
            value=float(r["value"]),
        ))
    return {tid: tuple(rows) for tid, rows in by_id.items()}


def _read_ae_factors(ws):
    """Read the optional ``ae_factors`` sheet.

    Each row is one (product, channel, coverage) -> factor (a runtime
    multiplier on the base rate). Optional axis columns
    ``{sex, age, issue_age, duration}`` let the factor vary along those
    dimensions (same schema-detection rules as the base rate tables); missing
    axes broadcast. ``channel`` empty matches the segment whose channel is
    blank (a single-segment workbook).

    Returns ``(factors, seg_axes)`` where ``factors`` is
    ``{(*seg_axis_values, coverage): callable(sex, issue_age, duration,
    issue_class, elapsed) -> factor}`` and ``seg_axes`` are the segment-axis
    columns the sheet declares -- any subset of the segments' axes, so an A/E
    calibrated coarsely (just ``product``) broadcasts over the finer
    routing axes. Missing / empty sheet -> ``({}, ())`` -> no A/E adjustment.
    """
    rows = list(_sheet_dicts(ws))
    if not rows:
        return {}, ()
    header = list(rows[0].keys())
    _require_row_cols(
        rows[0], ("product", "coverage", "factor"), sheet=ws.title,
    )
    # Rate axes (the factor's shape, like a rate table); everything else that is
    # not coverage / factor is a segment axis (which segment+coverage the
    # factor applies to).
    rate_axes = tuple(a for a in _RATE_AXES if a in header)
    if "age" in rate_axes and ("issue_age" in rate_axes or "duration" in rate_axes):
        raise ValueError(
            f"sheet {ws.title!r} mixes 'age' (attained) with "
            "'issue_age' / 'duration' (select schema) -- pick one"
        )
    seg_axes = tuple(c for c in header
                     if c not in _RATE_AXES and c not in ("coverage", "factor")
                     and not str(c).endswith("_name"))

    by_key: dict[tuple, list] = {}
    for r in rows:
        seg_key = tuple(str(r.get(a, "") or "").strip() for a in seg_axes)
        coverage = str(r["coverage"]).strip()
        key = seg_key + (coverage,)
        try:
            axes_key = tuple(int(r[a]) for a in rate_axes)
        except KeyError as exc:
            raise ValueError(
                f"sheet {ws.title!r} row for {key!r} is missing axis "
                f"column {exc.args[0]!r} (header declares rate axes {rate_axes!r})"
            ) from None
        by_key.setdefault(key, []).append((axes_key, float(r["factor"])))
    factors = {
        key: _build_rate_callable(rate_axes, entries, ws.title,
                                  "/".join(map(str, key)))
        for key, entries in by_key.items()
    }
    return factors, seg_axes


def _propagate_table_id(wrapper, inner, modifier_tag):
    """Carry the source table_id from an inner rate callable to a wrapper."""
    tid = getattr(inner, "_fcf_table_id", None)
    if tid is None:
        return
    wrapper._fcf_table_id = tid
    wrapper._fcf_sheet = getattr(inner, "_fcf_sheet", None)
    wrapper._fcf_modifiers = getattr(inner, "_fcf_modifiers", ()) + (modifier_tag,)


def _with_improvement(rate_fn, improvement_curve):
    """Wrap a rate callable to multiply by an annual improvement factor.

    ``improvement_curve`` is a ``(n_years,)`` array indexed by policy year
    (= duration). ``factor[0] = 1.0`` typically, decreasing for genuine
    improvement (mortality falls). Held flat past the curve's end.
    ``None`` returns ``rate_fn`` unchanged.
    """
    if rate_fn is None or improvement_curve is None:
        return rate_fn
    n = improvement_curve.shape[0]

    def improved(sex, issue_age, duration, issue_class, elapsed):
        d = np.asarray(duration, dtype=np.int64)
        idx = np.clip(d, 0, n - 1)
        return (rate_fn(sex, issue_age, duration, issue_class, elapsed)
                * improvement_curve[idx])
    _propagate_table_id(improved, rate_fn, "improvement")
    return improved


def _with_ae_factor(rate_fn, factor_fn):
    """Wrap a rate callable to multiply by an A/E factor at call time.

    ``factor_fn`` shares the ``(sex, issue_age, duration) -> array``
    signature; ``None`` (no factor configured for this coverage) returns
    ``rate_fn`` unchanged.
    """
    if factor_fn is None or rate_fn is None:
        return rate_fn

    def adjusted(sex, issue_age, duration, issue_class, elapsed):
        return (rate_fn(sex, issue_age, duration, issue_class, elapsed)
                * factor_fn(sex, issue_age, duration, issue_class, elapsed))
    _propagate_table_id(adjusted, rate_fn, "ae")
    return adjusted


def _with_age_shift(rate_fn, shift):
    """Wrap a rate callable to shift its ``issue_age`` argument by ``shift``.

    A positive shift treats every life as ``shift`` years older when looking
    up the base table; negative shifts make them younger. Returns ``rate_fn``
    unchanged when ``shift == 0`` (no allocation cost). ``rate_fn`` may be
    ``None`` (an optional rate the segment did not configure), in which case
    the wrapper is a no-op too.
    """
    if rate_fn is None or shift == 0:
        return rate_fn

    def shifted(sex, issue_age, duration, issue_class, elapsed):
        return rate_fn(sex, issue_age + shift, duration, issue_class, elapsed)
    _propagate_table_id(shifted, rate_fn, f"shift{shift:+d}")
    return shifted


def _surrender_value_col(ws) -> str:
    """Detect the surrender table's value column. ``amount`` -- a surrender
    amount by duration (per policy or per unit of the MP's
    surrender_base_amount -> an amount mode); ``factor`` -- a factor on
    cumulative premium (cum_premium_factor, the legacy default)."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    cols = {str(c) for c in header if c is not None}
    return "amount" if "amount" in cols else "factor"


def _axis_tables(ws, axis, *, value_col="rate"):
    """``{table_id: value array}`` from a sheet keyed by ``axis`` (0-based).

    ``value_col`` names the column carrying the per-axis value -- ``"rate"``
    for rate / probability sheets, ``"amount"`` for currency sheets
    (maintenance expense). The column-name distinction documents units;
    a probability and a currency amount should not share a column name.
    """
    by_id: dict[str, dict] = {}
    first = True
    for r in _sheet_dicts(ws):
        if first:
            _require_row_cols(
                r, ("table_id", axis, value_col), sheet=ws.title,
            )
            first = False
        tid = str(r["table_id"]).strip()
        try:
            k, v = int(r[axis]), float(r[value_col])
        except KeyError as exc:
            raise ValueError(
                f"sheet {ws.title!r} table {tid!r}: row is missing "
                f"column {exc.args[0]!r} (row has columns {sorted(r)})"
            ) from None
        bucket = by_id.setdefault(tid, {})
        if k in bucket:
            raise ValueError(
                f"sheet {ws.title!r} table {tid!r}: duplicate {axis}={k} "
                "(the later value silently overwrites the earlier); keep one "
                f"row per (table_id, {axis})"
            )
        bucket[k] = v
    return {tid: np.asarray([by_k[k] for k in range(len(by_k))], np.float64)
            for tid, by_k in by_id.items()}


# Recognised assumption-slot columns on the segments sheet. Every *other*
# column is a routing axis -- product, channel by convention, but a
# workbook may use any axes (just channel for a pricing run, or
# product x channel x risk_class). Detection mirrors the policies-frame
# attributes rule.
_SEGMENT_ASSUMPTION_COLS = frozenset({
    "mortality_table", "mortality_improvement_table", "lapse_table",
    "waiver_table", "surrender_value_table", "discount_table", "expense_table",
    "inflation_table", "settlement_table",
    "mortality_age_shift", "morbidity_age_shift", "waiver_age_shift",
    "ra_confidence", "mortality_cv", "morbidity_cv", "longevity_cv",
    "disability_cv", "expense_cv", "cost_of_capital_rate", "investment_return",
    "fund_fee", "ra_method", "state_model", "surrender_value_basis",
})


def read_basis(path: Path | str) -> "BasisRouter":
    """Read the basis workbook into a per-segment :class:`BasisRouter`.

    ``path`` is a single ``basis.xlsx`` workbook holding both the rate
    tables and the segment mapping (see the module header for the sheet
    layout). The ``segments`` sheet maps each (product, channel) to which
    tables it uses plus scalar parameters, with a ``_DEFAULTS`` row whose
    values blank cells inherit; the ``coverages`` sheet attaches
    rate-driven coverages to products.

    Returns a :class:`~fastcashflow.basis.BasisRouter` keyed by the segment
    axes -- ``(product, channel)`` by default, or whatever
    non-assumption columns the segments sheet declares (one axis, or three);
    ``.segment_axes`` records the axis names so :func:`~fastcashflow.gmm.measure`
    routes without a ``segment_by`` argument.

    v1: the discount and inflation tables are read but used flat (their
    first entry); the per-segment BasisRouter is returned for the caller to
    value segment by segment.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    def optional(sheet, reader):
        return reader(wb[sheet]) if sheet in wb.sheetnames else {}

    mortality_t = _flex_rate_table(_require_sheet(wb, "mortality_tables"))
    incidence_rate_t = optional("incidence_rate_tables", _flex_rate_table)
    waiver_t = optional("waiver_tables", _flex_rate_table)
    lapse_t = _flex_rate_table(_require_sheet(wb, "lapse_tables"))
    discount_t = _axis_tables(_require_sheet(wb, "discount_tables"), "year")
    inflation_t = optional(
        "inflation_tables", lambda w: _axis_tables(w, "year"),
    )
    if "ae_factors" in wb.sheetnames:
        ae_factors, ae_axes = _read_ae_factors(wb["ae_factors"])
    else:
        ae_factors, ae_axes = {}, ()
    improvement_t = optional(
        "improvement_tables",
        lambda w: _axis_tables(w, "year", value_col="factor"),
    )
    # Surrender value curves -- per-duration value, read from whichever
    # value column the sheet carries: ``amount`` (a currency amount, an
    # amount_per_* basis) or ``factor`` (a ratio on cumulative premium, the
    # cum_premium_factor basis). Optional; absent means lapse has no payout.
    # ``surrender_col_kind`` records which column was read so each segment's
    # ``surrender_value_basis`` can be checked against it (a factor read as
    # an amount, or vice versa, would silently mis-measure).
    surrender_col_kind = (
        _surrender_value_col(wb["surrender_value_tables"])
        if "surrender_value_tables" in wb.sheetnames else None
    )
    surrender_t = optional(
        "surrender_value_tables",
        lambda w: _axis_tables(w, "duration_month",
                               value_col=_surrender_value_col(w)),
    )
    # Expense ledger -- item form. Optional; per-segment ``expense_table``
    # in the segments sheet selects which table_id to attach.
    expense_t = optional("expense_tables", _read_expense_tables)
    # Claims run-off (settlement) patterns -- per-month weights summing to 1,
    # read as a month-indexed array (the ``Basis.settlement_pattern`` input).
    # Optional; per-segment ``settlement_table`` selects which table_id to
    # attach. Absent (or unattached) means every claim settles immediately.
    settlement_t = optional(
        "settlement_tables",
        lambda w: _axis_tables(w, "month", value_col="weight"),
    )

    defaults: dict = {}
    segments: list = []
    seg_rows = list(_sheet_dicts(_require_sheet(wb, "segments")))
    # Routing axes = every segments-sheet column that is not an assumption slot
    # and not a display label (``*_name``, report-only) -- order-independent, so
    # an axis column can sit anywhere among the assumption columns. The key tuple
    # reads them in column order; the default (no extra columns) is
    # (product, channel).
    axis_cols = tuple(
        c for c in (seg_rows[0].keys() if seg_rows else ())
        if c not in _SEGMENT_ASSUMPTION_COLS
        and c != "measurement_model"          # routing/accounting metadata, not an axis
        and not str(c).endswith("_name")
    )
    # An A/E axis that is not a segments-sheet column can never match a segment,
    # so the A/E would be silently discarded -- reject it up front.
    if ae_axes and seg_rows:
        seg_header = list(seg_rows[0].keys())
        unknown = [a for a in ae_axes if a not in seg_header]
        if unknown:
            raise ValueError(
                f"ae_factors sheet keys on axis column(s) {unknown} that are "
                "not in the segments sheet; the A/E would never match a segment "
                "and be silently discarded. Use segments-sheet axis columns "
                f"(have: {[c for c in seg_header if c in axis_cols]})."
            )
    if seg_rows:
        header = set(seg_rows[0].keys())
        for new, legacy in (("product", "product_code"),
                            ("channel", "channel_code")):
            if new not in header and legacy in header:
                raise ValueError(
                    f"segments sheet has column {legacy!r} but not {new!r} "
                    f"-- did you mean {new!r}? (the routing axes are now the "
                    "bare keys 'product' / 'channel', no '_code' suffix)"
                )
    for r in seg_rows:
        if str(r.get("product", "") or "").strip().lower() == "_defaults":
            defaults = r
        else:
            segments.append(r)
    # Coverages registry -- global, one row per coverage. The same code
    # plugs into any segment's contracts (a HEALTH policy and a TERM_LIFE
    # policy that both attach `CANCER` share the same incidence rate). When
    # a company genuinely needs product-specific calibrations of the same
    # disease, give them different coverage_codes (e.g. CANCER_HEALTH vs
    # CANCER_WHOLELIFE) -- the engine then treats them as separate coverages.
    #
    # Plan B (3-file split): this sheet carries only ``coverage`` +
    # ``rate_table`` -- the rate-driven entries. The pattern taxonomy
    # (DEATH / MORBIDITY / DIAGNOSIS / ANNUITY / MATURITY) moves to a
    # separate ``calculation_methods.csv`` file consumed by
    # :func:`read_model_points`, so the basis workbook is purely the
    # actuarial basis and the company catalogue lives elsewhere. Survival
    # entries (ANNUITY, MATURITY) never carry a ``rate_table`` and so do
    # not appear here. A death coverage's ``rate_table`` cell may point to
    # either an ``incidence_rate_tables`` entry or a ``mortality_tables``
    # entry: the engine reads the table as a per-coverage payment rate
    # independently of how the same id is also used for in-force decrement.
    rate_driven_coverages: list[tuple[str, str]] = []
    if "coverages" in wb.sheetnames:
        for r in _sheet_dicts(wb["coverages"]):
            rt = r.get("rate_table")
            code = str(r["coverage"]).strip()
            if rt in (None, ""):
                raise ValueError(
                    f"coverages row {code!r} has no rate_table; the "
                    "basis workbook only lists rate-driven coverages "
                    "(survival entries belong in calculation_methods.csv, "
                    "not here)"
                )
            rate_driven_coverages.append((code, str(rt).strip()))

    result = {}
    measurement_models = {}          # {seg_key: "GMM"|"PAA"|"VFA"} -- router metadata
    for seg in segments:
        product = str(seg.get("product", "") or "").strip()
        channel = str(seg.get("channel", "") or "").strip()
        seg_key = tuple(str(seg.get(c, "") or "").strip() for c in axis_cols)
        where = f"segments row {seg_key}"

        def cell(col):
            v = seg.get(col)
            if v is None or (isinstance(v, str) and not v.strip()):
                v = defaults.get(col)
            return None if (isinstance(v, str) and not v.strip()) else v

        def lookup(registry, col, optional_ref=False):
            tid = cell(col)
            tid = str(tid).strip() if tid is not None else None
            if tid is None:
                if optional_ref:
                    return None
                raise ValueError(f"{where}: {col!r} is required")
            if tid not in registry:
                raise ValueError(f"{where}: {col}={tid!r} is not registered")
            return registry[tid]

        def scalar(col, required=False):
            v = cell(col)
            if v is None and required:
                raise ValueError(f"{where}: {col!r} is required")
            return None if v is None else float(v)

        shift_mort = int(scalar("mortality_age_shift") or 0)
        shift_morb = int(scalar("morbidity_age_shift") or 0)
        shift_wvr = int(scalar("waiver_age_shift") or 0)

        def ae(coverage):
            ae_key = tuple(str(seg.get(a, "") or "").strip()
                           for a in ae_axes) + (coverage,)
            return ae_factors.get(ae_key)

        coverage_list = []
        for code, rate_table in rate_driven_coverages:
            # Death coverages share the mortality_tables namespace -- a
            # rate_table cell may name either an incidence_rate_tables or
            # a mortality_tables entry. Incidence wins on collision (the
            # convention rare in practice; calibrating the same code with
            # different tables under the same name is a workbook error).
            if rate_table in incidence_rate_t:
                rate_fn = incidence_rate_t[rate_table]
                shift = shift_morb
            elif rate_table in mortality_t:
                rate_fn = mortality_t[rate_table]
                shift = shift_mort
            else:
                raise ValueError(
                    f"coverage {code!r} of product {product!r}: "
                    f"rate_table {rate_table!r} is not registered. "
                    f"incidence_rate_tables has "
                    f"{_truncate_list(sorted(incidence_rate_t))}; "
                    f"mortality_tables has "
                    f"{_truncate_list(sorted(mortality_t))}"
                )
            rate_fn = _with_age_shift(rate_fn, shift)
            rate_fn = _with_ae_factor(rate_fn, ae(code))
            coverage_list.append(CoverageRate(code=code, rate=rate_fn))

        mortality_fn = lookup(mortality_t, "mortality_table")
        mortality_fn = _with_age_shift(mortality_fn, shift_mort)
        improvement_curve = lookup(
            improvement_t, "mortality_improvement_table", optional_ref=True,
        )
        mortality_fn = _with_improvement(mortality_fn, improvement_curve)

        waiver = lookup(waiver_t, "waiver_table", optional_ref=True)
        waiver_fn = _with_age_shift(waiver, shift_wvr)

        surrender_curve = lookup(
            surrender_t, "surrender_value_table", optional_ref=True,
        )
        # Row-form expense ledger -- the segments row points an
        # ``expense_table`` cell at one entry of the ``expense_tables``
        # sheet. Blank cell = no expense (empty ``expense_items`` tuple,
        # zero-expense projection).
        expense_items = lookup(expense_t, "expense_table", optional_ref=True)
        # Global economic inflation -- the segments row points an
        # ``inflation_table`` cell at one named scenario in the
        # ``inflation_tables`` sheet (analogous to ``discount_table``).
        # Blank cell = zero inflation (recurring expense items stay flat).
        inflation_curve = lookup(
            inflation_t, "inflation_table", optional_ref=True,
        )
        kwargs: dict = dict(
            mortality_annual=mortality_fn,
            lapse_annual=lookup(lapse_t, "lapse_table"),
            waiver_incidence_annual=waiver_fn,
            # Pass the full per-year discount through -- the engine
            # expands it to a per-month curve via fastcashflow.curves.
            # A one-row table reproduces the flat-scalar behaviour.
            discount_annual=lookup(discount_t, "discount_table"),
            expense_items=expense_items or (),
            expense_inflation=(
                inflation_curve if inflation_curve is not None else 0.0
            ),
            ra_confidence=scalar("ra_confidence", required=True),
            mortality_cv=scalar("mortality_cv", required=True),
            coverages=tuple(coverage_list),
            surrender_value_curve=surrender_curve,
            settlement_pattern=lookup(
                settlement_t, "settlement_table", optional_ref=True,
            ),
        )
        for opt_col in ("morbidity_cv", "longevity_cv", "disability_cv",
                        "expense_cv", "cost_of_capital_rate",
                        "investment_return", "fund_fee"):
            v = scalar(opt_col)
            if v is not None:
                kwargs[opt_col] = v
        method = cell("ra_method")
        if method is not None:
            kwargs["ra_method"] = str(method).strip()
        # Optional surrender_value_basis column -- how surrender_value_curve is
        # read: "cum_premium_factor" (default), "amount_per_policy", or
        # "amount_per_unit" (the latter needs a surrender_base_amount column on
        # the policies). Blank cell leaves the Basis default (cum_premium_factor).
        surr_basis = cell("surrender_value_basis")
        if surr_basis is not None:
            kwargs["surrender_value_basis"] = str(surr_basis).strip()
        # The surrender curve is read from one column kind (amount vs factor)
        # for the whole sheet, but each segment names how to interpret it via
        # surrender_value_basis. A mismatch -- an ``amount`` column used as a
        # cum_premium factor, or a ``factor`` column used as an amount --
        # silently mis-measures the surrender cash flow, so reject it here.
        if surrender_curve is not None and surrender_col_kind is not None:
            eff_basis = kwargs.get("surrender_value_basis", "cum_premium_factor")
            basis_wants_amount = eff_basis in ("amount_per_policy",
                                               "amount_per_unit")
            col_is_amount = surrender_col_kind == "amount"
            if basis_wants_amount != col_is_amount:
                raise ValueError(
                    f"{where}: surrender_value_basis={eff_basis!r} expects a "
                    f"{'amount' if basis_wants_amount else 'factor'} column "
                    f"but surrender_value_tables carries a "
                    f"{surrender_col_kind!r} column. Use a 'factor' column "
                    "for cum_premium_factor, or an 'amount' column for "
                    "amount_per_policy / amount_per_unit."
                )
        # Optional state_model column -- non-programmer actuary picks a
        # bundled topology by its registry key (e.g. "WAIVER"). Blank cell
        # leaves Basis.state_model = None; an unknown key is an
        # error with a hint listing the registered keys.
        state_model_key = cell("state_model")
        if state_model_key is not None:
            key = str(state_model_key).strip()
            try:
                kwargs["state_model"] = STATE_MODELS[key]
            except KeyError:
                raise ValueError(
                    f"{where}: state_model={key!r} is not in STATE_MODELS "
                    f"(known: {sorted(STATE_MODELS)})"
                ) from None
        if seg_key in result:
            raise ValueError(
                f"{where}: duplicate segment {seg_key} -- two rows route to "
                "the same Basis (the later silently overwrites the earlier). "
                "Remove the duplicate segment row."
            )
        result[seg_key] = Basis(**kwargs)
        mm = cell("measurement_model")          # router metadata, not a Basis kwarg
        if mm is not None:
            measurement_models[seg_key] = str(mm).strip().upper()
    return BasisRouter(result, segment_axes=axis_cols or ("product", "channel"),
                       measurement_models=measurement_models)


# ---------------------------------------------------------------------------
# Model points -- built from policies + coverages frames
# ---------------------------------------------------------------------------

def _read_state(col: pl.Series) -> IntArray:
    """Convert a model-point ``state`` column to engine state codes.

    Accepts the readable names a practitioner edits in a spreadsheet --
    ``active`` / ``waiver`` / ``paidup`` -- or the integer codes directly.
    Case, spaces, hyphens and underscores are ignored, so ``Paid-up`` and
    ``paid up`` read the same. A blank cell means an ordinary active contract.
    """
    if col.dtype == pl.String:
        # Normalised lookup -- canonical STATE_NAMES keys ("ACTIVE", "WAIVER",
        # "PAIDUP") are uppercase, but any spelling (case, spaces, hyphens,
        # underscores ignored) of the canonical name maps to the same code.
        normalised = {
            k.lower().replace("_", "").replace("-", "").replace(" ", ""): v
            for k, v in STATE_NAMES.items()
        }
        out = np.empty(len(col), dtype=np.int64)
        for i, v in enumerate(col):
            name = "" if v is None else str(v).strip().lower()
            name = name.replace(" ", "").replace("-", "").replace("_", "")
            if name == "":
                out[i] = STATE_ACTIVE
            elif name in normalised:
                out[i] = normalised[name]
            else:
                raise ValueError(
                    f"unknown contract state {v!r}; "
                    f"expected one of {sorted(STATE_NAMES)}"
                )
        return out
    out = col.fill_null(STATE_ACTIVE).to_numpy().astype(np.int64)
    valid = set(STATE_NAMES.values())
    bad = sorted(set(int(v) for v in out) - valid)
    if bad:
        raise ValueError(
            f"state column has unknown integer value(s) {bad}; "
            f"expected one of {sorted(valid)} (see STATE_NAMES)"
        )
    return out


def _warn_if_elapsed_months(columns) -> None:
    """Warn that ``elapsed_months`` on a policies frame is silently dropped.

    The static-spec policies frame holds inception-time facts (issue_age,
    term, sex ...). The in-force closing state lives in a separate
    ``inforce_state`` file -- :func:`read_inforce_state` is the only
    surface that fills the ``elapsed_months`` field of :class:`ModelPoints`.
    A column on the policies side is a common mistake (mixed roles) and
    would be silently ignored; the warning makes the source-of-truth
    boundary explicit.
    """
    if "elapsed_months" in columns:
        warnings.warn(
            "policies frame has 'elapsed_months' column; this reader "
            "ignores it. elapsed_months belongs in the in-force state file "
            "(see read_inforce_state / apply_inforce_state) -- the policies "
            "frame is the inception-time static spec.",
            UserWarning,
            stacklevel=3,
        )


def _parse_calculation_methods(path: Path | str) -> dict[str, CalculationMethod]:
    """Read a ``calculation_methods.csv`` taxonomy file into a dict.

    The file has two required columns -- ``coverage`` and
    ``calculation_method``. Any other column (e.g. a human-friendly label)
    is ignored, since the engine routes by the bare ``coverage`` key.
    Returns ``{coverage: CalculationMethod}``. Raises
    :class:`ValueError` for an unknown pattern (V1) and a duplicate code
    (V2); the messages name the offending row so the operator can fix
    the file without scrolling through it.
    """
    df = _read_frame(path)
    if "coverage" not in df.columns and "coverage_code" in df.columns:
        raise ValueError(
            "the calculation_methods file has column 'coverage_code' but not "
            "'coverage' -- the coverage key is now the bare 'coverage' "
            "(no '_code' suffix)"
        )
    for need in ("coverage", "calculation_method"):
        if need not in df.columns:
            raise ValueError(
                f"the calculation_methods file is missing required column "
                f"{need!r}"
            )
    result: dict[str, CalculationMethod] = {}
    valid = ", ".join(p.value for p in CalculationMethod)
    for row in df.iter_rows(named=True):
        code = str(row["coverage"]).strip()
        raw = str(row["calculation_method"]).strip()
        try:
            pattern = CalculationMethod(raw)
        except ValueError as exc:
            raise ValueError(
                f"calculation_methods row {code!r}: calculation_method={raw!r} "
                f"is not one of {{{valid}}}"
            ) from exc
        if code in result:
            raise ValueError(
                f"calculation_methods row {code!r}: duplicate coverage "
                "(every code may appear exactly once in the taxonomy)"
            )
        result[code] = pattern
    return result


# Policies-frame columns the engine recognises as fields (or the mp_id join
# key). Every *other* column on the policies frame is a grouping attribute --
# portfolio_id, profitability_group, risk_class, region, campaign_id, ... --
# read into ModelPoints.attributes for group()/group_of_contracts.
_POLICY_RESERVED_COLS = frozenset({
    "mp_id",
    "issue_age", "term_months", "premium",
    "sex", "count", "state", "issue_class", "elapsed_months", "issue_date",
    "premium_term_months", "premium_frequency_months",
    "annuity_frequency_months", "disability_income", "disability_benefit",
    "account_value", "minimum_crediting_rate", "minimum_death_benefit",
    "minimum_accumulation_benefit", "surrender_base_amount",
    "contract_boundary_months",
    "product", "channel",
})

# The columns ``read_vfa_model_points`` recognises (its allow-list plus the
# required fields). A column outside this set is ignored, so a typo is dropped.
_VFA_POLICY_COLS = frozenset({
    "mp_id", "issue_age", "term_months", "premium", "state",
    "sex", "count", "premium_term_months", "premium_frequency_months",
    "annuity_frequency_months", "maturity_benefit", "annuity_payment",
    "disability_income", "disability_benefit", "account_value",
    "minimum_crediting_rate", "minimum_death_benefit",
    "minimum_accumulation_benefit", "surrender_base_amount",
    "contract_boundary_months", "product", "channel",
})


def _within_edit_distance_1(a: str, b: str) -> bool:
    """True if ``a`` and ``b`` differ by at most one insert / delete / substitute
    (a Levenshtein distance <= 1) -- a cheap typo detector."""
    if a == b:
        return True
    la, lb = len(a), len(b)
    if abs(la - lb) > 1:
        return False
    if la == lb:                          # one substitution
        return sum(x != y for x, y in zip(a, b)) == 1
    if la > lb:                           # make ``a`` the shorter
        a, b = b, a
    i = j = 0
    skipped = False
    while i < len(a) and j < len(b):      # one insertion in the longer
        if a[i] == b[j]:
            i += 1
            j += 1
        elif skipped:
            return False
        else:
            skipped = True
            j += 1
    return True


def _warn_near_reserved_columns(columns, reserved, *, context: str) -> None:
    """Warn for a column that looks like a typo of a recognised engine field --
    a case-only difference or one edit away. Such a column is silently dropped
    (read as a grouping attribute, or ignored), so ``coun`` -> ``count`` would
    take its default (1) with no error. A warning, not an error, so a genuine
    attribute that happens to sit near a reserved name still reads."""
    for col in columns:
        c = str(col)
        if c in reserved or c.startswith("_"):
            continue
        cl = c.lower()
        for r in reserved:
            if _within_edit_distance_1(cl, r.lower()):
                warnings.warn(
                    f"{context}: column {c!r} looks like a typo of the field "
                    f"{r!r}; it is read as a grouping attribute, not as {r!r} "
                    f"(so {r!r} takes its default). Rename it if you meant {r!r}.",
                    UserWarning, stacklevel=3,
                )
                break


def _model_points_from_frames(pol: pl.DataFrame, cov: pl.DataFrame,
                       calculation_methods=None) -> ModelPoints:
    """Build a ``ModelPoints`` from a policies + coverages pair.

    The rate-driven coverage order is taken from the ``calculation_methods``
    catalogue, so the portfolio is read without the actuarial basis. The
    engine aligns its coverages to that order at measure time.
    """
    if calculation_methods is None:
        raise ValueError(
            "model points need the calculation_methods taxonomy -- "
            "the per-code pattern routes survival rows (ANNUITY / MATURITY) "
            "to scalar fields and rate-driven rows to the coverage CSR. "
            "Pass a calculation_methods.csv path to read_model_points."
        )
    for need in ("mp_id", "issue_age", "term_months"):
        if need not in pol.columns:
            raise ValueError(
                f"the policies frame is missing required column {need!r}"
            )
    # A leftover *_code column from the pre-rename schema would otherwise be
    # absorbed as a grouping attribute, leaving product / channel empty and
    # silently mis-routing under a segmented basis. Fail with the same hint
    # read_basis gives.
    for new, legacy in (("product", "product_code"), ("channel", "channel_code")):
        if new not in pol.columns and legacy in pol.columns:
            raise ValueError(
                f"the policies frame has column {legacy!r} but not {new!r} "
                f"-- the routing axes are now the bare keys 'product' / "
                f"'channel' (no '_code' suffix)"
            )
    if "coverage" not in cov.columns and "coverage_code" in cov.columns:
        raise ValueError(
            "the coverages frame has column 'coverage_code' but not 'coverage' "
            "-- the coverage key is now the bare 'coverage' (no '_code' suffix)"
        )
    for need in ("mp_id", "coverage", "amount"):
        if need not in cov.columns:
            raise ValueError(
                f"the coverages frame is missing required column {need!r}"
            )
    # An empty coverages frame otherwise fails cryptically at the policies
    # join (the all-null mp_id column infers a string dtype that does not
    # match the integer policies key). Every model point needs at least one
    # coverage row -- the benefit amounts and per-coverage premium live there.
    if cov.height == 0:
        raise ValueError(
            "the coverages frame is empty (0 rows); every model point needs "
            "at least one coverage row"
        )
    _warn_if_elapsed_months(pol.columns)
    n_mp = pol.height
    if n_mp == 0:
        raise ValueError(
            "the policies frame is empty (0 rows); there is nothing to measure"
        )
    # mp_id uniqueness -- a duplicate id would fan out the coverages join
    # (one-to-many) and silently inflate per-policy benefits.
    if pol["mp_id"].n_unique() != n_mp:
        dups = (pol.group_by("mp_id").len()
                   .filter(pl.col("len") > 1)["mp_id"].to_list())
        raise ValueError(
            f"policies frame has duplicate mp_id value(s) {dups[:10]}"
            f"{' (...)' if len(dups) > 10 else ''} -- mp_id must be unique"
        )
    # Premium double-source -- if both the coverages frame's ``premium``
    # column and the policies frame's ``premium`` are present, the
    # cov-side branch silently wins below. Reject up front so the operator
    # picks one source.
    if "premium" in cov.columns and "premium" in pol.columns:
        raise ValueError(
            "premium is specified twice -- 'premium' in the coverages "
            "frame and 'premium' in the policies frame. Pick one: "
            "the coverages-side column sums per coverage to the policy, "
            "the policies-side column is a flat per-policy amount."
        )
    # Coverage-rule columns -- waiting / reduction_end / reduction_factor
    # must arrive together. A reduction_factor without a reduction_end is
    # silently inert (the factor applies for ``t < 0`` months, i.e. never)
    # and almost certainly a user oversight.
    has_rend = "reduction_end" in cov.columns
    has_rfac = "reduction_factor" in cov.columns
    if has_rfac and not has_rend:
        raise ValueError(
            "coverages frame has 'reduction_factor' without 'reduction_end' "
            "-- the factor would never fire (reduction_end defaults to 0). "
            "Add a reduction_end column (months) or drop the factor column."
        )
    ctypes = {k: CalculationMethod(v) for k, v in calculation_methods.items()}
    # Rate-driven coverage order comes from the *catalogue* (calculation_methods),
    # not the basis -- so reading the portfolio needs no basis.
    # coverage_index integers index this order; the engine aligns
    # Basis.coverages to it at measure time (coverage.align_coverages).
    # Only the rate-driven codes that actually appear in this portfolio are
    # kept, in catalogue order.
    present_codes = set(cov["coverage"].to_list())
    rate_driven_codes = [c for c, m in ctypes.items()
                         if m in RATE_DRIVEN_METHODS and c in present_codes]
    code_to_cov_idx = {c: i for i, c in enumerate(rate_driven_codes)}

    # Resolve every coverage row to its policy index and coverage type.
    pol = pol.with_row_index("_mp")
    cmap = pl.DataFrame({
        "coverage": list(ctypes.keys()),
        "_type": [str(v) for v in ctypes.values()],
        "_cov_idx": [code_to_cov_idx.get(c, -1) for c in ctypes],
    })
    cov = (cov.join(pol.select("mp_id", "_mp"), on="mp_id", how="left")
              .join(cmap, on="coverage", how="left"))
    if cov["_mp"].null_count():
        bad = sorted({v for v in cov.filter(pl.col("_mp").is_null())
                                    ["mp_id"].to_list() if v is not None})
        raise ValueError(
            f"coverages frame references {len(bad)} unknown mp_id "
            f"value(s) not present in the policies frame: "
            f"{_truncate_list(bad)}"
        )
    if cov["_type"].null_count():
        bad = sorted({v for v in cov.filter(pl.col("_type").is_null())
                                    ["coverage"].to_list() if v is not None})
        raise ValueError(
            f"coverages frame references {len(bad)} coverage "
            f"value(s) not in the calculation_methods taxonomy: "
            f"{_truncate_list(bad)}"
        )

    mp = cov["_mp"].to_numpy()
    ctype = cov["_type"].to_numpy()
    cov_idx = cov["_cov_idx"].to_numpy().astype(np.int64)
    amount = cov["amount"].to_numpy().astype(np.float64)

    fields: dict[str, object] = dict(
        issue_age=pol["issue_age"].to_numpy(),
        term_months=pol["term_months"].to_numpy(),
    )
    for opt in ("sex", "count", "premium_term_months",
                "premium_frequency_months", "annuity_frequency_months",
                "disability_income", "disability_benefit", "issue_class",
                "surrender_base_amount", "contract_boundary_months"):
        if opt in pol.columns:
            fields[opt] = pol[opt].to_numpy()
    for opt in ("product", "channel"):
        if opt in pol.columns:
            fields[opt] = pol[opt].to_numpy()
    if "issue_date" in pol.columns:
        fields["issue_date"] = pol["issue_date"].to_numpy()
    if "state" in pol.columns:
        fields["state"] = _read_state(pol["state"])
    # Any policies column that is not a recognised engine field is a grouping
    # attribute (portfolio_id, profitability_group, risk_class, region, ...) --
    # one value per policy = one per model point. Available to group() /
    # group_of_contracts via ModelPoints.axis.
    attributes = {c: pol[c].to_numpy()
                  for c in pol.columns
                  if c not in _POLICY_RESERVED_COLS and not str(c).startswith("_")}
    # A column one edit away from a reserved field is almost certainly a typo
    # (``coun`` -> count, which would otherwise default to 1 = a 1000x error).
    _warn_near_reserved_columns(attributes, _POLICY_RESERVED_COLS,
                                context="read_model_points")
    if attributes:
        fields["attributes"] = attributes
    # Carry mp_id (the contract identity) as a dedicated field so
    # apply_inforce_state can join the period-close state on it instead of
    # trusting row order. It is a label, never read by the kernel.
    fields["mp_id"] = pol["mp_id"].to_numpy()

    def _by_policy(mask) -> np.ndarray:
        return np.bincount(mp[mask], weights=amount[mask], minlength=n_mp)

    fields["maturity_benefit"] = _by_policy(ctype == CalculationMethod.MATURITY)
    fields["annuity_payment"] = _by_policy(ctype == CalculationMethod.ANNUITY)

    # Premium -- the coverages frame carries it per coverage; sum to the policy.
    if "premium" in cov.columns:
        prem = cov["premium"].fill_null(0.0).to_numpy().astype(np.float64)
        fields["premium"] = np.bincount(mp, weights=prem, minlength=n_mp)
    elif "premium" in pol.columns:
        fields["premium"] = pol["premium"].to_numpy()
    else:
        # Neither source provided -- premium is silently zero. A genuine
        # paid-up portfolio is one valid case; a forgotten column is the
        # other. Warn so the latter doesn't slip through.
        warnings.warn(
            "model points have no premium source -- neither "
            "'premium' on the coverages frame nor 'premium' on the "
            "policies frame was found. premium defaults to zero; "
            "if this portfolio is not fully paid-up, add the column.",
            UserWarning,
            stacklevel=3,
        )
        fields["premium"] = np.zeros(n_mp)

    # Coverage list: the rate-driven coverages (codes 0..n-1 indexing
    # ``coverage_codes`` below). annuity / maturity are survival scalars and
    # not part of the CSR. Every rate-driven present code is in
    # ``rate_driven_codes`` by construction, so ``cov_idx >= 0`` here; a code
    # absent from the catalogue was already rejected (the ``_type`` null
    # check above). Whether the basis register a rate for each code is
    # checked at measure time (coverage.align_coverages, the V4 guard).
    is_cov = np.isin(ctype, RATE_DRIVEN_METHODS)
    order = np.argsort(mp[is_cov], kind="stable")
    cov_mp = mp[is_cov][order]
    fields["coverage_index"] = cov_idx[is_cov][order]
    fields["coverage_amount"] = amount[is_cov][order]

    # Optional per-coverage benefit rules -- a waiting period and a
    # reduced-benefit period, each CSR-aligned with coverage_index.
    for col, field, default in (("waiting", "coverage_waiting", 0),
                                ("reduction_end", "coverage_reduction_end", 0),
                                ("reduction_factor", "coverage_reduction_factor", 1.0),
                                ("step_month", "coverage_step_month", 0),
                                ("step_factor", "coverage_step_factor", 1.0),
                                ("escalation_annual", "coverage_escalation_annual", 0.0),
                                ("escalation_cap", "coverage_escalation_cap", 0.0)):
        if col in cov.columns:
            rule = cov[col].fill_null(default).to_numpy()
            fields[field] = rule[is_cov][order]

    fields["coverage_offset"] = np.concatenate((
        np.zeros(1, np.int64),
        np.cumsum(np.bincount(cov_mp, minlength=n_mp), dtype=np.int64),
    ))
    fields["calculation_methods"] = ctypes
    # The catalogue order the coverage_index integers were built against.
    # The engine aligns Basis.coverages to this at measure time.
    fields["coverage_codes"] = tuple(rate_driven_codes)
    return ModelPoints(**fields)


def read_model_points(
    path: Path | str,
    coverages: Path | str | None = None,
    calculation_methods: Path | str | dict[str, CalculationMethod] | None = None,
) -> ModelPoints:
    """Read model points from a parquet, CSV, Excel or feather file.

    Reads the portfolio **without any basis** -- the model points and
    the actuarial basis are separate inputs. The basis enters only at the
    engine call (``measure``), which aligns its coverages to the
    portfolio's coverage order.

    The portfolio is two frames -- a policies frame plus a coverages frame:

    * a policies frame (``mp_id``, ``issue_age``, ``term_months``, optional
      ``sex`` / ``count`` / ``state`` / ``issue_class`` / ``issue_date`` /
      ``premium`` / ``premium_term_months`` /
      ``premium_frequency_months`` / ``annuity_frequency_months`` /
      ``contract_boundary_months`` / ``product`` / ``channel``), one row
      per policy. Any *other*
      column is read as a grouping attribute (``portfolio_id``,
      ``profitability_group``, ``risk_class``, ``region``, ...) into
      :attr:`ModelPoints.attributes`, for :func:`~fastcashflow.group` /
      :func:`~fastcashflow.group_of_contracts`;
    * a coverages frame (``mp_id``, ``coverage``, ``amount``, optional
      ``premium`` / ``waiting`` / ``reduction_end`` / ``reduction_factor`` and
      the benefit step-up / escalation columns ``step_month`` / ``step_factor``
      (a benefit step at a duration) / ``escalation_annual`` / ``escalation_cap``
      (annual compounding growth, capped -- the escalating-benefits
      recipe in the cookbook)), one row per policy x coverage -- so per-coverage
      rules (waiting, reduction and escalation) ride along, which a flat
      one-row-per-policy file cannot carry.

    Pass them as ``read_model_points(policies, coverages=coverages_path,
    calculation_methods=...)``, or as a single ``.xlsx`` carrying ``policies``
    and ``coverages`` sheets. The normalised two-frame shape mirrors a policy
    table joined to a coverage table -- the form data arrives in from a policy
    system. ``calculation_methods`` is the company taxonomy file (CSV / parquet
    / feather / xlsx) -- the third side of the split between *portfolio*
    (policies + coverages), *basis* (basis.xlsx) and *catalogue*
    (calculation_methods.csv).

    The policies frame is the **inception-time static spec** -- issue_age,
    term, sex, and so on. The in-force closing state (elapsed_months,
    prior_csm, lock_in_rate) belongs in a separate file read by
    :func:`read_inforce_state`. An ``elapsed_months`` column on the
    policies side is ignored and a :class:`UserWarning` is emitted; do
    not encode the as-of date by mixing it into the static spec.
    """
    if isinstance(calculation_methods, (str, Path)):
        methods_dict = _parse_calculation_methods(calculation_methods)
    else:
        methods_dict = calculation_methods
    p = str(path)
    if coverages is None and p.endswith(".xlsx"):
        wb = openpyxl.load_workbook(p, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        if "policies" in sheets and "coverages" in sheets:
            return _model_points_from_frames(
                pl.read_excel(p, sheet_name="policies", engine="openpyxl"),
                pl.read_excel(p, sheet_name="coverages", engine="openpyxl"),
                methods_dict,
            )
    pol = _read_frame(path)
    if coverages is None:
        raise ValueError(
            f"{p!r} was read without a coverages frame. read_model_points "
            "needs a coverages frame: pass coverages=<path> (an mp_id / coverage / "
            "amount frame), or a single .xlsx carrying 'policies' and "
            "'coverages' sheets. A flat one-row-per-policy (wide) file cannot "
            "carry per-coverage waiting / reduction rules and is not accepted."
        )
    return _model_points_from_frames(
        pol, _read_frame(coverages), methods_dict,
    )


def read_vfa_model_points(
    path: Path | str,
    *,
    calculation_methods: "Path | str | dict[str, CalculationMethod] | None" = None,
) -> ModelPoints:
    """Read the account-value base of variable (VFA) contracts from a policies file.

    This reads the part measured under the VFA model: the account value and its
    guarantee floors (GMDB / GMAB), all named policy columns (``account_value``,
    ``minimum_death_benefit``, ``minimum_accumulation_benefit``,
    ``minimum_crediting_rate``). That base carries no coverage-code coverages,
    so it is a single policies frame. ``issue_age`` and ``term_months`` are
    required; the named policy / account columns are read if present.

    Protection riders attached to a variable product (death / cancer /
    hospitalisation rider) are separate coverages, read and measured on their own
    -- a policies + coverages book through :func:`read_model_points` (GMM). So a
    ``<coverage>_benefit`` column is rejected here: a coverage encoded as a
    column is the lossy wide form, and coverages belong in their own frame which
    can hold the per-coverage waiting / reduction rules a flat column cannot.
    """
    return _vfa_model_points_from_frame(_read_frame(path), calculation_methods)


def _vfa_model_points_from_frame(df, calculation_methods) -> ModelPoints:
    """Build a VFA :class:`ModelPoints` from a single policies frame.

    The frame-level body of :func:`read_vfa_model_points`, factored out so the
    out-of-core ``vfa.measure_stream`` can build model points from each parquet
    chunk slice without re-reading the whole file.
    """
    named_benefit = {"maturity_benefit", "disability_benefit",
                     "minimum_death_benefit", "minimum_accumulation_benefit"}
    coverage_cols = sorted(c for c in df.columns
                           if c.endswith("_benefit") and c not in named_benefit)
    if coverage_cols:
        raise ValueError(
            f"read_vfa_model_points got coverage benefit column(s) "
            f"{coverage_cols} -- account-value (VFA) contracts carry no "
            "coverage-code coverages. For a product with coverages, use "
            "read_model_points with a coverages frame."
        )
    for need in ("issue_age", "term_months"):
        if need not in df.columns:
            raise ValueError(
                f"the VFA policies file is missing required column {need!r}"
            )
    _warn_if_elapsed_months(df.columns)
    n_mp = df.height
    fields: dict[str, object] = dict(
        issue_age=df["issue_age"].to_numpy(),
        term_months=df["term_months"].to_numpy(),
        premium=(df["premium"].to_numpy()
                       if "premium" in df.columns else np.zeros(n_mp)),
    )
    for opt in ("sex", "count", "premium_term_months",
                "premium_frequency_months", "annuity_frequency_months",
                "maturity_benefit", "annuity_payment", "disability_income",
                "disability_benefit", "account_value", "minimum_crediting_rate",
                "minimum_death_benefit", "minimum_accumulation_benefit",
                "surrender_base_amount", "contract_boundary_months",
                "product", "channel", "mp_id"):
        if opt in df.columns:
            if opt == "minimum_crediting_rate":
                # A blank crediting-rate cell means no crediting guarantee (the
                # no-guarantee sentinel), not a 0% floor; an explicit 0.0
                # survives as a real 0% floor. Cast first so an all-blank column
                # (inferred as Null dtype) still fills.
                fields[opt] = (df[opt].cast(pl.Float64)
                               .fill_null(NO_GUARANTEE_RATE).to_numpy())
            else:
                fields[opt] = df[opt].to_numpy()
    if "state" in df.columns:
        fields["state"] = _read_state(df["state"])
    # The VFA reader has no grouping-attribute catch-all -- an unrecognised
    # column is dropped entirely, so a typo'd ``account_valu`` would leave
    # account_value 0 and the guarantee floor becomes the whole payout. Warn
    # on a near-match the same way the GMM reader does.
    _warn_near_reserved_columns(df.columns, _VFA_POLICY_COLS,
                                context="read_vfa_model_points")
    if isinstance(calculation_methods, (str, Path)):
        fields["calculation_methods"] = _parse_calculation_methods(calculation_methods)
    elif calculation_methods is not None:
        fields["calculation_methods"] = calculation_methods
    return ModelPoints(**fields)


def read_inforce_policies(
    path: Path | str,
    coverages: Path | str | None = None,
    calculation_methods: "Path | str | dict[str, CalculationMethod] | None" = None,
) -> "tuple[ModelPoints, InforceState]":
    """Read a single combined policies + in-force state file.

    A self-contained snapshot at a settlement date -- one file per
    valuation date, one row per surviving contract. Columns combine the
    permanent contract spec (``issue_age``, ``sex``, ``term_months``,
    premiums, benefits, ...) and the closing state from the prior period
    (``elapsed_months``, ``count``, ``prior_csm``, ``lock_in_rate``). This
    matches the Korean industry period-close file pattern -- one
    self-contained snapshot per period, no separate state file to keep
    in sync.

    Returns a ``(ModelPoints, InforceState)`` tuple. The ``ModelPoints``
    has the state's ``elapsed_months`` and ``count`` already folded in;
    the ``InforceState`` carries ``prior_csm`` and ``lock_in_rate`` for
    the period-close settlement call::

        mp, state = fcf.read_inforce_policies(
            "inforce_2026Q1.csv",
            coverages="coverages.csv",
            calculation_methods="calculation_methods.csv",
        )
        movement = fcf.gmm.settle(mp, state, basis, period_months=3)

    (For a diagnostics / run-off view without the movement lines, the same
    inputs feed :func:`fastcashflow.gmm.measure_inforce`.)

    For the two-file equivalent (separate ``policies.csv`` +
    ``inforce_state.csv``), see :func:`read_model_points` +
    :func:`read_inforce_state` + :func:`apply_inforce_state`. Both
    workflows produce valuation-ready inputs that give the same valuation:
    ``settle`` and ``measure_inforce`` re-align the state by mp_id
    internally, so the answer does not depend on the state file's row order. The returned
    ``InforceState`` is itself row-aligned to the model points here; in the
    two-file path the state object keeps its file order (only the model points
    are reordered by :func:`apply_inforce_state`), so call
    :func:`align_inforce_state` before slicing or reading ``state.prior_csm``
    directly. Pick the form that fits the company's extract pipeline.

    Required columns: ``mp_id``, ``elapsed_months``, ``count``,
    ``prior_csm``, ``lock_in_rate``, plus whatever the spec side of
    :func:`read_model_points` needs (``issue_age``, ``term_months``,
    optional ``sex``, premiums, ``<code>_benefit`` columns for wide form).
    Optional settlement columns: ``prior_count`` (the in-force count at the
    opening date -- both :func:`fastcashflow.gmm.settle` and
    :func:`fastcashflow.vfa.settle` need it) and ``prior_loss_component``
    (the prior period's closing loss component). Optional VFA state
    columns: ``account_value`` (the *observed* fund value at the valuation
    date -- it rides on the returned ``InforceState``; the snapshot has no
    separate inception fund column) and ``prior_account_value`` (the prior
    reporting date's observed fund value, which ``vfa.settle`` needs).
    Variance / movement analysis (:func:`roll_forward`,
    :func:`reconcile`) is unaffected -- mp_id-based matching across
    periods works the same regardless of which reader built each
    snapshot.
    """
    from fastcashflow.model_points import (
        InforceState, ModelPoints, apply_inforce_state,
    )

    df = _read_frame(path)
    needed = ("mp_id", "elapsed_months", "count", "prior_csm", "lock_in_rate")
    for col in needed:
        if col not in df.columns:
            raise ValueError(
                f"the in-force policies file is missing required column "
                f"{col!r}. The combined file carries the policies spec "
                "plus the closing-state columns "
                "(elapsed_months, count, prior_csm, lock_in_rate)."
            )
    # A uniform column collapses to the scalar locked-in rate; a cohort-aware
    # column (issue cohorts / GoCs with different inception rates, paragraph B72(b))
    # is carried per row -- gmm.settle partitions by rate.
    lock = df["lock_in_rate"].to_numpy().astype(np.float64)
    lock_in_rate = (float(lock[0]) if lock.size and np.all(lock == lock[0])
                    else lock)
    state = InforceState(
        mp_id=df["mp_id"].to_numpy(),
        elapsed_months=df["elapsed_months"].to_numpy().astype(np.int64),
        count=df["count"].to_numpy().astype(np.float64),
        prior_csm=df["prior_csm"].to_numpy().astype(np.float64),
        lock_in_rate=lock_in_rate,
        **_optional_state_columns(df),
    )

    # Drop the state-only columns before handing the frame to the
    # standard policies reader, which would otherwise warn about
    # ``elapsed_months`` on a policies frame and ignore the rest. ``count``
    # stays -- it is a valid policies column too, and ``apply_inforce_state``
    # will overwrite it with the state value below anyway. ``account_value``
    # also stays (a valid VFA policies column -- the inception fund value);
    # the state's observed fund value rides on the InforceState above.
    state_only = [c for c in _INFORCE_STATE_ONLY if c in df.columns]
    spec_df = df.drop(*state_only)

    if isinstance(calculation_methods, (str, Path)):
        methods_dict = _parse_calculation_methods(calculation_methods)
    else:
        methods_dict = calculation_methods
    if coverages is None:
        raise ValueError(
            "read_inforce_policies needs a coverages frame: pass coverages=<path> "
            "(an mp_id / coverage / amount frame). A flat one-row-per-policy "
            "(wide) file cannot carry per-coverage waiting / reduction rules and "
            "is not accepted."
        )
    mp = _model_points_from_frames(
        spec_df, _read_frame(coverages), methods_dict,
    )
    mp = apply_inforce_state(mp, state)
    return mp, state


def sample_data_dir() -> Path:
    """Return the on-disk path of the bundled sample data directory.

    The directory contains ``sample_basis.xlsx``, ``sample_policies.csv``
    and ``sample_coverages.csv`` -- the inputs behind
    :func:`load_sample_basis` and :func:`load_sample_model_points`.
    Use this to open the workbook in Excel and see what a complete
    fastcashflow input looks like before preparing your own.
    """
    return Path(str(resources.files("fastcashflow") / "sample_data"))


def load_sample_basis() -> "BasisRouter":
    """Read fastcashflow's bundled sample basis workbook.

    A filled-in workbook packaged with the library, the companion to
    :func:`load_sample_model_points`. See :func:`read_basis` for the
    workbook format. The bundled sample is a :class:`BasisRouter` over seven
    ``(product, channel)`` segments -- three products (``TERM_LIFE_A``,
    ``HEALTH_A``, ``WHOLE_LIFE_A``) across the ``FC`` / ``GA`` / ``TM``
    channels; resolve one segment (``router.resolve(segment)``) to use it as
    a single ``Basis``.
    """
    source = resources.files("fastcashflow") / "sample_data" / "sample_basis.xlsx"
    with resources.as_file(source) as path:
        return read_basis(path)


def load_sample_calculation_methods() -> dict[str, CalculationMethod]:
    """Read fastcashflow's bundled sample calculation-method taxonomy.

    The companion to :func:`load_sample_basis` and
    :func:`load_sample_model_points` -- the company-level catalogue that
    maps each ``coverage`` to its :class:`CalculationMethod`. The same
    file format every portfolio uses (see :func:`read_model_points`
    ``calculation_methods`` argument).
    """
    source = (
        resources.files("fastcashflow") / "sample_data"
        / "sample_calculation_methods.csv"
    )
    with resources.as_file(source) as path:
        return _parse_calculation_methods(path)


def load_sample_model_points() -> ModelPoints:
    """Read fastcashflow's bundled sample portfolio.

    A small portfolio -- a policies file, a coverages file and
    the calculation-method taxonomy -- packaged with the library, so the
    engine can be tried without preparing an input file. See
    :func:`read_model_points` for the file format. The coverage order
    comes from the ``calculation_methods`` catalogue; no basis are
    needed to read the portfolio.
    """
    patterns = load_sample_calculation_methods()
    base = resources.files("fastcashflow") / "sample_data"
    with resources.as_file(base / "sample_policies.csv") as policies, \
            resources.as_file(base / "sample_coverages.csv") as coverages:
        return read_model_points(
            policies, coverages=coverages, calculation_methods=patterns,
        )


def load_sample_inforce_state() -> "InforceState":
    """Read fastcashflow's bundled sample in-force state.

    Aligned row-for-row with :func:`load_sample_model_points`. Pair with
    :func:`apply_inforce_state` to fold ``elapsed_months`` and ``count``
    into the sample model points, then settle the period with
    :func:`fastcashflow.gmm.settle` (it reads ``prior_csm`` /
    ``prior_count`` / ``lock_in_rate``) or value the book diagnostically
    with :func:`fastcashflow.gmm.measure_inforce`.
    """
    base = resources.files("fastcashflow") / "sample_data"
    with resources.as_file(base / "sample_inforce_state.csv") as path:
        return read_inforce_state(path)


def load_sample_vfa_basis() -> Basis:
    """Bundled VFA (variable, account-value) basis -- a single basis.

    Built from the sample ``TERM_LIFE_A`` / ``FC`` basis (same mortality,
    lapse and discount) with the protection coverages dropped and the two
    VFA economic inputs set: ``investment_return`` (the underlying-items
    return the account value grows at) and ``fund_fee`` (the variable fee the
    entity keeps, which is the source of the CSM). Pair with
    :func:`load_sample_vfa_model_points`; ``vfa.measure`` takes a single
    :class:`Basis`.
    """
    source = resources.files("fastcashflow") / "sample_data" / "sample_vfa_basis.xlsx"
    with resources.as_file(source) as path:
        return read_basis(path).resolve(("VAR_ANNUITY_A", "BANCA"))


def load_sample_vfa_model_points() -> ModelPoints:
    """Bundled VFA sample -- variable annuities with minimum-rate, death and
    maturity guarantees.

    Three single-premium account-value contracts that share a 2% minimum
    crediting rate (``minimum_crediting_rate``, uniform across the rows so the
    stochastic time-value pass applies) and differ in their floors: one
    carries both a death (GMDB) and a maturity (GMAB) guarantee, one a
    maturity floor, one a death floor. Pair with
    :func:`load_sample_vfa_basis`; generate underlying-return scenarios
    to value the time value of the guarantees (see ``examples/vfa.py``).
    """
    source = resources.files("fastcashflow") / "sample_data" / "sample_vfa_policies.csv"
    with resources.as_file(source) as path:
        # Carry the coverage taxonomy so a basis's coverages align at measure
        # time, even though these account-value contracts hold no coverages.
        return read_vfa_model_points(
            path, calculation_methods=load_sample_calculation_methods(),
        )


def load_sample_paa_basis() -> Basis:
    """Bundled PAA (Premium Allocation Approach) basis -- a single basis.

    A short-tail group-accident cover: a scalar valuation discount, a flat
    short-term lapse, and a ``settlement_pattern`` spreading each incurred
    inpatient claim over four months (a claims run-off). Pair with
    :func:`load_sample_paa_model_points`; ``paa.measure`` takes a single
    :class:`Basis`. The settlement pattern is why the discount is scalar: a
    per-year discount curve combined with a settlement pattern is rejected
    (discounting each settlement to its payment date needs a flat rate).
    """
    source = resources.files("fastcashflow") / "sample_data" / "sample_paa_basis.xlsx"
    with resources.as_file(source) as path:
        return read_basis(path).resolve(("ACCIDENT_A", "GA"))


def load_sample_paa_model_points() -> ModelPoints:
    """Bundled PAA sample -- two onerous 12-month group-accident contracts.

    Each carries a single inpatient (MORBIDITY) claim coverage; the premium is
    set below break-even so the block is onerous at inception, exercising the
    PAA onerous test on the settlement-discounted claims. Pair with
    :func:`load_sample_paa_basis`; measure with ``paa.measure``.
    """
    base = resources.files("fastcashflow") / "sample_data"
    with resources.as_file(base / "sample_paa_policies.csv") as policies, \
            resources.as_file(base / "sample_paa_coverages.csv") as coverages:
        return read_model_points(
            policies, coverages=coverages,
            calculation_methods=load_sample_calculation_methods(),
        )


def _drop_sample_table(filename: str, dest: Path | str) -> Path:
    """Drop a packaged single-table sample file at ``dest``, converting to
    whatever format ``dest`` 's extension picks (``.csv`` / ``.parquet`` /
    ``.feather`` / ``.arrow``).

    ``dest`` may be a file path (used as-is) or a directory (file lands
    inside with its original ``sample_*.csv`` name)."""
    src = resources.files("fastcashflow") / "sample_data" / filename
    dest_path = Path(dest)
    if dest_path.is_dir():
        dest_path = dest_path / filename
    with resources.as_file(src) as src_path:
        if dest_path.suffix == src_path.suffix:
            # Same format as the source -- a byte-for-byte copy preserves
            # any formatting the workbook editor cared about.
            import shutil
            shutil.copy2(src_path, dest_path)
        else:
            # Different format -- read the source as a polars frame and
            # let _write_frame route to the right writer by extension.
            _write_frame(_read_frame(src_path), dest_path)
    return dest_path


def _save_sample_basis(path: Path | str) -> Path:
    """Drop the packaged sample basis workbook on disk at ``path``.

    Use this to bootstrap a workbook a reader can open in Excel, inspect,
    and then re-read with :func:`read_basis` -- the same call shape
    a real user types against their own file. The bundled sample carries
    seven (product, channel) segments across three products
    (``TERM_LIFE_A``, ``HEALTH_A``, ``WHOLE_LIFE_A``).

    Supported extension: ``.xlsx`` (the workbook carries multiple sheets,
    so single-table formats like CSV are not appropriate here).

    ``path`` may be a file (the workbook lands there) or a directory (the
    workbook lands inside with its original ``sample_basis.xlsx``
    name). Returns the resolved destination path.
    """
    import shutil
    src = (resources.files("fastcashflow")
           / "sample_data" / "sample_basis.xlsx")
    dest_path = Path(path)
    if dest_path.is_dir():
        dest_path = dest_path / "sample_basis.xlsx"
    if dest_path.suffix.lower() != ".xlsx":
        raise ValueError(
            f"_save_sample_basis: expected an .xlsx path, got "
            f"{str(path)!r}. The basis workbook carries multiple "
            "sheets (mortality_tables, lapse_tables, segments, ...) and "
            "single-table formats (csv / parquet / feather) cannot "
            "represent it. Use .xlsx."
        )
    with resources.as_file(src) as src_path:
        shutil.copy2(src_path, dest_path)
    return dest_path


def _save_sample_policies(path: Path | str) -> Path:
    """Drop the packaged sample policies file on disk at ``path``.

    The companion to :func:`_save_sample_coverages` and
    :func:`_save_sample_calculation_methods`. Use the three together with
    :func:`read_model_points` for a copy-paste workflow that mirrors how
    you would read your own files.

    Supported extensions: ``.csv``, ``.xlsx``, ``.parquet``, ``.feather``
    / ``.arrow``. The conversion runs through polars when the requested
    extension differs from the packaged ``.csv`` source. ``.xlsx`` is
    capped at 1,048,576 rows per sheet -- for production-scale
    portfolios use ``.parquet`` or ``.feather``.
    """
    return _drop_sample_table("sample_policies.csv", path)


def _save_sample_coverages(path: Path | str) -> Path:
    """Drop the packaged sample coverages file on disk at ``path``.

    Coverage entries -- one row per (model point, coverage)
    -- the companion to :func:`_save_sample_policies`. A
    portfolio has roughly ``n_mp x avg_coverages_per_mp`` rows here, so
    this is the file most likely to exceed the 1,048,576 row cap of
    ``.xlsx``.

    Supported extensions: ``.csv``, ``.xlsx``, ``.parquet``, ``.feather``
    / ``.arrow``. ``.parquet`` or ``.feather`` for production scale.
    """
    return _drop_sample_table("sample_coverages.csv", path)


def _save_sample_calculation_methods(path: Path | str) -> Path:
    """Drop the packaged sample calculation-method catalogue on disk at ``path``.

    The company catalogue file -- one row per ``coverage`` mapping
    it to its :class:`CalculationMethod`. Tens-to-hundreds of rows in
    practice; ``.xlsx`` row cap never binds.

    Supported extensions: ``.csv``, ``.xlsx``, ``.parquet``, ``.feather``
    / ``.arrow``.
    """
    return _drop_sample_table("sample_calculation_methods.csv", path)


def _save_sample_inforce_state(path: Path | str) -> Path:
    """Drop the packaged sample in-force state file on disk at ``path``.

    The dynamic state-at-valuation companion to the static
    :func:`_save_sample_policies` file: one row per ``mp_id`` carrying
    the closing state from the prior reporting period
    (``elapsed_months``, ``count``, ``prior_csm``, ``lock_in_rate``,
    ``prior_count``). Pair the dropped file with
    :func:`read_inforce_state` and feed the result through
    :func:`apply_inforce_state`; the period close runs through
    :func:`fastcashflow.gmm.settle` (it needs ``prior_count``, the opening
    in-force), the diagnostic / runoff view through
    :func:`fastcashflow.gmm.measure_inforce`.

    Supported extensions: ``.csv``, ``.xlsx``, ``.parquet``, ``.feather``
    / ``.arrow``. One row per contract, so the ``.xlsx`` row cap
    (~1M / sheet) binds at the same scale as the policies file.
    """
    return _drop_sample_table("sample_inforce_state.csv", path)


def _save_sample_inforce_policies(path: Path | str) -> Path:
    """Drop a combined policies + in-force state sample file on disk at ``path``.

    The companion to :func:`read_inforce_policies`. Each row carries
    the permanent spec (issue_age, sex, term_months, premium_term_months,
    product, channel) and the closing state from the prior
    period (elapsed_months, count, prior_csm, lock_in_rate, prior_count).
    Built on
    the fly by joining the packaged ``sample_policies.csv`` and
    ``sample_inforce_state.csv`` on ``mp_id``; ``count`` is the state
    value (post-decrement), not the inception count.

    Supported extensions: ``.csv``, ``.xlsx``, ``.parquet``, ``.feather``
    / ``.arrow``. ``.xlsx`` is capped at ~1M rows / sheet.
    """
    base = resources.files("fastcashflow") / "sample_data"
    with resources.as_file(base / "sample_policies.csv") as policies, \
            resources.as_file(base / "sample_inforce_state.csv") as state:
        pol = pl.read_csv(policies)
        st = pl.read_csv(state)
    # ``count`` is on both files; drop the inception count from policies so
    # the state's post-decrement count is the one that survives the join.
    pol = pol.drop("count")
    combined = pol.join(st, on="mp_id", how="inner")
    dest_path = Path(path)
    if dest_path.is_dir():
        dest_path = dest_path / "sample_inforce_policies.csv"
    _write_frame(combined, dest_path)
    return dest_path


# ---------------------------------------------------------------------------
# Economic scenarios
# ---------------------------------------------------------------------------

def _optional_state_columns(df) -> dict:
    """Collect the optional per-MP state columns from a state-carrying
    frame -- absent columns stay ``None``. ``account_value`` feeds
    ``vfa.measure_inforce`` / ``vfa.settle``; the ``prior_*`` columns are
    the prior reporting date's figures the period-close settle entry
    points (``gmm.settle`` / ``vfa.settle``) need."""
    out: dict = {}
    for col in ("account_value", "prior_count", "prior_account_value",
                "prior_loss_component"):
        if col in df.columns:
            out[col] = df[col].to_numpy().astype(np.float64)
    return out


def read_inforce_state(path: Path | str) -> "InforceState":
    """Read an in-force state file -- the per-MP closing state from the
    prior reporting period.

    The file has one row per model point with columns ``mp_id``,
    ``elapsed_months``, ``count``, ``prior_csm`` and ``lock_in_rate``,
    plus the optional settlement columns ``prior_count`` and
    ``prior_loss_component`` (the prior reporting date's figures
    :func:`fastcashflow.gmm.settle` / :func:`fastcashflow.vfa.settle` need)
    and the VFA columns ``account_value`` (observed fund value at the
    valuation date) and ``prior_account_value``. Reads ``.parquet``,
    ``.csv``, ``.xlsx`` or ``.feather`` / ``.arrow`` via :func:`_read_frame`.

    Pair with :func:`apply_inforce_state` to join the state onto a
    :class:`ModelPoints` built from the static policies file, then settle
    the period with :func:`fastcashflow.gmm.settle` and the returned
    :class:`InforceState` (or value the book diagnostically with
    :func:`fastcashflow.gmm.measure_inforce`).

    ``lock_in_rate`` is required to be uniform across rows in v1 -- the
    engine takes a scalar locked-in rate. Cohort-aware per-MP rates are
    a future extension; for now the reader errors out if the column is
    not constant rather than silently dropping the per-row detail.
    """
    from fastcashflow.model_points import InforceState
    df = _read_frame(path)
    needed = ("mp_id", "elapsed_months", "count", "prior_csm", "lock_in_rate")
    for col in needed:
        if col not in df.columns:
            raise ValueError(
                f"the in-force state file is missing required column {col!r}"
            )
    # Uniform -> scalar; a cohort-aware column (per-row, paragraph B72(b)) is carried
    # per row, which gmm.settle partitions by rate.
    lock = df["lock_in_rate"].to_numpy().astype(np.float64)
    lock_in_rate = (float(lock[0]) if lock.size and np.all(lock == lock[0])
                    else lock)
    return InforceState(
        mp_id=df["mp_id"].to_numpy(),
        elapsed_months=df["elapsed_months"].to_numpy().astype(np.int64),
        count=df["count"].to_numpy().astype(np.float64),
        prior_csm=df["prior_csm"].to_numpy().astype(np.float64),
        lock_in_rate=lock_in_rate,
        **_optional_state_columns(df),
    )


def read_scenarios(path: Path | str) -> FloatArray:
    """Read a stochastic scenario set from a file.

    The file is a 2-D table -- one row per scenario, one column per
    projection month, every cell a rate or return. Reads ``.parquet``,
    ``.csv``, ``.xlsx`` or ``.feather`` / ``.arrow`` via :func:`_read_frame`.

    Returns a numpy ``float64`` array of shape ``(n_scenarios, n_time)``,
    or ``(n_scenarios,)`` when the file has a single column (flat-rate
    scenarios). The result is what :func:`measure_stochastic` and
    :func:`measure_tvog` accept as their ``scenarios`` / ``return_scenarios``
    input.

    Calibration -- Hull-White, Vasicek, regime-switching, climate paths,
    etc. -- is left to a separate scenario-generator step; this reader is
    just the storage / handover layer. For large scenario sets (thousands
    of paths) prefer ``.parquet`` or ``.feather`` over ``.xlsx``.
    """
    df = _read_frame(path)
    arr = df.to_numpy().astype(np.float64)
    if arr.shape[1] == 1:
        return arr[:, 0]
    return arr


# ---------------------------------------------------------------------------
# Measurement results
# ---------------------------------------------------------------------------

@singledispatch
def write_measurement(
    measurement,
    path: Path | str,
    *,
    ids: np.ndarray | None = None,
) -> None:
    """Write a measurement's per-model-point headline results to parquet / CSV.

    One row per model point, in model-point order. Pass ``ids`` for a leading
    ``id`` column so the results join back to policies. Dispatches on the
    measurement type -- GMM writes ``bel`` / ``ra`` / ``csm`` /
    ``loss_component``, PAA writes ``lrc`` / ``loss_component``, VFA adds
    ``variable_fee`` / ``time_value``, reinsurance held writes ``bel`` /
    ``ra`` / ``csm``. A mixed-portfolio
    :class:`~fastcashflow.portfolio.PortfolioMeasurement` writes one file
    per model present (``results.parquet`` becomes ``results-gmm.parquet`` /
    ``results-paa.parquet`` / ...), each with an ``id`` column joining its
    rows back to the portfolio. A new model registers its columns with
    ``@write_measurement.register`` in the module that defines its measurement
    type (so io.py stays free of the engine import).
    """
    raise TypeError(
        f"write_measurement does not handle {model_tag(measurement)}; pass a "
        f"{' / '.join(supported_model_tags(write_measurement))} measurement or "
        "a portfolio measurement"
    )


def _write_measurement_columns(
    columns: dict[str, np.ndarray], path: Path | str, ids: np.ndarray | None
) -> None:
    """Shared writer for the registered ``write_measurement`` implementations:
    an optional leading ``id`` column, then the model's headline columns."""
    out: dict[str, np.ndarray] = {}
    if ids is not None:
        out["id"] = np.asarray(ids)
    out.update(columns)
    _write_frame(pl.DataFrame(out), path)


def measure_stream(
    input_path: Path | str,
    output_dir: Path | str,
    basis: Basis | dict[tuple[str, str], Basis],
    *,
    coverages: Path | str | None = None,
    calculation_methods: Path | str | dict[str, CalculationMethod] | None = None,
    chunk_size: int = 20_000_000,
    backend: str = "cpu",
    id_column: str | None = None,
    validate_unique_mp_id: bool = True,
) -> int:
    """Stream a valuation through a parquet file one chunk at a time.

    Reads the input in chunks of ``chunk_size`` model points, values each
    chunk with the fused fast path (``measure(..., full=False)``), and writes
    the results as a parquet dataset -- one ``part-NNNNN.parquet`` file per
    chunk -- under ``output_dir``. Peak memory is a single chunk, so this
    scales past what an in-memory run could hold.

    The input is a policies + coverages pair, mirroring
    :func:`read_model_points`: ``input_path`` is the policies parquet and
    ``coverages`` the coverages parquet. Each chunk of policies pulls its
    coverage rows by ``mp_id``, so sorting the coverages file by ``mp_id``
    lets the parquet reader prune row groups. A flat one-row-per-policy
    (wide) file is not accepted -- it cannot carry the per-coverage waiting
    and reduction rules.

    ``basis`` may be a single :class:`Basis` (uniform portfolio) or a
    ``{(product, channel): Basis}`` dict, exactly as ``measure``.
    With a dict each chunk routes its model points to their segment's basis,
    so the policies parquet must carry ``product`` / ``channel``
    columns.

    ``id_column`` names the policies column written as the result ``id`` (so the
    output parquet joins back to a business key); it defaults to ``mp_id``. The
    coverages are always joined on ``mp_id`` regardless.

    ``validate_unique_mp_id`` (default ``True``) scans the whole policies file
    once up front and rejects a duplicate ``mp_id`` -- the same data error
    :func:`read_model_points` raises, which a chunk-by-chunk read would
    otherwise miss when the same id falls in different chunks. Set it ``False``
    to skip the scan when the upstream extract already guarantees uniqueness.

    Returns the total number of model points processed.
    """
    # Lazy import -- only ``measure_stream`` actually drives a valuation, so we
    # keep the engine import off the I/O hot path. A script that only reads
    # model points or writes results never pays the engine import cost.
    from fastcashflow._measurement.gmm import measure

    return _stream_policies_coverages(
        input_path, output_dir, coverages=coverages,
        calculation_methods=calculation_methods, chunk_size=chunk_size,
        id_column=id_column, validate_unique_mp_id=validate_unique_mp_id,
        measure_fn=lambda mp: measure(mp, basis, full=False, backend=backend),
    )


def _stream_policies_coverages(
    input_path: Path | str,
    output_dir: Path | str,
    *,
    coverages: Path | str | None,
    calculation_methods,
    chunk_size: int,
    id_column: str | None,
    validate_unique_mp_id: bool,
    measure_fn,
) -> int:
    """Shared out-of-core driver for the policies + coverages models.

    Reads ``input_path`` (policies parquet) in ``chunk_size`` blocks, pulls each
    block's coverage rows from ``coverages`` by ``mp_id``, builds the
    :class:`ModelPoints`, and writes ``measure_fn(model_points)`` to one
    ``part-NNNNN.parquet`` per chunk. ``measure_fn`` is the only model-specific
    piece -- ``gmm`` / ``paa`` / ``reinsurance`` pass their own measure closure.
    Returns the number of model points processed.
    """
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    if input_path.suffix != ".parquet":
        raise ValueError(
            f"measure_stream streams parquet input only; got {str(input_path)!r}"
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    if any(output_dir.glob("part-*.parquet")):
        raise ValueError(
            f"output directory {str(output_dir)!r} already contains part "
            "files; use a fresh directory"
        )

    if isinstance(calculation_methods, (str, Path)):
        methods_dict = _parse_calculation_methods(calculation_methods)
    else:
        methods_dict = calculation_methods
    scan = pl.scan_parquet(input_path)
    n_total = scan.select(pl.len()).collect().item()
    processed = 0

    # The result id is written from ``id_column`` (a business key the output
    # joins back on); coverages still join on ``mp_id``. Validate up front so a
    # typo'd id_column -- or a missing mp_id -- is a clear error, not a polars
    # ColumnNotFoundError leaking from a per-chunk read.
    schema_names = scan.collect_schema().names()
    if "mp_id" not in schema_names:
        raise ValueError(
            f"measure_stream: the policies file {str(input_path)!r} has no "
            "'mp_id' column; mp_id is the contract identity and the coverages "
            "join key (it is required even when id_column names a different "
            "result id)."
        )
    id_col = id_column if id_column is not None else "mp_id"
    if id_col not in schema_names:
        raise ValueError(
            f"measure_stream: id_column {id_col!r} is not a column of the "
            f"policies file {str(input_path)!r}"
        )

    # mp_id is the contract identity and the coverages join key. A chunk-by-chunk
    # read only sees one chunk's ids at a time, so a duplicate that straddles two
    # chunks would pass silently (and write a duplicate result id). Scan the
    # whole file once -- the same uniqueness read_model_points enforces in memory.
    if validate_unique_mp_id:
        dups = (
            scan.select("mp_id").group_by("mp_id").len()
            .filter(pl.col("len") > 1).head(5).collect()
        )
        if dups.height:
            raise ValueError(
                f"measure_stream: duplicate mp_id in {str(input_path)!r} (e.g. "
                f"{dups['mp_id'].to_list()}); mp_id is the contract identity / "
                "coverages join key and must be unique across the whole file. "
                "Pass validate_unique_mp_id=False to skip this scan when the "
                "upstream extract already guarantees uniqueness."
            )

    if coverages is not None:
        # chunk the policies, pull each chunk's coverage rows.
        cov_scan = pl.scan_parquet(Path(coverages))
        for part, offset in enumerate(range(0, n_total, chunk_size)):
            pol = scan.slice(offset, chunk_size).collect()
            ids = pol[id_col]
            cov = cov_scan.join(
                pol.lazy().select("mp_id"), on="mp_id", how="semi"
            ).collect()
            model_points = _model_points_from_frames(pol, cov, methods_dict)
            write_measurement(
                measure_fn(model_points),
                output_dir / f"part-{part:05d}.parquet",
                ids=ids.to_numpy(),
            )
            processed += model_points.n_mp
        return processed

    raise ValueError(
        "measure_stream needs a coverages frame: pass coverages=<parquet path> "
        "(an mp_id / coverage / amount frame). A flat one-row-per-policy "
        "(wide) file cannot carry per-coverage waiting / reduction rules and is "
        "not accepted."
    )


def _stream_validate(input_path: Path, output_dir: Path, id_column: str | None,
                     validate_unique_mp_id: bool,
                     entry: str = "measure_stream"):
    """Shared up-front checks for the streaming drivers: parquet input, empty
    output dir, mp_id present + unique, id_column present. Returns
    ``(scan, n_total, id_col)``."""
    if input_path.suffix != ".parquet":
        raise ValueError(
            f"{entry} streams parquet input only; got {str(input_path)!r}")
    output_dir.mkdir(parents=True, exist_ok=True)
    if any(output_dir.glob("part-*.parquet")):
        raise ValueError(
            f"output directory {str(output_dir)!r} already contains part "
            "files; use a fresh directory")
    scan = pl.scan_parquet(input_path)
    n_total = scan.select(pl.len()).collect().item()
    schema_names = scan.collect_schema().names()
    if "mp_id" not in schema_names:
        raise ValueError(
            f"{entry}: the policies file {str(input_path)!r} has no "
            "'mp_id' column; mp_id is the contract identity.")
    id_col = id_column if id_column is not None else "mp_id"
    if id_col not in schema_names:
        raise ValueError(
            f"{entry}: id_column {id_col!r} is not a column of the "
            f"policies file {str(input_path)!r}")
    if validate_unique_mp_id:
        dups = (scan.select("mp_id").group_by("mp_id").len()
                .filter(pl.col("len") > 1).head(5).collect())
        if dups.height:
            raise ValueError(
                f"{entry}: duplicate mp_id in {str(input_path)!r} (e.g. "
                f"{dups['mp_id'].to_list()}); mp_id must be unique across the "
                "whole file. Pass validate_unique_mp_id=False to skip this scan.")
    return scan, n_total, id_col


def _stream_single_file(
    input_path: Path | str,
    output_dir: Path | str,
    *,
    chunk_size: int,
    id_column: str | None,
    validate_unique_mp_id: bool,
    build_mp,
    measure_fn,
) -> int:
    """Out-of-core driver for a single-frame model (no coverages join), e.g. the
    VFA account-value book. ``build_mp(frame)`` turns each parquet chunk slice
    into a :class:`ModelPoints`; ``measure_fn(model_points)`` measures it. Writes
    one ``part-NNNNN.parquet`` per chunk; returns the model points processed."""
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    scan, n_total, id_col = _stream_validate(
        input_path, output_dir, id_column, validate_unique_mp_id)
    processed = 0
    for part, offset in enumerate(range(0, n_total, chunk_size)):
        pol = scan.slice(offset, chunk_size).collect()
        ids = pol[id_col]
        model_points = build_mp(pol)
        write_measurement(measure_fn(model_points),
                          output_dir / f"part-{part:05d}.parquet",
                          ids=ids.to_numpy())
        processed += model_points.n_mp
    return processed


# ---------------------------------------------------------------------------
# Out-of-core settlement -- the stream variant of gmm/vfa.settle
# ---------------------------------------------------------------------------

# The closing-state columns of a combined in-force file (the period-close
# snapshot layout of read_inforce_policies). Split off each chunk frame
# before the spec readers see it; ``count`` and ``account_value`` stay --
# they are valid policies columns too, and apply_inforce_state /
# vfa.settle take their state-side values from the InforceState.
_INFORCE_STATE_ONLY = ("elapsed_months", "prior_csm", "lock_in_rate",
                       "prior_count", "prior_account_value",
                       "prior_loss_component")

_STATE_REQUIRED = ("mp_id", "elapsed_months", "count", "prior_csm",
                   "lock_in_rate")


def _state_from_chunk(df, lock_in_rate: float | None) -> "InforceState":
    """Build the per-chunk :class:`InforceState` from a state-carrying frame
    slice. A float ``lock_in_rate`` is the globally validated scalar -- the
    chunk's own column is not re-read (uniformity was checked across the whole
    file, the v1 scalar contract). ``None`` means a cohort-aware book (paragraph
    B72(b)): the chunk's own per-row ``lock_in_rate`` column is carried instead,
    so the downstream ``gmm.settle`` partitions by rate; no global uniformity
    scan applies."""
    from fastcashflow.model_points import InforceState

    lir = (df["lock_in_rate"].to_numpy().astype(np.float64)
           if lock_in_rate is None else lock_in_rate)
    return InforceState(
        mp_id=df["mp_id"].to_numpy(),
        elapsed_months=df["elapsed_months"].to_numpy().astype(np.int64),
        count=df["count"].to_numpy().astype(np.float64),
        prior_csm=df["prior_csm"].to_numpy().astype(np.float64),
        lock_in_rate=lir,
        **_optional_state_columns(df),
    )


def _settle_stream_driver(
    input_path: Path | str,
    output_dir: Path | str,
    *,
    state_path: Path | str | None,
    chunk_size: int,
    id_column: str | None,
    validate_unique_mp_id: bool,
    build_mp,
    settle_fn,
    entry: str,
    cohort_aware_lock_in: bool = False,
) -> int:
    """Shared out-of-core driver for the settlement stream.

    Reads the in-force book in ``chunk_size`` blocks, assembles each block's
    ``(ModelPoints, InforceState)`` pair, and writes
    ``settle_fn(model_points, state)`` -- a per-MP settlement movement --
    to one ``part-NNNNN.parquet`` per chunk through the movement write
    arms. Two input layouts:

    * ``state_path is None`` -- ONE combined file: policies spec plus the
      closing-state columns in the same parquet (the period-close snapshot
      of :func:`read_inforce_policies`); the state columns are split off
      each chunk before ``build_mp`` sees the spec.
    * ``state_path`` given -- TWO files: a policies parquet plus a state
      parquet, semi-joined per chunk on ``mp_id``. A semi-join hides both
      missing and surplus rows, so the GLOBAL id sets are checked for
      bidirectional equality up front.

    ``lock_in_rate`` must be uniform across the whole book (the v1 scalar
    contract) -- validated globally, since a per-chunk check would pass a
    book whose rates differ only across chunks. ``cohort_aware_lock_in``
    lifts that restriction (only ``gmm.settle`` partitions by rate, paragraph
    B72(b)): the per-chunk state carries its own per-row ``lock_in_rate``
    column and ``settle_fn`` settles each rate cohort -- the global scan is
    skipped. Returns the number of model points processed.
    """
    from fastcashflow.model_points import apply_inforce_state

    if chunk_size < 1:
        raise ValueError(f"chunk_size must be >= 1, got {chunk_size}")
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    scan, n_total, id_col = _stream_validate(
        input_path, output_dir, id_column, validate_unique_mp_id,
        entry=entry)
    schema_names = scan.collect_schema().names()

    if state_path is None:
        missing = [c for c in _STATE_REQUIRED if c not in schema_names]
        if missing:
            raise ValueError(
                f"{entry}: the combined in-force file {str(input_path)!r} is "
                f"missing closing-state column(s) {missing}. Add them (the "
                "read_inforce_policies snapshot layout) or pass "
                "state_path=<state parquet> for the two-file layout."
            )
        state_scan = None
        lock_scan = scan
    else:
        state_path = Path(state_path)
        if state_path.suffix != ".parquet":
            raise ValueError(
                f"{entry} streams parquet state input only; got "
                f"{str(state_path)!r}")
        state_scan = pl.scan_parquet(state_path)
        st_names = state_scan.collect_schema().names()
        missing = [c for c in _STATE_REQUIRED if c not in st_names]
        if missing:
            raise ValueError(
                f"{entry}: the state file {str(state_path)!r} is missing "
                f"column(s) {missing} (the read_inforce_state layout)."
            )
        dups = (state_scan.select("mp_id").group_by("mp_id").len()
                .filter(pl.col("len") > 1).head(5).collect())
        if dups.height:
            raise ValueError(
                f"{entry}: duplicate mp_id in the state file "
                f"{str(state_path)!r} (e.g. {dups['mp_id'].to_list()}); one "
                "state row per contract."
            )
        # A semi-join silently STARVES a policies row with no state row and
        # silently IGNORES a state row with no policies row -- guard the
        # global id sets in both directions before any chunk is cut. Join on
        # the STRING form of mp_id: the in-memory join (align_inforce_state)
        # compares ids as strings, so an integer-id policies file matches a
        # string-id state file here too instead of a polars SchemaError.
        id_key = pl.col("mp_id").cast(pl.String).alias("__mp_id_str")
        pol_keys = scan.select(id_key)
        state_keys = state_scan.select(id_key)
        starved = (pol_keys.join(state_keys, on="__mp_id_str", how="anti")
                   .select(pl.len()).collect().item())
        ignored = (state_keys.join(pol_keys, on="__mp_id_str", how="anti")
                   .select(pl.len()).collect().item())
        if starved or ignored:
            raise ValueError(
                f"{entry}: the policies file and the state file must cover "
                f"exactly the same contracts -- {starved} policies row(s) "
                f"have no state row and {ignored} state row(s) have no "
                "policies row. A per-chunk semi-join would silently drop "
                "them; fix the extracts so the mp_id sets match."
            )
        lock_scan = state_scan

    if cohort_aware_lock_in:
        # The per-chunk state carries its own per-row lock_in_rate column;
        # gmm.settle partitions each chunk by rate (paragraph B72(b)). No global
        # uniformity scan -- a mixed book is the supported case.
        lock = None
    else:
        # v1 scalar lock-in, validated across the WHOLE book (the in-memory
        # readers check the same thing per file).
        locks = (lock_scan.select(pl.col("lock_in_rate").unique())
                 .collect()["lock_in_rate"].to_numpy())
        if locks.size > 1:
            raise NotImplementedError(
                f"{entry}: lock_in_rate must be uniform across rows in v1 "
                f"(found {locks.size} distinct values, e.g. "
                f"{np.sort(locks)[:3].tolist()}); per-MP (cohort-aware) lock-in "
                "rates are a future extension"
            )
        lock = float(locks[0]) if locks.size else 0.0

    processed = 0
    for part, offset in enumerate(range(0, n_total, chunk_size)):
        pol = scan.slice(offset, chunk_size).collect()
        ids = pol[id_col]
        if state_scan is None:
            sdf = pol
        else:
            id_key = pl.col("mp_id").cast(pl.String).alias("__mp_id_str")
            sdf = (state_scan.with_columns(id_key)
                   .join(pol.lazy().select(id_key), on="__mp_id_str",
                         how="semi")
                   .drop("__mp_id_str").collect())
        # Split the closing-state columns off the spec in BOTH layouts: in
        # the combined file they belong to the state object built above; in
        # the two-file layout a stray stale state column on the policies
        # file must not be absorbed as a grouping attribute (the state file
        # is the only state authority).
        spec = pol.drop([c for c in _INFORCE_STATE_ONLY
                         if c in pol.columns])
        state = _state_from_chunk(sdf, lock)
        model_points = apply_inforce_state(build_mp(spec), state)
        write_measurement(settle_fn(model_points, state),
                          output_dir / f"part-{part:05d}.parquet",
                          ids=ids.to_numpy())
        processed += model_points.n_mp
    return processed


def settle_stream(
    input_path: Path | str,
    output_dir: Path | str,
    basis: Basis | dict[tuple[str, str], Basis],
    *,
    coverages: Path | str | None = None,
    calculation_methods: Path | str | dict[str, CalculationMethod] | None = None,
    state_path: Path | str | None = None,
    period_months: int | None = None,
    chunk_size: int = 200_000,
    id_column: str | None = None,
    validate_unique_mp_id: bool = True,
) -> int:
    """Stream a paragraph-44 period close through a parquet file, chunk by
    chunk.

    The out-of-core variant of :func:`fastcashflow.gmm.settle`: reads the
    in-force book in ``chunk_size`` blocks, settles each block, and writes
    the per-MP settlement movements as a parquet dataset -- one
    ``part-NNNNN.parquet`` per chunk under ``output_dir``, every movement
    line plus the ``measurement_basis`` marker. Peak memory is one chunk's
    projection, so a book whose per-MP movements would not fit in memory
    still closes. Returns the number of model points processed.

    Input layouts (both produce identical output):

    * **One combined file** (primary): ``input_path`` carries the policies
      spec plus the closing-state columns (``elapsed_months``, ``count``,
      ``prior_csm``, ``lock_in_rate``, ``prior_count``,
      ``prior_loss_component``) -- the period-close snapshot of
      :func:`read_inforce_policies`.
    * **Two files**: ``input_path`` is the standard policies parquet and
      ``state_path`` the state parquet (the :func:`read_inforce_state`
      layout), semi-joined per chunk on ``mp_id``. The global id sets must
      match in both directions (validated up front -- a semi-join would
      silently drop a mismatch); a duplicate state ``mp_id`` is rejected
      like a duplicate policies ``mp_id``.

    ``coverages`` is the per-contract coverage parquet (required, as in
    :func:`~fastcashflow.gmm.measure_stream`). ``lock_in_rate`` may vary by
    row -- a cohort-aware book whose issue cohorts / GoCs locked in different
    inception rates (paragraph B72(b)): each chunk's settle partitions by rate, so
    the streamed close equals the in-memory :func:`~fastcashflow.gmm.settle`
    per contract.

    **Chaining on disk**: each part carries the closing-state columns --
    ``count``, ``lock_in_rate``, ``elapsed_months`` and the closing
    balances -- so the next period's state file is assembled from the
    parts alone: ``prior_csm <- csm_closing``, ``prior_loss_component <-
    loss_component_closing``, ``prior_count <- count``, then advance
    ``elapsed_months`` / ``count`` to the next observation. The disk side
    of :meth:`SettlementMovement.closing_inputs()
    <fastcashflow.gmm.SettlementMovement.closing_inputs>`.
    """
    from fastcashflow._measurement.gmm import settle

    build_mp = _coverages_build_mp(coverages, calculation_methods,
                                   entry="gmm.settle_stream")
    return _settle_stream_driver(
        input_path, output_dir, state_path=state_path, chunk_size=chunk_size,
        id_column=id_column, validate_unique_mp_id=validate_unique_mp_id,
        build_mp=build_mp,
        settle_fn=lambda mp, st: settle(mp, st, basis,
                                        period_months=period_months),
        entry="gmm.settle_stream",
        cohort_aware_lock_in=True,
    )


def _coverages_build_mp(coverages, calculation_methods, *, entry):
    """The coverages-based ``build_mp`` shared by the GMM / PAA / reinsurance
    settlement streams: scan the per-contract coverage parquet once and, for
    each policies chunk, semi-join its coverages and assemble the model points.
    The coverages frame is required -- a flat one-row-per-policy (wide) file
    cannot carry per-coverage waiting / reduction rules."""
    if coverages is None:
        raise ValueError(
            f"{entry} needs a coverages frame: pass coverages=<parquet path> "
            "(an mp_id / coverage / amount frame). A flat one-row-per-policy "
            "(wide) file cannot carry per-coverage waiting / reduction rules "
            "and is not accepted.")
    if isinstance(calculation_methods, (str, Path)):
        methods_dict = _parse_calculation_methods(calculation_methods)
    else:
        methods_dict = calculation_methods
    cov_scan = pl.scan_parquet(Path(coverages))

    def build_mp(spec):
        cov = cov_scan.join(
            spec.lazy().select("mp_id"), on="mp_id", how="semi"
        ).collect()
        return _model_points_from_frames(spec, cov, methods_dict)

    return build_mp
