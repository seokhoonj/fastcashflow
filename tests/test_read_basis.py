"""Basis workbook reader.

A single ``basis.xlsx`` carries the segment mapping plus the named rate
tables. The ``segments`` sheet has a ``_DEFAULTS`` row whose values blank
cells inherit, and one row per (product, channel) segment; the reader returns
one ``Basis`` per segment. See docs/basis-format.md.
"""
import fastcashflow as fcf
import numpy as np

from fastcashflow import ModelPoints
from fastcashflow.gmm import measure


def test_segments_resolve():
    """The sample workbook resolves to several (product, channel) segments."""
    basis = fcf.samples.basis()
    # The sample carries three products on FC/GA (HEALTH also adds TM).
    assert set(basis.segments) >= {
        ("TERM_LIFE_A", "FC"), ("TERM_LIFE_A", "GA"),
        ("HEALTH_A", "FC"), ("HEALTH_A", "GA"), ("HEALTH_A", "TM"),
        ("WHOLE_LIFE_A", "FC"), ("WHOLE_LIFE_A", "GA"),
    }


def test_defaults_inherited():
    """Blank cells in a segment row inherit from the ``_DEFAULTS`` row."""
    basis = fcf.samples.basis()
    ga, fc = basis.resolve(("TERM_LIFE_A", "GA")), basis.resolve(("TERM_LIFE_A", "FC"))
    # ra_confidence / mortality_cv / morbidity_cv live only on the defaults row
    assert ga.ra_confidence == 0.75 and fc.ra_confidence == 0.75
    assert ga.mortality_cv == 0.10 and fc.mortality_cv == 0.10
    assert ga.morbidity_cv == 0.12 and fc.morbidity_cv == 0.12
    # Shared economic curve -- inherited identically (a per-year forward
    # curve derived from the government-bond spot rates, not a flat scalar).
    assert np.allclose(np.asarray(ga.discount_annual),
                       np.asarray(fc.discount_annual))
    # Shared maintenance row in both segments' expense ledgers --
    # 90_000 per-policy; the 2% inflation is the global economic
    # assumption on the Basis object, not on the row itself.
    for basis in (ga, fc):
        maint = [r for r in basis.expense_items
                 if r.basis == "gamma_fixed"]
        assert len(maint) == 1
        assert maint[0].value == 90_000.0
        assert basis.expense_inflation == 0.02


def test_channel_segmented_lapse():
    """GA and FC reference different lapse tables -- the per-segment table
    reference. GA persistency is worse than FC."""
    basis = fcf.samples.basis()
    dur = np.arange(6)
    zero = np.zeros_like(dur)
    ga_lapse = basis.resolve(("TERM_LIFE_A", "GA")).lapse_annual(zero, zero, dur, zero, zero)
    fc_lapse = basis.resolve(("TERM_LIFE_A", "FC")).lapse_annual(zero, zero, dur, zero, zero)
    assert np.all(ga_lapse > fc_lapse)


def test_per_segment_acquisition_amount():
    """Acquisition cost differs per segment row (GA vs FC commission) --
    each segment points to its own expense_table_id in the
    ``expense_tables`` sheet."""
    basis = fcf.samples.basis()
    for (key, expected_acq) in (
        (("TERM_LIFE_A", "GA"),   800_000.0),
        (("TERM_LIFE_A", "FC"),   700_000.0),
        (("HEALTH_A",    "FC"),   800_000.0),
        (("HEALTH_A",    "GA"),   950_000.0),
        (("HEALTH_A",    "TM"),   250_000.0),
        (("WHOLE_LIFE_A","FC"), 1_600_000.0),
        (("WHOLE_LIFE_A","GA"), 1_900_000.0),
    ):
        rows = basis.resolve(key).expense_items
        acq = [r for r in rows if r.basis == "alpha_fixed"]
        assert len(acq) == 1, f"{key}: expected one per_policy_init row"
        assert acq[0].value == expected_acq, (key, expected_acq, acq[0].value)


def test_every_segment_has_expense_items():
    """The sample workbook attaches an ``expense_table`` to every segment;
    the loader populates ``Basis.expense_items`` on each."""
    basis = fcf.samples.basis()
    for basis in basis.segments.values():
        assert basis.expense_items                       # populated


def test_coverages_resolved():
    """Rate-driven coverages resolve from ``incidence_rate_tables`` (or
    ``mortality_tables`` for the general death coverage); the pattern
    taxonomy now lives in ``calculation_methods.csv`` (read by
    :func:`load_sample_calculation_methods`), no longer on the
    :class:`Basis`. The order matches the workbook's ``coverages``
    sheet rows; the engine treats every entry as an ordinary rate-driven
    coverage (no slot reserved)."""
    from fastcashflow import CalculationMethod

    basis = fcf.samples.basis()
    basis = basis.resolve(("TERM_LIFE_A", "GA"))
    assert [r.code for r in basis.coverages] == [
        "DEATH", "INPATIENT", "CANCER", "ADB", "DISEASE_DEATH",
    ]
    assert fcf.samples.calculation_methods() == {
        "DEATH":         CalculationMethod.DEATH,
        "INPATIENT":     CalculationMethod.MORBIDITY,
        "CANCER":        CalculationMethod.DIAGNOSIS,
        "ADB":           CalculationMethod.DEATH,
        "DISEASE_DEATH": CalculationMethod.DEATH,
        "ANNUITY":       CalculationMethod.ANNUITY,
        "MATURITY":      CalculationMethod.MATURITY,
    }


def test_resolved_basis_values():
    """A resolved ``Basis`` runs through ``value`` and ``measure``; the
    GA and FC segments give different BEL because lapse differs (channel
    segmentation actually bites the valuation)."""
    
    basis = fcf.samples.basis()
    mp = ModelPoints.single(issue_age=40, benefits={0: 100_000_000.0},
                            premium=50_000.0, term_months=120,
                            calculation_methods=fcf.samples.calculation_methods())
    # Strip surrender for the fused / detailed equivalence assertion below so
    # this test isolates the channel-segmentation effect on BEL. (The fast
    # path does carry surrender now -- see test_surrender_parity.py; here we
    # just want the two paths to agree on the no-surrender baseline.)
    import dataclasses
    basis_ga_no_surr = dataclasses.replace(
        basis.resolve(("TERM_LIFE_A", "GA")), surrender_value_curve=None)
    basis_fc_no_surr = dataclasses.replace(
        basis.resolve(("TERM_LIFE_A", "FC")), surrender_value_curve=None)
    ga = measure(mp, basis_ga_no_surr, full=False).bel[0]
    fc = measure(mp, basis_fc_no_surr, full=False).bel[0]
    assert np.isfinite(ga) and np.isfinite(fc)
    assert not np.isclose(ga, fc)
    # fused and detailed paths agree (when surrender is disabled).
    assert np.isclose(measure(mp, basis_ga_no_surr).bel_path[0, 0], ga)


# ---------------------------------------------------------------------------
# state_model column + STATE_MODELS registry (U+W)
# ---------------------------------------------------------------------------


def test_state_model_column_resolves_to_registry_entry():
    """The sample workbook's ``_DEFAULTS`` row carries
    ``state_model = WAIVER`` -- both segments inherit and resolve to
    ``STATE_MODELS['WAIVER']``.
    """
    from fastcashflow import STATE_MODELS
    basis = fcf.samples.basis()
    for basis in basis.segments.values():
        assert basis.state_model is STATE_MODELS["WAIVER"]


def test_state_model_column_blank_keeps_none():
    """A segment row with a blank ``state_model`` cell falls back to the
    defaults row; an empty defaults cell leaves ``Basis.state_model``
    as ``None`` (matching the engine's pre-registry behaviour).
    """
    import openpyxl, tempfile, shutil
    import importlib.resources as resources
    from fastcashflow import read_basis
    # Copy the sample workbook and clear the defaults row's state_model.
    sample = resources.files("fastcashflow").joinpath(
        "sample_data/sample_basis.xlsx")
    with tempfile.TemporaryDirectory() as d:
        dst = f"{d}/blank_state.xlsx"
        shutil.copy(sample, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["segments"]
        col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "state_model":
                col = c; break
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).value = None
        wb.save(dst)
        basis = read_basis(dst)
        for basis in basis.segments.values():
            assert basis.state_model is None


def test_state_model_unknown_key_raises():
    """An unrecognised state_model key is rejected at read time, with a
    hint listing the registry contents."""
    import openpyxl, tempfile, shutil, pytest
    import importlib.resources as resources
    from fastcashflow import read_basis
    sample = resources.files("fastcashflow").joinpath(
        "sample_data/sample_basis.xlsx")
    with tempfile.TemporaryDirectory() as d:
        dst = f"{d}/bad_state.xlsx"
        shutil.copy(sample, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["segments"]
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "state_model":
                ws.cell(row=2, column=c).value = "NOT_A_MODEL"
                break
        wb.save(dst)
        with pytest.raises(ValueError, match="NOT_A_MODEL"):
            read_basis(dst)


def test_surrender_column_basis_mismatch_raises():
    """A surrender_value_tables sheet whose value column (factor vs amount)
    contradicts the segment's surrender_value_basis is rejected at read time
    -- reading a factor as an amount (or vice versa) silently mis-measures.
    The sample carries a ``factor`` column and no surrender_value_basis (so
    every segment defaults to cum_premium_factor); renaming the column to
    ``amount`` without changing the basis is the mismatch."""
    import openpyxl, tempfile, shutil, pytest
    import importlib.resources as resources
    from fastcashflow import read_basis
    sample = resources.files("fastcashflow").joinpath(
        "sample_data/sample_basis.xlsx")
    with tempfile.TemporaryDirectory() as d:
        dst = f"{d}/bad_surrender.xlsx"
        shutil.copy(sample, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["surrender_value_tables"]
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "factor":
                ws.cell(row=1, column=c).value = "amount"   # now amount column
                break
        wb.save(dst)
        # amount column but cum_premium_factor basis (the default) -> mismatch
        with pytest.raises(ValueError, match="surrender_value_basis"):
            read_basis(dst)
