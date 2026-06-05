"""read_basis detects flexible segment axes -- N (or 1) columns, not just
(product, channel) -- and measure routes by them with no extra
argument. The axes are the leading non-assumption, non-``*_name`` columns.
"""
import dataclasses

import numpy as np
import openpyxl
import pytest

import fastcashflow as fcf
from fastcashflow import read_basis
from fastcashflow.io import SegmentedBasis


def _export(tmp_path):
    fcf.samples.export(tmp_path, template="gmm")
    return tmp_path / "basis.xlsx"


def _insert_risk_class(path, value_per_row="A"):
    """Insert a ``risk_class`` axis column after channel on the segments sheet."""
    wb = openpyxl.load_workbook(path)
    ws = wb["segments"]
    header = [c.value for c in ws[1]]
    after = header.index("channel") + 1          # 1-based col of channel
    ws.insert_cols(after + 1)
    ws.cell(row=1, column=after + 1, value="risk_class")
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=after + 1, value=value_per_row)
    wb.save(path)


def test_read_basis_returns_segmented_basis_default_axes(tmp_path):
    basis = read_basis(_export(tmp_path))
    assert isinstance(basis, SegmentedBasis)
    assert basis.segment_axes == ("product", "channel")
    assert all(isinstance(k, tuple) and len(k) == 2 for k in basis)   # 2-tuple keys
    mp = fcf.samples.model_points()
    assert fcf.gmm.measure(mp, basis).bel.shape[0] == mp.n_mp          # routes, no segment_by


def test_read_basis_detects_extra_axis(tmp_path):
    path = _export(tmp_path)
    _insert_risk_class(path)
    basis = read_basis(path)
    assert basis.segment_axes == ("product", "channel", "risk_class")
    assert all(len(k) == 3 and k[2] == "A" for k in basis)            # 3-tuple keys


def test_measure_routes_by_file_declared_axes(tmp_path):
    path = _export(tmp_path)
    _insert_risk_class(path)
    basis = read_basis(path)
    mp = fcf.samples.model_points()
    mp = dataclasses.replace(mp, attributes={"risk_class": np.array(["A"] * mp.n_mp)})
    m = fcf.gmm.measure(mp, basis)              # segment_by auto = the 3 file axes
    assert m.bel.shape[0] == mp.n_mp


def test_read_basis_detects_axis_after_assumption_column(tmp_path):
    """Axis detection is order-independent -- an axis among the assumption
    columns is still picked up (not only the leading columns)."""
    path = _export(tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["segments"]
    header = [c.value for c in ws[1]]
    after = header.index("mortality_table") + 1      # insert AFTER an assumption column
    ws.insert_cols(after + 1)
    ws.cell(row=1, column=after + 1, value="risk_class")
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=after + 1, value="A")
    wb.save(path)
    basis = read_basis(path)
    assert "risk_class" in basis.segment_axes         # detected despite position
    assert all(k[-1] == "A" for k in basis)


def test_read_basis_rejects_ae_axis_not_in_segments(tmp_path):
    """An A/E keyed on an axis absent from segments would silently never match."""
    path = _export(tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb.create_sheet("ae_factors")
    ws.append(["product", "region", "coverage", "factor"])  # region not a segment axis
    ws.append(["TERM_LIFE_A", "SEOUL", "DEATH", 1.1])
    wb.save(path)
    with pytest.raises(ValueError, match="not in the segments sheet"):
        read_basis(path)
