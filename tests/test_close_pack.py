"""Phase-3 reporting layer: write_close_pack -- the multi-sheet xlsx artifact.

write_close_pack serialises a ClosePackage to a multi-sheet workbook (the
aggregate IFRS 17 statements) plus an optional per-model-point parquet sidecar
(Excel's row limit keeps the per-MP detail out of the workbook).
"""
import dataclasses

import numpy as np
import openpyxl
import polars as pl
import pytest

import fastcashflow as fcf
from fastcashflow import InforceState, ModelPoints
from fastcashflow.closing import close
from fastcashflow.disclosure import write_close_pack
from fastcashflow.movement import (
    GMMSettlementReconciliation, ReinsuranceSettlementReconciliation)
from fastcashflow.report import Report
from conftest import PATTERNS, make_death_basis


def _build(cls, **over):
    kw = {}
    for f in dataclasses.fields(cls):
        if f.name in over:
            kw[f.name] = over[f.name]
        elif f.name == "period_months":
            kw[f.name] = 12
        elif f.name == "revenue_basis":
            kw[f.name] = "time"
        else:
            kw[f.name] = 0.0
    return cls(**kw)


def _report(revenue):
    revenue = np.asarray(revenue, dtype=float)
    z = np.zeros_like(revenue)
    return Report(
        insurance_revenue=revenue, insurance_service_expense=z,
        insurance_service_result=revenue, insurance_finance_expense=z,
        bel_finance_expense=z, ra_finance_expense=z, csm_finance_expense=z,
        loss_component=np.zeros(revenue.shape[0]), csm_opening=z,
        csm_accretion=z, csm_release=z, csm_closing=z)


def _sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return header, [dict(zip(header, r)) for r in rows[1:]]


def _settlement_movement():
    """A real GMM settlement movement + its reconciliation (for the sidecar)."""
    basis = make_death_basis(
        mortality_q=0.02, lapse_q=0.03, discount_annual=0.03,
        ra_confidence=0.75, mortality_cv=0.10)
    surv = fcf.gmm.measure(
        ModelPoints(issue_age=np.array([40]), premium=np.array([100.0]),
                    term_months=np.array([36]), benefits={0: np.array([1e6])},
                    count=np.array([1.0]), calculation_methods=PATTERNS),
        basis, full=True).cashflows.inforce[0]
    eo, ec, scale = 12, 24, 1000.0
    mp = ModelPoints(
        issue_age=np.array([40]), premium=np.array([100.0]),
        term_months=np.array([36]), benefits={0: np.array([1e6])},
        count=np.array([scale * surv[ec]]), elapsed_months=np.array([ec]),
        mp_id=np.array(["P0"]), product=np.array(["A"]),
        calculation_methods=PATTERNS)
    st = InforceState(
        mp_id=np.array(["P0"]), elapsed_months=np.array([ec]),
        count=np.array([scale * surv[ec]]), prior_csm=np.array([0.0]),
        lock_in_rate=0.03, prior_count=np.array([scale * surv[eo]]))
    mv = fcf.gmm.settle(mp, st, basis, period_months=12)
    return mv, fcf.reconcile([mv])[0]


def test_close_pack_writes_the_aggregate_sheets(tmp_path):
    gmm = _build(GMMSettlementReconciliation, bel_closing=700.0, ra_closing=200.0,
                 csm_closing=100.0, lic_closing=300.0, bel_interest=4.0)
    held = _build(ReinsuranceSettlementReconciliation, bel_closing=-150.0,
                  bel_interest=-1.0)
    pack = close([gmm, held], group_ids=["GoC-1", "RE-1"])
    out = tmp_path / "close_pack.xlsx"
    write_close_pack(pack, out)

    wb = openpyxl.load_workbook(out)
    # service result absent (no reports given)
    assert wb.sheetnames == ["00_Index", "01_SoFP", "03_Finance",
                             "04_Reconciliation"]
    # SoFP carries the kind x component grid; net carrying amount foots
    header, rows = _sheet_rows(wb["01_SoFP"])
    assert header == ("kind", "component", "opening", "change", "closing")
    net_total = [r for r in rows if r["kind"] == "Net" and r["component"] == "Total"]
    # one signed frame: the -150 reinsurance recoverable is added in (reduces net)
    assert net_total[0]["closing"] == pytest.approx(1300.0 + (-150.0))


def test_reconciliation_sheet_materialises_rich_audit_columns(tmp_path):
    gmm = _build(GMMSettlementReconciliation, bel_closing=100.0, lic_closing=10.0)
    pack = close([gmm], group_ids=["GoC-1"])
    out = tmp_path / "pack.xlsx"
    write_close_pack(pack, out)
    wb = openpyxl.load_workbook(out)
    header, rows = _sheet_rows(wb["04_Reconciliation"])
    for col in ("model", "block", "line", "amount", "line_code",
                "ifrs17_paragraph", "is_memo", "sort_order"):
        assert col in header
    # the audit anchor came from the line registry join
    opening = [r for r in rows if r["block"] == "BEL" and r["line"] == "Opening"]
    assert opening[0]["line_code"] == "bel_opening"
    assert opening[0]["ifrs17_paragraph"] == "100(a)"


def test_service_result_sheet_present_when_reports_given(tmp_path):
    gmm = _build(GMMSettlementReconciliation, bel_closing=100.0)
    pack = close([gmm], reports=[_report(np.array([[10.0] * 12]))])
    out = tmp_path / "pack.xlsx"
    write_close_pack(pack, out)
    wb = openpyxl.load_workbook(out)
    assert "02_Service_Result" in wb.sheetnames


def test_index_sheet_lists_models_and_sidecar(tmp_path):
    mv, recon = _settlement_movement()
    pack = close([recon], group_ids=["GoC-1"])
    out = tmp_path / "pack.xlsx"
    write_close_pack(pack, out, movements=mv)

    # the per-MP sidecar parquet was written beside the workbook
    sidecar = tmp_path / "pack_permp.parquet"
    assert sidecar.exists()
    side = pl.read_parquet(sidecar)
    assert side.height >= 1

    wb = openpyxl.load_workbook(out)
    _header, rows = _sheet_rows(wb["00_Index"])
    index = {r["item"]: r["value"] for r in rows}
    assert index["Reporting period (months)"] == "12"
    assert "gmm" in index["Models"]
    assert index["Per-MP detail"] == "pack_permp.parquet"


def test_sidecar_naming_keys_off_call_shape(tmp_path):
    """A single movement -> a bare sidecar; a list -> indexed, even for one."""
    mv, recon = _settlement_movement()
    pack = close([recon])
    out = tmp_path / "pack.xlsx"
    write_close_pack(pack, out, movements=[mv])      # list, one entry
    assert (tmp_path / "pack_permp_0.parquet").exists()
    assert not (tmp_path / "pack_permp.parquet").exists()
    wb = openpyxl.load_workbook(out)
    _header, rows = _sheet_rows(wb["00_Index"])
    index = {r["item"]: r["value"] for r in rows}
    assert index["Per-MP detail"] == "pack_permp_0.parquet"


def test_close_pack_rejects_non_xlsx(tmp_path):
    gmm = _build(GMMSettlementReconciliation, bel_closing=1.0)
    pack = close([gmm])
    with pytest.raises(ValueError, match="xlsx"):
        write_close_pack(pack, tmp_path / "pack.parquet")
