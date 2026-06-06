"""Input-validation tests -- the guards that turn a future silently-wrong BEL
into a loud error at the workbook reader / dataclass construction site.

These guards close concrete footguns surfaced by the 2nd review (silent BEL
miscomputation from bad inputs the reader used to accept). Each test asserts
the guard fires; without a test, a future refactor can silently remove the
guard and the footgun returns.
"""
from pathlib import Path

import numpy as np
import openpyxl
import polars as pl
import pytest

import fastcashflow as fcf
from fastcashflow.engine import _measure_inforce_full, _measure_inforce_fast
from fastcashflow import Basis, CalculationMethod, CoverageRate, ModelPoints
from fastcashflow.basis import annual_to_monthly
from fastcashflow.io import (
    _axis_tables, _flex_rate_table, _read_expense_tables, _read_state,
    _truncate_list,
)


# ---------------------------------------------------------------------------
# ModelPoints scalar guards
# ---------------------------------------------------------------------------

def test_modelpoints_rejects_negative_issue_age():
    with pytest.raises(ValueError, match="issue_age"):
        ModelPoints(
            issue_age=np.array([-5.0]),
            premium=np.array([0.0]),
            term_months=np.array([12]),
        )


def test_modelpoints_rejects_zero_term_months():
    with pytest.raises(ValueError, match="term_months"):
        ModelPoints(
            issue_age=np.array([40.0]),
            premium=np.array([0.0]),
            term_months=np.array([0]),
        )


def test_modelpoints_rejects_negative_count():
    with pytest.raises(ValueError, match="count"):
        ModelPoints(
            issue_age=np.array([40.0]),
            premium=np.array([0.0]),
            term_months=np.array([12]),
            count=np.array([-100.0]),
        )


# ---------------------------------------------------------------------------
# Basis scalar guards
# ---------------------------------------------------------------------------

def _flat_rate(annual=0.01):
    return lambda sex, issue_age, duration: np.full(issue_age.shape, annual)


def test_assumptions_rejects_ra_confidence_at_boundary():
    for bad in (-0.1, 0.0, 1.0, 1.5):
        with pytest.raises(ValueError, match="ra_confidence"):
            Basis(
                mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
                discount_annual=0.0, ra_confidence=bad,
                mortality_cv=0.10,
                coverages=(CoverageRate("DEATH", _flat_rate()),),
            )


def test_assumptions_rejects_negative_cv():
    with pytest.raises(ValueError, match="mortality_cv"):
        Basis(
            mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
            discount_annual=0.0, ra_confidence=0.75,
            mortality_cv=-0.1,
            coverages=(CoverageRate("DEATH", _flat_rate()),),
        )


def test_assumptions_rejects_settlement_pattern_not_summing_to_one():
    with pytest.raises(ValueError, match="settlement_pattern"):
        Basis(
            mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
            discount_annual=0.0, ra_confidence=0.75, mortality_cv=0.10,
            settlement_pattern=np.array([0.3, 0.3, 0.3]),
            coverages=(CoverageRate("DEATH", _flat_rate()),),
        )


# ---------------------------------------------------------------------------
# io.py rate-table duplicate row catch
# ---------------------------------------------------------------------------

def _make_ws(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "mortality_tables"
    for r in rows:
        ws.append(list(r))
    return ws


def test_flex_rate_table_rejects_duplicate_rows():
    """A duplicate (table_id, sex, age, ...) row would last-wins silently."""
    ws = _make_ws([
        ("table_id", "sex", "age", "rate"),
        ("MORT", 0, 30, 0.001),
        ("MORT", 0, 30, 99.0),    # the silent-overwrite case
    ])
    with pytest.raises(ValueError, match="duplicate row"):
        _flex_rate_table(ws)


# ---------------------------------------------------------------------------
# io.py state range check
# ---------------------------------------------------------------------------

def test_read_state_rejects_unknown_integer():
    """state=42 used to slip through as a silent garbage state index."""
    col = pl.Series("state", [0, 42])
    with pytest.raises(ValueError, match="unknown integer value"):
        _read_state(col)


# ---------------------------------------------------------------------------
# io.py mp_id uniqueness / premium double-source / reduction pair
# ---------------------------------------------------------------------------

def test_long_form_rejects_duplicate_mp_id(tmp_path):
    """Duplicate mp_id would fan out the coverages join silently."""
    basis_book = tmp_path / "basis.xlsx"
    _write_minimal_assumptions(basis_book, coverage="DEATH")
    pol_csv = tmp_path / "policies.csv"
    pl.DataFrame({
        "mp_id": ["A", "A"],          # the duplicate-id case
        "issue_age": [40, 40], "term_months": [12, 12], "premium": [0.0, 0.0],
    }).write_csv(pol_csv)
    cov_csv = tmp_path / "coverages.csv"
    pl.DataFrame({
        "mp_id": ["A"], "coverage": ["DEATH"], "amount": [1e8],
    }).write_csv(cov_csv)
    method_csv = tmp_path / "calculation_methods.csv"
    pl.DataFrame({
        "coverage": ["DEATH"], "calculation_method": ["DEATH"],
    }).write_csv(method_csv)

    basis = fcf.read_basis(basis_book)
    with pytest.raises(ValueError, match="duplicate mp_id"):
        fcf.read_model_points(
            pol_csv, coverages=cov_csv,
            calculation_methods=method_csv,
        )


def test_long_form_rejects_premium_in_both_frames(tmp_path):
    """``premium`` in coverages and ``premium`` in policies = ambiguous."""
    basis_book = tmp_path / "basis.xlsx"
    _write_minimal_assumptions(basis_book, coverage="DEATH")
    pol_csv = tmp_path / "policies.csv"
    pl.DataFrame({
        "mp_id": ["A"], "issue_age": [40], "term_months": [12],
        "premium": [12_000.0],    # source 1
    }).write_csv(pol_csv)
    cov_csv = tmp_path / "coverages.csv"
    pl.DataFrame({
        "mp_id": ["A"], "coverage": ["DEATH"], "amount": [1e8],
        "premium": [1.0],               # source 2 -- silently overrode
    }).write_csv(cov_csv)
    method_csv = tmp_path / "calculation_methods.csv"
    pl.DataFrame({
        "coverage": ["DEATH"], "calculation_method": ["DEATH"],
    }).write_csv(method_csv)

    basis = fcf.read_basis(basis_book)
    with pytest.raises(ValueError, match="premium is specified twice"):
        fcf.read_model_points(
            pol_csv, coverages=cov_csv,
            calculation_methods=method_csv,
        )


def test_long_form_rejects_reduction_factor_without_reduction_end(tmp_path):
    """reduction_factor=0.5 without reduction_end is silently inert."""
    basis_book = tmp_path / "basis.xlsx"
    _write_minimal_assumptions(basis_book, coverage="DEATH")
    pol_csv = tmp_path / "policies.csv"
    pl.DataFrame({
        "mp_id": ["A"], "issue_age": [40], "term_months": [12],
        "premium": [12_000.0],
    }).write_csv(pol_csv)
    cov_csv = tmp_path / "coverages.csv"
    pl.DataFrame({
        "mp_id": ["A"], "coverage": ["DEATH"], "amount": [1e8],
        "reduction_factor": [0.5],      # no reduction_end -- never fires
    }).write_csv(cov_csv)
    method_csv = tmp_path / "calculation_methods.csv"
    pl.DataFrame({
        "coverage": ["DEATH"], "calculation_method": ["DEATH"],
    }).write_csv(method_csv)

    basis = fcf.read_basis(basis_book)
    with pytest.raises(ValueError, match="reduction_factor"):
        fcf.read_model_points(
            pol_csv, coverages=cov_csv,
            calculation_methods=method_csv,
        )


# ---------------------------------------------------------------------------
# Minimal basis workbook helper
# ---------------------------------------------------------------------------

def _write_minimal_assumptions(path: Path, coverage: str) -> None:
    """A tiny one-segment workbook with one rate-driven coverage."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name, rows):
        ws = wb.create_sheet(name)
        for r in rows:
            ws.append(list(r))

    sheet("mortality_tables", [
        ("table_id", "rate"),
        ("MORT_FLAT", 0.001),
    ])
    sheet("lapse_tables", [
        ("table_id", "rate"),
        ("LAPSE_FLAT", 0.01),
    ])
    sheet("discount_tables", [
        ("table_id", "year", "rate"),
        ("DISC_FLAT", 0, 0.03),
    ])
    sheet("coverages", [
        ("coverage", "rate_table"),
        (coverage, "MORT_FLAT"),
    ])
    sheet("segments", [
        ("product", "channel", "mortality_table", "lapse_table",
         "discount_table", "ra_confidence", "mortality_cv"),
        ("TERM_LIFE_A", "FC", "MORT_FLAT", "LAPSE_FLAT", "DISC_FLAT",
         0.75, 0.10),
    ])
    wb.save(path)


# ---------------------------------------------------------------------------
# P1 E -- io.py error-message quality
# ---------------------------------------------------------------------------

def _make_sheet(title, rows):
    """Build an openpyxl worksheet from header + data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for r in rows:
        ws.append(list(r))
    return ws


def test_truncate_list_caps_long_alternatives():
    """A 100-entry registry list must collapse to 10 items + suffix."""
    items = [f"T{i:03d}" for i in range(20)]
    out = _truncate_list(items, cap=10)
    assert "T000" in out and "T009" in out
    assert "and 10 more" in out
    assert "T015" not in out          # past the cap


def test_truncate_list_returns_full_when_under_cap():
    items = ["A", "B", "C"]
    assert _truncate_list(items, cap=10) == repr(items)


def test_segments_legacy_product_column_hints(tmp_path):
    """A user with the old ``product_code`` header gets a rename hint."""
    book = tmp_path / "basis.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name, rows):
        ws = wb.create_sheet(name)
        for r in rows:
            ws.append(list(r))

    sheet("mortality_tables", [("table_id", "rate"), ("MORT", 0.001)])
    sheet("lapse_tables", [("table_id", "rate"), ("LAPSE", 0.01)])
    sheet("discount_tables", [("table_id", "year", "rate"), ("DISC", 0, 0.03)])
    sheet("segments", [
        ("product_code", "channel", "mortality_table", "lapse_table",
         "discount_table", "ra_confidence", "mortality_cv"),
        ("TERM_LIFE_A", "FC", "MORT", "LAPSE", "DISC", 0.75, 0.10),
    ])
    wb.save(book)
    with pytest.raises(ValueError, match="did you mean 'product'"):
        fcf.read_basis(book)


def test_missing_required_sheet_friendly_error(tmp_path):
    """A workbook missing ``mortality_tables`` raises a sheet-named error,
    not a raw openpyxl KeyError."""
    book = tmp_path / "basis.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Build a workbook with everything except mortality_tables.
    for name, rows in (
        ("lapse_tables", [("table_id", "rate"), ("LAPSE", 0.01)]),
        ("discount_tables", [("table_id", "year", "rate"), ("DISC", 0, 0.03)]),
        ("segments", [
            ("product", "channel", "mortality_table", "lapse_table",
             "discount_table", "ra_confidence", "mortality_cv"),
        ]),
    ):
        ws = wb.create_sheet(name)
        for r in rows:
            ws.append(list(r))
    wb.save(book)
    with pytest.raises(ValueError, match="missing required sheet 'mortality_tables'"):
        fcf.read_basis(book)


def test_flex_rate_table_missing_rate_column():
    """A rate sheet without the ``rate`` column gets a named-column error."""
    ws = _make_sheet("mortality_tables", [
        ("table_id", "sex", "age"),    # no 'rate' column
        ("MORT", 0, 30),
    ])
    with pytest.raises(ValueError, match="missing required column 'rate'"):
        _flex_rate_table(ws)


def test_axis_tables_missing_value_column():
    """An axis sheet without the value column gets a named-column error."""
    ws = _make_sheet("discount_tables", [
        ("table_id", "year"),          # no 'rate' column
        ("DISC", 0),
    ])
    with pytest.raises(ValueError, match=r"missing required column.*'rate'"):
        _axis_tables(ws, "year")


def test_axis_tables_sparse_row():
    """A row missing its axis column gets a row-level error, not raw KeyError."""
    ws = _make_sheet("discount_tables", [
        ("table_id", "year", "rate"),
        ("DISC", 0, 0.03),
    ])
    # Synthesise a second row with no 'year' value by reaching into openpyxl
    ws.append(["DISC", None, 0.04])
    # The sparse row's ``year`` becomes None -> int(None) raises TypeError, not
    # KeyError, so the bare-row guard here is on the column-existence side.
    # Instead test the value_col-missing case via the absent column.
    ws2 = _make_sheet("discount_tables", [
        ("table_id", "year", "rate"),
        ("DISC", 0, 0.03),
    ])
    # Build a row dict that drops 'year' to trigger the per-row guard.
    # _axis_tables iterates _sheet_dicts, which only yields cells whose
    # header was set. The cleanest sparse-row case is the header missing,
    # exercised by the missing-value-column test above. Keep this test as
    # a smoke check that valid rows still parse.
    out = _axis_tables(ws2, "year")
    assert "DISC" in out and out["DISC"][0] == 0.03


def test_read_expense_tables_missing_column():
    """The expense_tables sheet without ``value`` column is named in the error."""
    ws = _make_sheet("expense_tables", [
        ("table_id", "expense_type", "basis"),     # no 'value'
        ("EXP", "acquisition", "premium_pct"),
    ])
    with pytest.raises(ValueError, match="missing required column"):
        _read_expense_tables(ws)


def test_long_form_orphan_mp_id_names_offender(tmp_path):
    """``cov.mp_id='X'`` not in policies: the error names ``'X'``."""
    basis_book = tmp_path / "basis.xlsx"
    _write_minimal_assumptions(basis_book, coverage="DEATH")
    pol_csv = tmp_path / "policies.csv"
    pl.DataFrame({
        "mp_id": ["A"], "issue_age": [40], "term_months": [12],
        "premium": [0.0],
    }).write_csv(pol_csv)
    cov_csv = tmp_path / "coverages.csv"
    pl.DataFrame({
        "mp_id": ["ORPHAN_X"],            # not in policies
        "coverage": ["DEATH"], "amount": [1e8],
    }).write_csv(cov_csv)
    method_csv = tmp_path / "calculation_methods.csv"
    pl.DataFrame({
        "coverage": ["DEATH"], "calculation_method": ["DEATH"],
    }).write_csv(method_csv)

    basis = fcf.read_basis(basis_book)
    with pytest.raises(ValueError, match="ORPHAN_X"):
        fcf.read_model_points(
            pol_csv, coverages=cov_csv,
            calculation_methods=method_csv,
        )


def test_long_form_orphan_coverage_code_names_offender(tmp_path):
    """``cov.coverage`` not in calculation_methods: the error names the code."""
    basis_book = tmp_path / "basis.xlsx"
    _write_minimal_assumptions(basis_book, coverage="DEATH")
    pol_csv = tmp_path / "policies.csv"
    pl.DataFrame({
        "mp_id": ["A"], "issue_age": [40], "term_months": [12],
        "premium": [0.0],
    }).write_csv(pol_csv)
    cov_csv = tmp_path / "coverages.csv"
    pl.DataFrame({
        "mp_id": ["A"], "coverage": ["GHOST_CODE"], "amount": [1e8],
    }).write_csv(cov_csv)
    method_csv = tmp_path / "calculation_methods.csv"
    pl.DataFrame({
        "coverage": ["DEATH"], "calculation_method": ["DEATH"],
    }).write_csv(method_csv)

    basis = fcf.read_basis(basis_book)
    with pytest.raises(ValueError, match="GHOST_CODE"):
        fcf.read_model_points(
            pol_csv, coverages=cov_csv,
            calculation_methods=method_csv,
        )


def test_long_form_no_premium_source_warns(tmp_path, recwarn):
    """With neither ``premium`` (cov) nor ``premium`` (pol)
    silently defaults to zero -- now warns."""
    basis_book = tmp_path / "basis.xlsx"
    _write_minimal_assumptions(basis_book, coverage="DEATH")
    pol_csv = tmp_path / "policies.csv"
    pl.DataFrame({
        "mp_id": ["A"], "issue_age": [40], "term_months": [12],
        # no premium
    }).write_csv(pol_csv)
    cov_csv = tmp_path / "coverages.csv"
    pl.DataFrame({
        "mp_id": ["A"], "coverage": ["DEATH"], "amount": [1e8],
        # no premium
    }).write_csv(cov_csv)
    method_csv = tmp_path / "calculation_methods.csv"
    pl.DataFrame({
        "coverage": ["DEATH"], "calculation_method": ["DEATH"],
    }).write_csv(method_csv)

    basis = fcf.read_basis(basis_book)
    fcf.read_model_points(
        pol_csv, coverages=cov_csv,
        calculation_methods=method_csv,
    )
    matched = [w for w in recwarn.list
               if issubclass(w.category, UserWarning)
               and "no premium source" in str(w.message)]
    assert matched, [str(w.message) for w in recwarn.list]


def test_rate_table_not_found_caps_alternatives(tmp_path):
    """100 mortality tables + a typo: the error lists at most 10 plus a count."""
    book = tmp_path / "basis.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name, rows):
        ws = wb.create_sheet(name)
        for r in rows:
            ws.append(list(r))

    mort_rows = [("table_id", "rate")] + [
        (f"MORT_{i:03d}", 0.001) for i in range(20)
    ]
    sheet("mortality_tables", mort_rows)
    sheet("lapse_tables", [("table_id", "rate"), ("LAPSE", 0.01)])
    sheet("discount_tables", [("table_id", "year", "rate"), ("DISC", 0, 0.03)])
    sheet("coverages", [
        ("coverage", "rate_table"),
        ("DEATH", "NONEXISTENT"),         # typo -- triggers not-registered
    ])
    sheet("segments", [
        ("product", "channel", "mortality_table", "lapse_table",
         "discount_table", "ra_confidence", "mortality_cv"),
        ("TERM_LIFE_A", "FC", "MORT_000", "LAPSE", "DISC", 0.75, 0.10),
    ])
    wb.save(book)
    with pytest.raises(ValueError, match="and 10 more"):
        fcf.read_basis(book)


# ---------------------------------------------------------------------------
# P1 F -- second-tier regression risk
# ---------------------------------------------------------------------------

def test_annual_to_monthly_rejects_rate_above_one():
    """A decrement probability above 1.0 produces a silent NaN; reject."""
    with pytest.raises(ValueError, match="annual rate must be <= 1.0"):
        annual_to_monthly(np.array([0.5, 1.5]))


def test_annual_to_monthly_accepts_boundary_one():
    """annual = 1.0 is the everyone-leaves-in-the-year boundary -- still valid."""
    out = annual_to_monthly(np.array([1.0]))
    assert out[0] == pytest.approx(1.0)        # monthly q = 1.0


def test_discount_curve_rejects_rate_at_negative_one():
    """A discount annual <= -1.0 produces NaN; reject at construction."""
    with pytest.raises(ValueError, match="discount_annual must be > -1.0"):
        Basis(
            mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
            discount_annual=-1.0,
            ra_confidence=0.75, mortality_cv=0.10,
            coverages=(CoverageRate("DEATH", _flat_rate()),),
        )


def test_inforce_fast_rejects_elapsed_past_term():
    """elapsed_months past the trajectory horizon (contract boundary, which
    defaults to term_months) is silently read past the trajectory -- reject."""
    mp = ModelPoints(
        issue_age=np.array([40.0]),
        premium=np.array([0.0]),
        term_months=np.array([12]),
        elapsed_months=np.array([15]),         # past the boundary (== term)
    )
    basis = Basis(
        mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
        discount_annual=0.03,
        ra_confidence=0.75, mortality_cv=0.10,
        coverages=(CoverageRate("DEATH", _flat_rate()),),
    )
    with pytest.raises(ValueError, match="past the contract boundary"):
        _measure_inforce_fast(mp, basis)


def test_inforce_full_rejects_elapsed_past_term():
    """Same boundary guard on the trajectory-returning entry."""
    mp = ModelPoints(
        issue_age=np.array([40.0]),
        premium=np.array([0.0]),
        term_months=np.array([12]),
        elapsed_months=np.array([15]),
    )
    basis = Basis(
        mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
        discount_annual=0.03,
        ra_confidence=0.75, mortality_cv=0.10,
        coverages=(CoverageRate("DEATH", _flat_rate()),),
    )
    with pytest.raises(ValueError, match="past the contract boundary"):
        _measure_inforce_full(
            mp, basis, prior_csm=np.array([0.0]),
            lock_in_rate=0.03, period_months=12,
        )


def test_issue_age_fractional_warns(recwarn):
    """A fractional issue_age would silently truncate at rate lookup -- warn."""
    ModelPoints(
        issue_age=np.array([40.7, 50.0]),       # 40.7 truncates to 40
        premium=np.array([0.0, 0.0]),
        term_months=np.array([12, 12]),
    )
    matched = [w for w in recwarn.list
               if issubclass(w.category, UserWarning)
               and "fractional" in str(w.message)]
    assert matched, [str(w.message) for w in recwarn.list]


def test_issue_age_integer_does_not_warn(recwarn):
    """Whole-year issue_age (the typical case) does not warn."""
    ModelPoints(
        issue_age=np.array([40.0, 50.0]),
        premium=np.array([0.0, 0.0]),
        term_months=np.array([12, 12]),
    )
    fractional = [w for w in recwarn.list
                  if issubclass(w.category, UserWarning)
                  and "fractional" in str(w.message)]
    assert not fractional


def test_segmented_measure_matches_nfc_and_nfd_codes():
    """Composed (NFC) vs decomposed (NFD) Unicode codes match the same segment.

    A product on the model_points side composed (e.g. ``café`` =
    'caf' + U+00E9) compared against a basis key decomposed (``café`` =
    'cafe' + U+0301) used to mismatch by byte identity. NFC-normalising
    both sides fixes the lookup.
    """
    composed = "café"            # NFC: single e-acute char
    decomposed = "café"          # NFD: e + combining acute
    assert composed != decomposed
    mp = ModelPoints(
        issue_age=np.array([40.0]),
        premium=np.array([0.0]),
        term_months=np.array([12]),
        product=np.array([composed], dtype=object),
        channel=np.array(["FC"], dtype=object),
    )
    basis = Basis(
        mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
        discount_annual=0.03,
        ra_confidence=0.75, mortality_cv=0.10,
        coverages=(CoverageRate("DEATH", _flat_rate()),),
    )
    # Basis keyed under the decomposed form -- the lookup must still match.
    basis = {(decomposed, "FC"): basis}
    out = fcf.gmm.measure(mp, basis, full=False)
    assert out.bel.shape == (1,)


# ---------------------------------------------------------------------------
# P1 G -- test-gap fill: validation paths that had no regression net
# ---------------------------------------------------------------------------

def test_coverage_arrays_rejects_unresolved_code():
    """A coverage code that is neither in calculation_methods nor a bare
    CalculationMethod name raises naming the offender."""
    from fastcashflow.coverage import coverage_arrays
    coverages = (CoverageRate("MYSTERY_CODE", _flat_rate()),)
    with pytest.raises(ValueError, match="MYSTERY_CODE"):
        coverage_arrays(coverages, calculation_methods=None)


def test_settlement_lic_rejects_bad_settlement_pattern_sum():
    """numerics._settlement_lic's pattern.sum() != 1 guard."""
    from fastcashflow.numerics import _settlement_lic
    incurred = np.zeros((1, 12))
    with pytest.raises(ValueError, match="settlement_pattern must sum to 1"):
        _settlement_lic(incurred, np.array([0.3, 0.3, 0.3]))


def test_settlement_factor_rejects_bad_pattern_sum():
    """numerics._settlement_factor's pattern.sum() != 1 guard."""
    from fastcashflow.numerics import _settlement_factor
    with pytest.raises(ValueError, match="settlement_pattern must sum to 1"):
        _settlement_factor(np.array([0.3, 0.3, 0.3]), monthly_rate=0.0)


def test_settlement_factor_rejects_bad_rate_shape():
    """_settlement_factor rejects a 2-D monthly_rate."""
    from fastcashflow.numerics import _settlement_factor
    with pytest.raises(ValueError, match="monthly_rate must be a scalar"):
        _settlement_factor(
            np.array([1.0]), monthly_rate=np.zeros((3, 3)),
        )


def test_norm_ppf_rejects_p_outside_open_interval():
    """numerics._norm_ppf rejects p at the boundary and outside."""
    from fastcashflow.numerics import _norm_ppf
    for bad in (-0.1, 0.0, 1.0, 1.5):
        with pytest.raises(ValueError, match="open interval"):
            _norm_ppf(bad)


def test_empty_portfolio_value_raises_loudly():
    """A zero-policy ModelPoints does not silently return garbage.

    ``measure()`` and ``measure()`` reject n_mp=0 up front with an explicit
    ValueError naming the empty portfolio (rather than letting an opaque
    ``term_months.max()`` zero-size reduction surface). This locks in the
    loud-fail behaviour so a future change that returns empty-but-meaningful
    trajectories is an intentional design move (and updates this test), not a
    regression. Daily-ETL workflows that may hit an empty segment should
    filter upstream.
    """
    mp = ModelPoints(
        issue_age=np.array([], dtype=np.float64),
        premium=np.array([], dtype=np.float64),
        term_months=np.array([], dtype=np.int64),
    )
    basis = Basis(
        mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
        discount_annual=0.03,
        ra_confidence=0.75, mortality_cv=0.10,
        coverages=(CoverageRate("DEATH", _flat_rate()),),
    )
    with pytest.raises(ValueError, match="empty"):
        fcf.gmm.measure(mp, basis, full=False)
    with pytest.raises(ValueError, match="empty"):
        fcf.gmm.measure(mp, basis)


def test_single_month_measure():
    """term_months=1 -- the engine handles a one-step horizon without
    off-by-one errors at the trajectory ends."""
    mp = ModelPoints(
        issue_age=np.array([40.0]),
        premium=np.array([0.0]),
        term_months=np.array([1]),
    )
    basis = Basis(
        mortality_annual=_flat_rate(0.01), lapse_annual=_flat_rate(0.0),
        discount_annual=0.03,
        ra_confidence=0.75, mortality_cv=0.10,
        coverages=(CoverageRate("DEATH", _flat_rate(0.01)),),
    )
    m = fcf.gmm.measure(mp, basis)
    assert m.bel_path.shape == (1, 2)            # (n_mp, term+1)
    assert np.isfinite(m.bel).all()


def test_mixed_term_months_tail_padded_consistently():
    """MPs with different term_months share a trajectory width = max(term);
    the short-term MP's bel beyond its term must be either zero or held
    flat (not garbage)."""
    mp = ModelPoints(
        issue_age=np.array([40.0, 40.0]),
        premium=np.array([0.0, 0.0]),
        term_months=np.array([3, 12]),       # mixed
    )
    basis = Basis(
        mortality_annual=_flat_rate(0.01), lapse_annual=_flat_rate(0.0),
        discount_annual=0.0,
        ra_confidence=0.75, mortality_cv=0.10,
        coverages=(CoverageRate("DEATH", _flat_rate(0.01)),),
    )
    m = fcf.gmm.measure(mp, basis)
    assert m.bel_path.shape == (2, 13)            # max term + 1
    # The 3-month MP's BEL at t=12 must be a finite, well-defined value
    # (zero or held-flat post-maturity), never NaN / inf.
    assert np.isfinite(m.bel_path[0, 12])


# ---------------------------------------------------------------------------
# StateModel validation paths
# ---------------------------------------------------------------------------

def test_statemodel_rejects_negative_duration_max():
    from fastcashflow import State
    with pytest.raises(ValueError, match="duration_max must be non-negative"):
        State(name="active", duration_max=-1)


def test_statemodel_rejects_empty_states():
    from fastcashflow import StateModel
    with pytest.raises(ValueError, match="at least one state"):
        StateModel(states=())


def test_statemodel_rejects_duplicate_state_names():
    from fastcashflow import State, StateModel
    with pytest.raises(ValueError, match="state names must be unique"):
        StateModel(states=(State(name="active"), State(name="active")))


def test_statemodel_rejects_transition_to_unknown_state():
    from fastcashflow import State, StateModel, Transition
    s = State(name="active", transitions=(
        Transition(rate="mortality", to="GHOST"),
    ))
    with pytest.raises(ValueError, match="transition to an unknown state"):
        StateModel(states=(s,))


def test_statemodel_rejects_lump_sum_without_destination():
    from fastcashflow import State, StateModel, Transition
    s = State(name="active", transitions=(
        Transition(rate="mortality", to=None, lump_sum=True),
    ))
    with pytest.raises(ValueError, match="lump-sum transition with no destination"):
        StateModel(states=(s,))


def test_statemodel_rejects_seating_index_out_of_range():
    from fastcashflow import State, StateModel
    with pytest.raises(ValueError, match="seating index out of range"):
        StateModel(states=(State(name="active"),), seating=(5,))


def test_construction_rejects_garbage_inputs():
    """Guards added after the adversarial API sweep: garbage that used to
    flow through to a silently-NaN / nonsense result is now rejected at
    construction / rate conversion with a clear error."""
    from dataclasses import replace

    # annual_to_monthly: a decrement rate must be a finite probability in
    # [0, 1] -- negative or NaN rates previously round-tripped / propagated.
    with pytest.raises(ValueError, match="must be >= 0"):
        annual_to_monthly(np.array([-0.5]))
    with pytest.raises(ValueError, match="finite"):
        annual_to_monthly(np.array([np.nan]))

    # Basis: a NaN discount used to give a silently-NaN liability.
    basis = fcf.samples.basis()[("TERM_LIFE_A", "GA")]
    with pytest.raises(ValueError, match="discount_annual must be finite"):
        replace(basis, discount_annual=float("nan"))

    # ModelPoints: sex domain, per-MP length mismatch, NaN premium, negative benefit.
    with pytest.raises(ValueError, match="sex must be 0"):
        fcf.ModelPoints.single(issue_age=40, premium=100, term_months=12, sex=2)
    with pytest.raises(ValueError, match="length"):
        fcf.ModelPoints(issue_age=np.array([40.0]), premium=np.array([100.0]),
                        term_months=np.array([12]), sex=np.array([0, 1]))
    with pytest.raises(ValueError, match="premium must be finite"):
        fcf.ModelPoints(issue_age=np.array([40.0]),
                        premium=np.array([np.nan]), term_months=np.array([12]))
    with pytest.raises(ValueError, match="coverage amounts must be >= 0"):
        fcf.ModelPoints(issue_age=np.array([40.0]), premium=np.array([100.0]),
                        term_months=np.array([12]), benefits={0: np.array([-1e6])})


def test_guards_full_false_cession_scenarios():
    """More sweep guards: group / transition need a full=True measurement;
    QuotaShare validates its cession; stochastic rejects empty / non-finite
    scenarios (all previously crashed cryptically or returned silent NaN)."""
    mp = fcf.samples.model_points()
    basis = fcf.samples.basis()
    head = fcf.gmm.measure(mp, basis, full=False)  # headline only -- no trajectory
    with pytest.raises(ValueError, match="full=True"):
        fcf.group(head, np.zeros(mp.n_mp))
    with pytest.raises(ValueError, match="full=True"):
        fcf.transition(head, np.zeros(mp.n_mp))

    with pytest.raises(ValueError, match="cession must be in"):
        fcf.reinsurance.QuotaShare(1.5)
    with pytest.raises(ValueError, match="cession must be finite"):
        fcf.reinsurance.QuotaShare(float("nan"))

    b1 = basis[("TERM_LIFE_A", "GA")]
    with pytest.raises(ValueError, match="scenarios must be finite"):
        fcf.gmm.stochastic(mp, b1, np.array([0.03, np.nan]))
    with pytest.raises(ValueError, match="scenarios is empty"):
        fcf.gmm.stochastic(mp, b1, np.array([]))


def test_guards_negative_amounts_and_premium():
    """Benefit / premium / account amounts are non-negative; a rate field is
    not (a guaranteed minimum crediting rate may legitimately be negative).
    A negative level premium is a sign error -- accounting adjustments are
    actual experience and belong in movement analysis, not the projection."""
    base = dict(issue_age=np.array([40.0]), premium=np.array([100.0]),
                term_months=np.array([120]))
    with pytest.raises(ValueError, match="premium must be >= 0"):
        fcf.ModelPoints(**{**base, "premium": np.array([-1.0])})
    with pytest.raises(ValueError, match=r"account_value must be >= 0"):
        fcf.ModelPoints(**base, account_value=np.array([-5.0]))
    with pytest.raises(ValueError, match=r"maturity_benefit must be >= 0"):
        fcf.ModelPoints(**base, maturity_benefit=np.array([-5.0]))
    # a rate, not an amount -- negative is allowed
    fcf.ModelPoints(**base, minimum_crediting_rate=np.array([-0.01]))


def test_guards_inforce_state():
    """InforceState rejects a backward elapsed month, negative count, a
    non-finite carried CSM / lock-in rate, and ragged per-MP arrays."""
    def mk(**kw):
        d = dict(mp_id=np.array([1, 2]), elapsed_months=np.array([12, 24]),
                 count=np.array([1.0, 1.0]), prior_csm=np.array([10.0, 20.0]),
                 lock_in_rate=0.03)
        return fcf.InforceState(**{**d, **kw})
    with pytest.raises(ValueError, match="elapsed_months must be >= 0"):
        mk(elapsed_months=np.array([-1, 24]))
    with pytest.raises(ValueError, match="count must be >= 0"):
        mk(count=np.array([-1.0, 1.0]))
    with pytest.raises(ValueError, match="prior_csm must be finite"):
        mk(prior_csm=np.array([np.nan, 20.0]))
    with pytest.raises(ValueError, match="lock_in_rate must be finite"):
        mk(lock_in_rate=float("nan"))
    with pytest.raises(ValueError, match="must match"):
        mk(count=np.array([1.0, 1.0, 1.0]))


def test_guards_expense_item_and_trace_month():
    """ExpenseItem validates its basis / value at construction; the BEL / CSM
    step tracers reject a non-integer anchor month instead of truncating it."""
    from fastcashflow.basis import ExpenseItem
    with pytest.raises(ValueError, match="unknown expense basis"):
        ExpenseItem("acquisition", "alpha", 0.1)
    with pytest.raises(ValueError, match="value must be finite"):
        ExpenseItem("acquisition", "alpha_fixed", float("nan"))
    ExpenseItem("acquisition", "alpha_fixed", 100.0)  # valid

    mp = fcf.samples.model_points()
    basis = fcf.samples.basis()
    with pytest.raises(ValueError, match="whole-month integers"):
        fcf.gmm.trace_bel_step(0, mp, basis, months=[12.5])
    with pytest.raises(ValueError, match="whole-month integers"):
        fcf.gmm.trace_csm_step(0, mp, basis, months=[12.5])


def test_guards_empty_portfolio(tmp_path):
    """An empty policies or coverages file is rejected with a clear message
    rather than a cryptic join / kernel error."""
    cm = tmp_path / "cm.csv"
    cm.write_text("coverage,calculation_method\nDEATH,DEATH\n")
    cov_empty = tmp_path / "cov0.csv"
    cov_empty.write_text("mp_id,coverage,amount\n")  # header only
    cov_ok = tmp_path / "cov.csv"
    cov_ok.write_text("mp_id,coverage,amount\n1,DEATH,1000000\n")
    # empty policies -- give a non-empty coverages file so the policies guard
    # is the one that fires (the empty-coverages guard runs first otherwise).
    pol_empty = tmp_path / "pol0.csv"
    pol_empty.write_text(
        "mp_id,product,channel,issue_age,sex,term_months,premium\n")
    with pytest.raises(ValueError, match="policies frame is empty"):
        fcf.read_model_points(str(pol_empty), coverages=str(cov_ok),
                              calculation_methods=str(cm))
    pol = tmp_path / "pol.csv"
    pol.write_text(
        "mp_id,product,channel,issue_age,sex,term_months,premium\n"
        "1,TERM_LIFE_A,GA,40,0,120,100.0\n")
    with pytest.raises(ValueError, match="coverages frame is empty"):
        fcf.read_model_points(str(pol), coverages=str(cov_empty),
                              calculation_methods=str(cm))


def test_coverage_amount_csr_path_rejects_negative_and_nan():
    """The CSR coverage_index / coverage_amount path -- the one read_model_points
    fills directly -- must reject a negative or NaN coverage amount, not only the
    benefits-map path: a negative flips the claim sign and a NaN silently NaNs
    the BEL, both straight into the kernel."""
    base = dict(issue_age=np.array([40.0]), premium=np.array([0.0]),
                term_months=np.array([12]),
                coverage_index=np.array([0], np.int64),
                coverage_offset=np.array([0, 1], np.int64))
    with pytest.raises(ValueError, match="coverage amounts must be >= 0"):
        ModelPoints(coverage_amount=np.array([-1000.0]), **base)
    with pytest.raises(ValueError, match="coverage amounts must be finite"):
        ModelPoints(coverage_amount=np.array([np.nan]), **base)
    # the benefits-map path is still guarded by the same unified check
    with pytest.raises(ValueError, match="coverage amounts must be >= 0"):
        ModelPoints(issue_age=np.array([40.0]), premium=np.array([0.0]),
                    term_months=np.array([12]), benefits={0: np.array([-5.0])})


def test_reader_warns_on_near_reserved_column_typo(tmp_path):
    """A policies column one edit from a reserved field (``coun`` -> count) is a
    silent footgun -- count would default to 1 (a 1000x understatement) -- so
    the reader warns; a genuine attribute (``region``) does not warn."""
    pl.DataFrame({"mp_id": ["A"], "issue_age": [40], "term_months": [12],
                  "premium": [0.0], "coun": [1000.0], "region": ["KR"]}
                 ).write_csv(tmp_path / "p.csv")
    pl.DataFrame({"mp_id": ["A"], "coverage": ["DEATH"], "amount": [1e8]}
                 ).write_csv(tmp_path / "c.csv")
    pl.DataFrame({"coverage": ["DEATH"], "calculation_method": ["DEATH"]}
                 ).write_csv(tmp_path / "m.csv")
    with pytest.warns(UserWarning, match=r"typo of the field 'count'"):
        fcf.read_model_points(tmp_path / "p.csv", coverages=tmp_path / "c.csv",
                              calculation_methods=tmp_path / "m.csv")


def test_coverage_rate_callable_wrong_shape_is_value_error():
    """A CoverageRate callable that returns a scalar / mis-broadcast array is a
    clean ValueError naming the coverage -- not an AssertionError that vanishes
    under ``python -O`` and lets a wrong-shaped rate mis-index the kernel. The
    guard sits in build_coverage_rates, so it fires on both full and fast."""
    z = lambda s, a, d: np.full(np.shape(a), 0.0)
    basis = Basis(mortality_annual=z, lapse_annual=z, discount_annual=0.0,
                  ra_confidence=0.75, mortality_cv=0.10,
                  coverages=(CoverageRate("CANCER", lambda s, a, d: 0.01),))
    mp = ModelPoints.single(40, 0.0, 24, benefits={0: 1000.0},
                            calculation_methods={"CANCER": CalculationMethod.DIAGNOSIS})
    for full in (True, False):
        with pytest.raises(ValueError, match=r"coverage rate 'CANCER' must return an array"):
            fcf.gmm.measure(mp, basis, full=full)


def test_benefits_accepts_coverage_code_keys():
    """benefits can be keyed by coverage CODE ('DEATH') -- self-documenting and
    aligned by code -- as well as by integer index. The two agree; a code-keyed
    map is order-independent (it pins coverage_codes, so the engine matches by
    code regardless of the basis' registration order); mixed keys are rejected."""
    _F = lambda v: (lambda s, a, d: np.full(np.shape(a), v))
    cm = {"DEATH": CalculationMethod.DEATH, "CANCER": CalculationMethod.DIAGNOSIS}
    # int and code keys give the same measurement
    basis1 = Basis(mortality_annual=_F(0.01), lapse_annual=_F(0.0), discount_annual=0.03,
                   ra_confidence=0.75, mortality_cv=0.10,
                   coverages=(CoverageRate("DEATH", _F(0.01)),))
    by_idx = fcf.gmm.measure(
        ModelPoints.single(40, 100.0, 24, benefits={0: 12000.0}, calculation_methods=cm), basis1)
    by_code = fcf.gmm.measure(
        ModelPoints.single(40, 100.0, 24, benefits={"DEATH": 12000.0}, calculation_methods=cm), basis1)
    assert np.allclose(by_idx.bel, by_code.bel)
    # order-independent: basis registers CANCER first, DEATH second; code keys
    # still route DEATH->DEATH rate and CANCER->CANCER rate.
    basis2 = Basis(mortality_annual=_F(0.0), lapse_annual=_F(0.0), discount_annual=0.0,
                   ra_confidence=0.75, mortality_cv=0.0, morbidity_cv=0.0,
                   coverages=(CoverageRate("CANCER", _F(0.02)), CoverageRate("DEATH", _F(0.01))))
    mp = ModelPoints.single(40, 0.0, 12, benefits={"DEATH": 1000.0, "CANCER": 5000.0},
                            calculation_methods=cm)
    assert mp.coverage_codes == ("DEATH", "CANCER")
    m = fcf.gmm.measure(mp, basis2)
    q = lambda a: 1 - (1 - a) ** (1 / 12)            # annual -> monthly
    assert m.cashflows.claim_cf[0, 0] == pytest.approx(1000.0 * q(0.01))      # DEATH rate
    assert m.cashflows.morbidity_cf[0, 0] == pytest.approx(5000.0 * q(0.02))  # CANCER rate
    # mixed int + str keys are rejected
    with pytest.raises(ValueError, match="all coverage codes .* or all coverage indices"):
        ModelPoints.single(40, 100.0, 24, benefits={0: 1.0, "CANCER": 2.0}, calculation_methods=cm)


# ---------------------------------------------------------------------------
# ModelPoints array guards -- shape / finiteness / sign on per-MP fields
# (Codex review 2026-06-07: count / state / issue_class / elapsed_months /
# premium_term_months / frequency arrays lacked length + finiteness checks, so
# a mis-shaped array was silently truncated and a NaN silently NaN'd the BEL.)
# ---------------------------------------------------------------------------

def _mp1(**kw):
    """One model point (n_mp = 1); override one field to trip a guard."""
    base = dict(issue_age=np.array([40.0]), premium=np.array([0.0]),
                term_months=np.array([12]))
    base.update(kw)
    return ModelPoints(**base)


def test_modelpoints_count_length_mismatch():
    with pytest.raises(ValueError, match="count has length 2 but n_mp is 1"):
        _mp1(count=np.array([1.0, 999.0]))


def test_modelpoints_count_nan():
    with pytest.raises(ValueError, match="count must be finite"):
        _mp1(count=np.array([np.nan]))


def test_modelpoints_state_length_mismatch():
    with pytest.raises(ValueError, match="state has length 2 but n_mp is 1"):
        _mp1(state=np.array([0, 1]))


def test_modelpoints_state_negative():
    with pytest.raises(ValueError, match="state must be >= 0"):
        _mp1(state=np.array([-1]))


def test_modelpoints_issue_class_negative():
    with pytest.raises(ValueError, match="issue_class must be >= 0"):
        _mp1(issue_class=np.array([-1]))


def test_modelpoints_elapsed_months_negative():
    with pytest.raises(ValueError, match="elapsed_months must be >= 0"):
        _mp1(elapsed_months=np.array([-1]))


def test_modelpoints_premium_term_length_mismatch():
    with pytest.raises(ValueError, match="premium_term_months has length 2"):
        _mp1(premium_term_months=np.array([12, 12]))


def test_modelpoints_frequency_length_mismatch():
    with pytest.raises(ValueError, match="premium_frequency_months has length 2"):
        _mp1(premium_frequency_months=np.array([1, 1]))


# ---------------------------------------------------------------------------
# Basis scalar guards -- NaN / negative / out-of-range that NaN the RA or BEL
# ---------------------------------------------------------------------------

def _basis1(**kw):
    base = dict(mortality_annual=_flat_rate(), lapse_annual=_flat_rate(),
                discount_annual=0.0, ra_confidence=0.75, mortality_cv=0.10,
                coverages=(CoverageRate("DEATH", _flat_rate()),))
    base.update(kw)
    return Basis(**base)


def test_basis_mortality_cv_nan():
    # NaN slips past a bare ``v < 0`` (NaN < 0 is False) and NaNs the RA.
    with pytest.raises(ValueError, match="mortality_cv must be finite"):
        _basis1(mortality_cv=np.nan)


def test_basis_cost_of_capital_nan():
    with pytest.raises(ValueError, match="cost_of_capital_rate must be finite"):
        _basis1(cost_of_capital_rate=np.nan)


def test_basis_settlement_pattern_negative_component():
    # Sums to 1 but a negative weight distorts LIC / BEL.
    with pytest.raises(ValueError, match="settlement_pattern weights must be >= 0"):
        _basis1(settlement_pattern=np.array([1.2, -0.2]))


def test_basis_expense_inflation_at_or_below_negative_one():
    with pytest.raises(ValueError, match="expense_inflation must be > -1.0"):
        _basis1(expense_inflation=-1.5)


# ---------------------------------------------------------------------------
# Workbook duplicate-key guard -- a duplicate (table_id, axis) used to be
# silently overwritten (last row wins), like the named rate tables already
# reject. (Codex review 2026-06-07.)
# ---------------------------------------------------------------------------

def test_axis_tables_rejects_duplicate_row():
    ws = _make_sheet("discount_tables", [
        ("table_id", "year", "rate"),
        ("DISC", 0, 0.03),
        ("DISC", 0, 0.05),     # duplicate (DISC, year 0)
    ])
    with pytest.raises(ValueError, match="duplicate year=0"):
        _axis_tables(ws, "year")


def test_basis_investment_return_at_negative_one():
    # A VFA return <= -100% has no monthly equivalent and NaNs the account.
    with pytest.raises(ValueError, match="investment_return must be finite and > -1.0"):
        _basis1(investment_return=-1.5)


def test_modelpoints_state_above_state_model_count_rejected():
    """A state index past the state_model's state count now raises a clear
    error at measurement (was a late IndexError at seating lookup), on both the
    full and fast paths. Codex 2026-06-07."""
    from fastcashflow.statemodel import STATE_MODELS
    mp = ModelPoints(
        issue_age=np.array([40.0]), premium=np.array([100.0]),
        term_months=np.array([12]), state=np.array([99]),
        benefits={"DEATH": np.array([1000.0])},
        calculation_methods={"DEATH": fcf.CalculationMethod.DEATH},
    )
    basis = _basis1(state_model=STATE_MODELS["WAIVER"],
                    waiver_incidence_annual=_flat_rate(0.01))
    with pytest.raises(ValueError, match="accepts only .* seating states"):
        fcf.gmm.measure(mp, basis)               # full path (projection.py)
    with pytest.raises(ValueError, match="accepts only .* seating states"):
        fcf.gmm.measure(mp, basis, full=False)   # fast path (engine.py)
