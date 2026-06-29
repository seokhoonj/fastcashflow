"""Hand-calculation validation of the semi-Markov path.

Tests are intentionally tiny -- one contract, a couple of months, simple
rates -- so each BEL can be derived by hand and matched to ``measure()``.
"""
from __future__ import annotations

import numpy as np
import pytest

import fastcashflow as fcf
from fastcashflow.state_model import StateModel, State, Transition
from conftest import PATTERNS, annual_from_monthly as _annual


def _cancer_reincidence_model(sojourn_tracking_months: int) -> StateModel:
    return StateModel(states=(
        State("healthy", pays_premium=True, transitions=(
            Transition("mortality"),
            Transition("ci_incidence", to="post_first"),
            Transition("lapse"),
        )),
        State("post_first", sojourn_tracking_months=sojourn_tracking_months, transitions=(
            Transition("mortality"),
            Transition("ci_reincidence", to="post_second",
                       pays_lump_sum=True, sojourn_dependent=True),
        )),
        State("post_second", transitions=(
            Transition("mortality"),
        )),
    ), seating=(0, 1, 2))


def _flat_assumptions(*, ci_reincidence_fn) -> fcf.Basis:
    return fcf.Basis(
        mortality_annual=lambda s, a, d: np.full(d.shape, _annual(0.001)),
        lapse_annual=lambda s, a, d: np.full(d.shape, 0.0),
        ci_incidence_annual=lambda s, a, d: np.full(d.shape, _annual(0.005)),
        ci_reincidence_annual=ci_reincidence_fn,
        discount_annual=0.0,
        ra_confidence=0.5,
        mortality_cv=0.0,
        state_model=_cancer_reincidence_model(12),
        coverages=(fcf.CoverageRate("DEATH", lambda s, a, d: np.full(d.shape, _annual(0.001))),),
    )


def _single_contract(term_months: int, *, death_benefit: float = 10_000_000.0,
                     reincidence_benefit: float = 5_000_000.0) -> fcf.ModelPoints:
    return fcf.ModelPoints(
        issue_age=np.array([40], dtype=np.int64),
        benefits={"DEATH": np.array([death_benefit])},
        premium=np.array([0.0]),
        term_months=np.array([term_months], dtype=np.int64),
        disability_benefit=np.array([reincidence_benefit]),
        calculation_methods=PATTERNS,
    )


def test_one_month_only_death_claim():
    """Term = 1 month, reincidence rate = 0 everywhere: the BEL collapses to
    one month of pure death-claim cost. Hand calculation:

        in-force = 1.0 at t = 0
        claim rate per unit IF = mortality_monthly * death_benefit
                               = 0.001 * 10_000_000 = 10_000
        pc = 1.0 * 10_000 * dm  (dm = mid-month discount factor at t=0)

    With discount_annual = 0, dm = 1, so pc = 10_000.
    No other PV components fire (no premium, no annuity, no reincidence,
    no maturity), so bel = pc = 10_000.
    """
    basis = _flat_assumptions(
        ci_reincidence_fn=lambda s, a, p, sd: np.zeros_like(sd, dtype=float),
    )
    v = fcf.gmm.measure(_single_contract(1), basis, full=False)
    assert np.isclose(v.bel[0], 10_000.0), v.bel[0]


def test_one_month_with_reincidence_in_exclusion():
    """At t = 0 only the healthy state has in-force; the reincidence
    transition operates on post_first occupancy which is still zero.
    Even with a nonzero reincidence rate the first-month BEL must equal
    the pure-death-claim value of the prior test.
    """
    basis = _flat_assumptions(
        ci_reincidence_fn=lambda s, a, p, sd: np.full_like(sd, _annual(0.02),
                                                           dtype=float),
    )
    v = fcf.gmm.measure(_single_contract(1), basis, full=False)
    assert np.isclose(v.bel[0], 10_000.0), v.bel[0]


def test_two_month_first_diagnosis_no_reincidence():
    """Two months. Reincidence rate = 0 throughout, so post_first cohorts
    only drain via mortality. Hand calculation:

      t = 0: occ = {h:1.0, p1:0, p2:0}
              claim PV += 1.0 * 0.001 * 10M = 10_000

              edges:
                healthy -> post_first  prob = 0.999 * 0.005 = 0.004995
                healthy stays          prob = 0.999 * 0.995 = 0.994005
              after step:
                h          = 0.994005
                p1[tau=0]  = 0.004995

      t = 1: ift = 0.994005 + 0.004995 = 0.999
              claim PV += 0.999 * 0.001 * 10M = 9_990
              (mortality rides every state via the DEATH coverage)

    Total bel = 10_000 + 9_990 = 19_990.
    """
    basis = _flat_assumptions(
        ci_reincidence_fn=lambda s, a, p, sd: np.zeros_like(sd, dtype=float),
    )
    v = fcf.gmm.measure(_single_contract(2), basis, full=False)
    assert np.isclose(v.bel[0], 19_990.0), v.bel[0]


def test_one_month_reincidence_active_via_seating():
    """Seat the contract directly on post_first (ss = 1, cohort 0) and place
    the reincidence rate outside its exclusion window. Term = 1 month.

      t = 0: occ = {h:0, p1[tau=0]:1.0, p2:0}
        Death-claim PV (mortality on the whole portfolio) = 1.0 * 0.001 * 10M
                                                          = 10_000.
        Reincidence rate at cohort 0 is set to monthly 0.02 (after the
        prior mortality is taken in competing-decrement order):
            flow = 0.999 * 0.02 = 0.01998
        Reincidence lump-sum PV = flow * reincidence_benefit * dm
                               = 0.01998 * 5_000_000 = 99_900.
      bel = 10_000 + 99_900 = 109_900.
    """
    # Reincidence rate = 0.02 monthly = _annual(0.02) annual, but only at
    # cohort 0; later cohorts unused in a one-month term.
    def ci_rein(s, a, p, sd):
        return np.full_like(sd, _annual(0.02), dtype=float)

    basis = _flat_assumptions(ci_reincidence_fn=ci_rein)
    mp = fcf.ModelPoints(
        issue_age=np.array([40], dtype=np.int64),
        benefits={"DEATH": np.array([10_000_000.0])},
        premium=np.array([0.0]),
        term_months=np.array([1], dtype=np.int64),
        disability_benefit=np.array([5_000_000.0]),
        state=np.array([1], dtype=np.int64),    # seat on post_first,
        calculation_methods=PATTERNS,
    )
    v = fcf.gmm.measure(mp, basis, full=False)
    assert np.isclose(v.bel[0], 109_900.0), v.bel[0]


def test_reincidence_rate_zero_in_exclusion_window():
    """Seat on post_first cohort 0. With a 12-month exclusion (rate = 0 for
    sd < 12), the first month must look identical to the case where the
    rate is zero everywhere.
    """
    def ci_rein_with_excl(s, a, p, sd):
        return np.where(sd < 12, 0.0, _annual(0.02))

    def ci_rein_all_zero(s, a, p, sd):
        return np.zeros_like(sd, dtype=float)

    mp = fcf.ModelPoints(
        issue_age=np.array([40], dtype=np.int64),
        benefits={"DEATH": np.array([10_000_000.0])},
        premium=np.array([0.0]),
        term_months=np.array([1], dtype=np.int64),
        disability_benefit=np.array([5_000_000.0]),
        state=np.array([1], dtype=np.int64),
        calculation_methods=PATTERNS,
    )
    v_excl = fcf.gmm.measure(mp, _flat_assumptions(ci_reincidence_fn=ci_rein_with_excl), full=False)
    v_zero = fcf.gmm.measure(mp, _flat_assumptions(ci_reincidence_fn=ci_rein_all_zero), full=False)
    assert np.isclose(v_excl.bel[0], v_zero.bel[0])


# ---------------------------------------------------------------------------
# measure() <-> measure() parity (semi-Markov detailed projection)
# ---------------------------------------------------------------------------
#
# project_cashflows() drives measure() and used to be Markov-only. A
# cohort-aware detailed kernel mirrors measure()'s semi-Markov
# path. These tests confirm the two paths still produce identical headline
# numbers on the cancer-reincidence model, across single contracts and a
# mixed portfolio.


def _reincidence_assumptions(*, sojourn_tracking_months, exclusion_months,
                              reincidence_monthly):
    def ci_rein(s, a, p, sd):
        return np.where(sd < exclusion_months, 0.0, _annual(reincidence_monthly))
    return fcf.Basis(
        mortality_annual=lambda s, a, d: np.full(d.shape, _annual(0.001)),
        lapse_annual=lambda s, a, d: np.full(d.shape, _annual(0.005)),
        ci_incidence_annual=lambda s, a, d: np.full(d.shape, _annual(0.005)),
        ci_reincidence_annual=ci_rein,
        discount_annual=0.03,
        expense_inflation=0.02,
        expense_items=(
            fcf.ExpenseItem("acquisition", "per_policy",    200_000.0),
            fcf.ExpenseItem("maintenance", "per_policy",  40_000.0),
        ),
        ra_confidence=0.75,
        mortality_cv=0.10,
        state_model=_cancer_reincidence_model(sojourn_tracking_months),
        coverages=(fcf.CoverageRate("DEATH", lambda s, a, d: np.full(d.shape, _annual(0.001))),),
    )


def test_measure_value_agree_single_contract():
    """One contract, 36-month term, mid-exclusion -- measure().bel_path[:,0] must
    equal measure().bel within floating-point tolerance.
    """
    basis = _reincidence_assumptions(sojourn_tracking_months=12, exclusion_months=6,
                                     reincidence_monthly=0.01)
    mp = _single_contract(36)
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)


def test_measure_value_agree_mixed_portfolio():
    """50-contract portfolio across sexes / ages / terms / starting states.
    """
    rng = np.random.default_rng(7)
    n = 50
    mp = fcf.ModelPoints(
        issue_age=rng.integers(30, 55, n).astype(np.int64),
        sex=rng.integers(0, 2, n).astype(np.int64),
        benefits={"DEATH": rng.integers(10, 80, n) * 1_000_000.0},
        premium=np.zeros(n),
        term_months=rng.integers(60, 180, n).astype(np.int64),
        disability_benefit=rng.integers(5, 30, n) * 1_000_000.0,
        state=rng.integers(0, 3, n).astype(np.int64),
        calculation_methods=PATTERNS,
    )
    basis = _reincidence_assumptions(sojourn_tracking_months=24, exclusion_months=12,
                                     reincidence_monthly=0.008)
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)


def test_measure_value_agree_long_cohort():
    """Same portfolio shape as the mixed test but with a deeper cohort grid
    (D = 60) to exercise the long-tail absorbing semantics.
    """
    rng = np.random.default_rng(11)
    n = 20
    mp = fcf.ModelPoints(
        issue_age=rng.integers(30, 55, n).astype(np.int64),
        sex=rng.integers(0, 2, n).astype(np.int64),
        benefits={"DEATH": rng.integers(10, 80, n) * 1_000_000.0},
        premium=np.zeros(n),
        term_months=np.full(n, 120, dtype=np.int64),
        disability_benefit=rng.integers(5, 30, n) * 1_000_000.0,
        calculation_methods=PATTERNS,
    )
    basis = _reincidence_assumptions(sojourn_tracking_months=60, exclusion_months=24,
                                     reincidence_monthly=0.012)
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)



# ---------------------------------------------------------------------------
# Coverage rules + diagnosis coverages combined with semi-Markov (P)
# ---------------------------------------------------------------------------
#
# Real cancer-reincidence products combine the state-duration mechanism
# (reincidence exclusion) with contract-level coverage rules (a policy-level waiting /
# reduction period) and additional diagnosis coverages. State duration and policy
# duration are orthogonal axes that must work on the same contract.


def _reincidence_assumptions_with_extra_coverage(sojourn_tracking_months, exclusion_months,
                                                  extra_is_diagnosis, extra_rate):
    """Reincidence model plus one extra coverage whose rate is constant per month.

    ``extra_is_diagnosis`` picks between a single-payment diagnosis coverage
    (claims run off a depleting not-yet-diagnosed pool) and a recurring
    health coverage (claim_rate accumulates each month).
    """
    from fastcashflow.basis import CoverageRate

    def extra_fn(sex, age, dur):
        return np.full(dur.shape, _annual(extra_rate))

    base = _reincidence_assumptions(sojourn_tracking_months=sojourn_tracking_months,
                                     exclusion_months=exclusion_months,
                                     reincidence_monthly=0.01)
    return fcf.Basis(
        mortality_annual=base.mortality_annual,
        lapse_annual=base.lapse_annual,
        ci_incidence_annual=base.ci_incidence_annual,
        ci_reincidence_annual=base.ci_reincidence_annual,
        discount_annual=base.discount_annual,
        expense_items=base.expense_items,
        ra_confidence=base.ra_confidence,
        mortality_cv=base.mortality_cv,
        morbidity_cv=0.10,
        state_model=base.state_model,
        coverages=(
            fcf.CoverageRate(code="DEATH", rate=base.mortality_annual),
            fcf.CoverageRate(code="EXTRA", rate=extra_fn),
        ),
    )


def _portfolio_with_rule_coverage(n, seed, extra_waiting, extra_reduction_end,
                                  extra_reduction_factor,
                                  extra_is_diagnosis=False):
    """A small portfolio with one death coverage (rule-free) and one extra
    coverage (carrying the per-coverage rule). The DEATH coverage is at
    cov_idx 0, the extra coverage at cov_idx 1 -- their integer codes are
    positions in ``basis.coverages`` in registration order.
    """
    rng = np.random.default_rng(seed)
    # Build coverage_index / coverage_amount: two coverages per mp (DEATH then EXTRA).
    death_amount = rng.integers(10, 80, n) * 1_000_000.0
    extra_amount = rng.integers(3, 15, n) * 1_000_000.0
    coverage_index = np.empty(n * 2, np.int64)
    coverage_amount = np.empty(n * 2)
    coverage_offset = np.arange(0, n * 2 + 1, 2, np.int64)
    coverage_waiting = np.zeros(n * 2, np.int64)
    coverage_reduction_end = np.zeros(n * 2, np.int64)
    coverage_reduction_factor = np.ones(n * 2)
    for i in range(n):
        coverage_index[2 * i] = 0    # DEATH
        coverage_index[2 * i + 1] = 1  # EXTRA (first registered after DEATH)
        coverage_amount[2 * i] = death_amount[i]
        coverage_amount[2 * i + 1] = extra_amount[i]
        coverage_waiting[2 * i + 1] = extra_waiting
        coverage_reduction_end[2 * i + 1] = extra_reduction_end
        coverage_reduction_factor[2 * i + 1] = extra_reduction_factor
    extra_pattern = (fcf.CalculationMethod.DIAGNOSIS if extra_is_diagnosis
                     else fcf.CalculationMethod.MORBIDITY)
    return fcf.ModelPoints(
        issue_age=rng.integers(30, 55, n).astype(np.int64),
        sex=rng.integers(0, 2, n).astype(np.int64),
        premium=np.zeros(n),
        term_months=np.full(n, 60, dtype=np.int64),
        disability_benefit=rng.integers(5, 30, n) * 1_000_000.0,
        coverage_index=coverage_index,
        coverage_amount=coverage_amount,
        coverage_offset=coverage_offset,
        coverage_waiting=coverage_waiting,
        coverage_reduction_end=coverage_reduction_end,
        coverage_reduction_factor=coverage_reduction_factor,
        calculation_methods={"DEATH": fcf.CalculationMethod.DEATH, "EXTRA": extra_pattern},
    )


def test_semi_markov_with_waiting_period_on_coverage():
    """Reincidence model + recurring coverage with a 3-month waiting period.
    measure() and measure() must agree.
    """
    basis = _reincidence_assumptions_with_extra_coverage(
        sojourn_tracking_months=12, exclusion_months=6,
        extra_is_diagnosis=False, extra_rate=0.0008,
    )
    mp = _portfolio_with_rule_coverage(
        n=30, seed=13,
        extra_waiting=3, extra_reduction_end=0, extra_reduction_factor=1.0,
        extra_is_diagnosis=False,
    )
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)


def test_semi_markov_with_diagnosis_coverage():
    """Diagnosis coverage on top of the reincidence state machine. The
    coverage's claim runs off a depleting not-yet-diagnosed pool that must
    apply to the cohort-aware in-force trajectory.
    """
    basis = _reincidence_assumptions_with_extra_coverage(
        sojourn_tracking_months=12, exclusion_months=6,
        extra_is_diagnosis=True, extra_rate=0.0008,
    )
    mp = _portfolio_with_rule_coverage(
        n=30, seed=17,
        extra_waiting=0, extra_reduction_end=0, extra_reduction_factor=1.0,
        extra_is_diagnosis=True,
    )
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)


def test_semi_markov_with_diagnosis_and_waiting_and_reduction():
    """All three axes at once: state-duration reincidence + policy-duration
    waiting + reduction on a diagnosis coverage. Each lives on its own axis
    and the engine has to combine them correctly.
    """
    basis = _reincidence_assumptions_with_extra_coverage(
        sojourn_tracking_months=12, exclusion_months=6,
        extra_is_diagnosis=True, extra_rate=0.001,
    )
    mp = _portfolio_with_rule_coverage(
        n=25, seed=19,
        extra_waiting=6, extra_reduction_end=24, extra_reduction_factor=0.5,
        extra_is_diagnosis=True,
    )
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)


# ---------------------------------------------------------------------------
# DI recovery (R) -- semi-Markov re-entry from disabled to active
# ---------------------------------------------------------------------------
#
# Disability-income products are the canonical motivation for semi-Markov:
# the disabled -> active recovery (termination) rate is sharply duration-
# dependent. fastcashflow models this with a sojourn_dependent transition
# back to the source state and a four-arg ``disability_recovery_annual``
# rate. The disabled state is also a benefit state -- ``disability_income``
# is paid each month its occupancy is held.


def _di_model(sojourn_tracking_months: int) -> StateModel:
    return StateModel(states=(
        State("active", pays_premium=True, transitions=(
            Transition("mortality"),
            Transition("waiver_incidence", to="disabled"),
            Transition("lapse"),
        )),
        State("disabled", pays_periodic_benefit=True, sojourn_tracking_months=sojourn_tracking_months,
              transitions=(
                  Transition("mortality"),
                  Transition("disability_recovery", to="active",
                             sojourn_dependent=True),
              )),
    ), seating=(0, 1, 1))


def _di_assumptions(*, sojourn_tracking_months, recovery_monthly):
    def recovery(s, a, p, sd):
        return np.full(sd.shape, _annual(recovery_monthly), dtype=float)
    return fcf.Basis(
        mortality_annual=lambda s, a, d: np.full(d.shape, _annual(0.001)),
        lapse_annual=lambda s, a, d: np.full(d.shape, _annual(0.005)),
        waiver_incidence_annual=lambda s, a, d: np.full(
            d.shape, _annual(0.003)),
        disability_recovery_annual=recovery,
        discount_annual=0.03,
        ra_confidence=0.75,
        mortality_cv=0.10,
        disability_cv=0.20,
        state_model=_di_model(sojourn_tracking_months),
        coverages=(fcf.CoverageRate("DEATH", lambda s, a, d: np.full(d.shape, _annual(0.001))),),
    )


def test_di_recovery_hand_calc_one_month_seated_on_disabled():
    """Seat the contract directly on disabled (ss = 1, cohort 0). With
    monthly mortality 0.001, monthly recovery 0.05 at cohort 0, monthly
    disability income 1.0M, no discount and a one-month term:

      t = 0: occ[disabled][cohort 0] = 1.0
        Death claim (DEATH coverage, no death_benefit set here): 0
        Disability income paid: benefit_occ * disability_income * dm
                              = 1.0 * 1_000_000 * 1 = 1_000_000

    BEL = 1_000_000 (only disability income; no premium since seated on
    disabled which is pays_premium=False, no maturity, no death claim).
    """
    basis = fcf.Basis(
        mortality_annual=lambda s, a, d: np.full(d.shape, _annual(0.001)),
        lapse_annual=lambda s, a, d: np.full(d.shape, 0.0),
        waiver_incidence_annual=lambda s, a, d: np.full(d.shape, 0.0),
        disability_recovery_annual=lambda s, a, p, sd: np.full(
            sd.shape, _annual(0.05), dtype=float),
        discount_annual=0.0,
        ra_confidence=0.5,
        mortality_cv=0.0,
        disability_cv=0.0,
        state_model=_di_model(12),
        coverages=(fcf.CoverageRate("DEATH", lambda s, a, d: np.full(d.shape, _annual(0.001))),),
    )
    mp = fcf.ModelPoints(
        issue_age=np.array([45], dtype=np.int64),
        benefits={"DEATH": np.array([0.0])},
        premium=np.array([0.0]),
        term_months=np.array([1], dtype=np.int64),
        disability_income=np.array([1_000_000.0]),
        state=np.array([1], dtype=np.int64),
        calculation_methods=PATTERNS,
    )
    v = fcf.gmm.measure(mp, basis, full=False)
    assert np.isclose(v.bel[0], 1_000_000.0), v.bel[0]


def test_di_recovery_higher_rate_drains_disabled_occupancy_faster():
    """A higher recovery rate must drain the disabled cohort faster:
    after the same number of months the inforce-on-disabled (i.e. the
    benefit_occ that pays disability income) is strictly smaller with a
    higher recovery rate.
    """
    mp = fcf.ModelPoints(
        issue_age=np.array([45], dtype=np.int64),
        benefits={"DEATH": np.array([0.0])},
        premium=np.array([0.0]),
        term_months=np.array([24], dtype=np.int64),
        disability_income=np.array([1_000_000.0]),
        state=np.array([1], dtype=np.int64),
        calculation_methods=PATTERNS,
    )
    low = fcf.gmm.measure(mp, _di_assumptions(
        sojourn_tracking_months=24, recovery_monthly=0.01))
    high = fcf.gmm.measure(mp, _di_assumptions(
        sojourn_tracking_months=24, recovery_monthly=0.10))
    # By t = 6 months the high-recovery scenario has visibly drained the
    # disabled occupancy more than the low-recovery one.
    assert high.cashflows.disability_cf[0, 6] < low.cashflows.disability_cf[0, 6]
    # Over the whole term, total disability paid is smaller too.
    assert high.cashflows.disability_cf.sum() < low.cashflows.disability_cf.sum()


def test_di_recovery_measure_value_agree_mixed_portfolio():
    """50-contract DI portfolio with a duration-tapered recovery rate.
    measure() and measure() must agree.
    """
    def recovery(s, a, p, sd):
        # DI valuation-table shape: high recovery in early months, dropping
        # off sharply with claim duration.
        return np.where(sd < 3, _annual(0.20),
                        np.where(sd < 12, _annual(0.05),
                                 _annual(0.01)))
    basis = fcf.Basis(
        mortality_annual=lambda s, a, d: np.full(d.shape, _annual(0.001)),
        lapse_annual=lambda s, a, d: np.full(d.shape, _annual(0.005)),
        waiver_incidence_annual=lambda s, a, d: np.full(
            d.shape, _annual(0.003)),
        disability_recovery_annual=recovery,
        discount_annual=0.03,
        expense_inflation=0.02,
        expense_items=(
            fcf.ExpenseItem("acquisition", "per_policy",    200_000.0),
            fcf.ExpenseItem("maintenance", "per_policy",  40_000.0),
        ),
        ra_confidence=0.75,
        mortality_cv=0.10,
        disability_cv=0.20,
        state_model=_di_model(36),
        coverages=(fcf.CoverageRate("DEATH", lambda s, a, d: np.full(d.shape, _annual(0.001))),),
    )
    rng = np.random.default_rng(23)
    n = 50
    mp = fcf.ModelPoints(
        issue_age=rng.integers(30, 55, n).astype(np.int64),
        sex=rng.integers(0, 2, n).astype(np.int64),
        benefits={"DEATH": rng.integers(10, 80, n) * 1_000_000.0},
        premium=rng.integers(2, 10, n) * 10_000.0,
        term_months=rng.integers(60, 180, n).astype(np.int64),
        disability_income=rng.integers(3, 10, n) * 100_000.0,
        state=rng.integers(0, 2, n).astype(np.int64),
        calculation_methods=PATTERNS,
    )
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)


def _portfolio_with_rule_and_diagnosis_coverages(n, seed, rule_waiting,
                                                  rule_reduction_end,
                                                  rule_reduction_factor):
    """Portfolio with three coverages per mp: DEATH (rule-free) + a
    recurring coverage carrying a waiting/reduction rule (cov_idx = 1) + a
    diagnosis coverage with no rules (cov_idx = 2). Exercises both the
    coverage-rule pass and the diagnosis pass on the same model points.
    """
    rng = np.random.default_rng(seed)
    death_amount = rng.integers(10, 80, n) * 1_000_000.0
    recur_amount = rng.integers(3, 15, n) * 1_000_000.0
    diag_amount = rng.integers(2, 10, n) * 1_000_000.0
    coverage_index = np.empty(n * 3, np.int64)
    coverage_amount = np.empty(n * 3)
    coverage_offset = np.arange(0, n * 3 + 1, 3, np.int64)
    coverage_waiting = np.zeros(n * 3, np.int64)
    coverage_reduction_end = np.zeros(n * 3, np.int64)
    coverage_reduction_factor = np.ones(n * 3)
    for i in range(n):
        coverage_index[3 * i + 0] = 0   # DEATH
        coverage_index[3 * i + 1] = 1   # recurring coverage (rule)
        coverage_index[3 * i + 2] = 2   # diagnosis coverage
        coverage_amount[3 * i + 0] = death_amount[i]
        coverage_amount[3 * i + 1] = recur_amount[i]
        coverage_amount[3 * i + 2] = diag_amount[i]
        coverage_waiting[3 * i + 1] = rule_waiting
        coverage_reduction_end[3 * i + 1] = rule_reduction_end
        coverage_reduction_factor[3 * i + 1] = rule_reduction_factor
    return fcf.ModelPoints(
        issue_age=rng.integers(30, 55, n).astype(np.int64),
        sex=rng.integers(0, 2, n).astype(np.int64),
        premium=np.zeros(n),
        term_months=np.full(n, 60, dtype=np.int64),
        disability_benefit=rng.integers(5, 30, n) * 1_000_000.0,
        coverage_index=coverage_index,
        coverage_amount=coverage_amount,
        coverage_offset=coverage_offset,
        coverage_waiting=coverage_waiting,
        coverage_reduction_end=coverage_reduction_end,
        coverage_reduction_factor=coverage_reduction_factor,
        calculation_methods={
            "DEATH": fcf.CalculationMethod.DEATH,
            "recur": fcf.CalculationMethod.MORBIDITY,
            "diag":  fcf.CalculationMethod.DIAGNOSIS,
        },
    )


def test_semi_markov_with_rule_and_diagnosis_coverages_together():
    """The strongest parity case for the semi-Markov inforce-trajectory
    caching: a single portfolio where both the coverage-rule pass and
    the diagnosis pass fire on every contract. If the cached trajectory
    is mis-saved or mis-read by either pass, BEL will diverge between
    measure() and measure().
    """
    from fastcashflow.basis import CoverageRate

    def recur_rate(sex, age, dur):
        return np.full(dur.shape, _annual(0.0006))

    def diag_rate(sex, age, dur):
        return np.full(dur.shape, _annual(0.0009))

    base = _reincidence_assumptions(sojourn_tracking_months=12, exclusion_months=6,
                                     reincidence_monthly=0.01)
    basis = fcf.Basis(
        mortality_annual=base.mortality_annual,
        lapse_annual=base.lapse_annual,
        ci_incidence_annual=base.ci_incidence_annual,
        ci_reincidence_annual=base.ci_reincidence_annual,
        discount_annual=base.discount_annual,
        expense_items=base.expense_items,
        ra_confidence=base.ra_confidence,
        mortality_cv=base.mortality_cv,
        morbidity_cv=0.10,
        state_model=base.state_model,
        coverages=(
            CoverageRate(code="DEATH", rate=base.mortality_annual),
            CoverageRate(code="recur", rate=recur_rate),
            CoverageRate(code="diag", rate=diag_rate),
        ),
    )
    mp = _portfolio_with_rule_and_diagnosis_coverages(
        n=40, seed=29,
        rule_waiting=3, rule_reduction_end=12,
        rule_reduction_factor=0.6,
    )
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.allclose(m.bel_path[:, 0], v.bel)


def test_workbook_elapsed_axis_drives_semi_markov_reincidence(tmp_path):
    """End-to-end: a rate sheet with an ``elapsed`` column is loaded via the
    schema-flex reader, plugged into ``ci_reincidence_annual``, and fed
    through the semi-Markov engine. Swapping the sheet to all-zero
    reincidence changes the BEL -- proving the ``elapsed`` axis
    actually flows from the workbook into the ``(sex, age, year, cohort)``
    evaluation."""
    import openpyxl
    from fastcashflow.io import _flex_rate_table

    # Sheet with the new sojourn axis. Six-month exclusion window then a
    # flat 5% recurrence -- a recurrence-cancer waiting period of one half-year.
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("rates")
    ws.append(["table_id", "elapsed", "rate"])
    for elapsed in range(6):
        ws.append(["CAN_RE", elapsed, 0.0])      # exclusion window
    for elapsed in range(6, 24):
        ws.append(["CAN_RE", elapsed, _annual(0.05)])
    ws_zero = wb.create_sheet("rates_zero")
    ws_zero.append(["table_id", "elapsed", "rate"])
    for elapsed in range(24):
        ws_zero.append(["CAN_RE_Z", elapsed, 0.0])
    p = tmp_path / "rates.xlsx"
    wb.save(p)
    reload = openpyxl.load_workbook(p)
    reincidence_fn = _flex_rate_table(reload["rates"])["CAN_RE"]
    zero_fn = _flex_rate_table(reload["rates_zero"])["CAN_RE_Z"]

    basis = _flat_assumptions(ci_reincidence_fn=reincidence_fn)
    mp = _single_contract(term_months=24, death_benefit=10_000_000.0,
                          reincidence_benefit=5_000_000.0)
    # measure / value parity is the existing semi-Markov contract -- a
    # workbook-sourced reincidence rate keeps it.
    m, v = fcf.gmm.measure(mp, basis), fcf.gmm.measure(mp, basis, full=False)
    assert np.isclose(m.bel_path[0, 0], v.bel[0])
    # Swap to a zero-rate sheet -- the BEL must move because the elapsed
    # axis really drives the reincidence claim outflow.
    basis_zero = _flat_assumptions(ci_reincidence_fn=zero_fn)
    v_zero = fcf.gmm.measure(mp, basis_zero, full=False)
    assert not np.isclose(v.bel[0], v_zero.bel[0])


def test_report_service_expense_includes_disability_cf():
    """report()'s insurance service expense includes the disability income /
    lump-sum flow (a protection benefit, B120-B124), not only death / morbidity
    / expense. A DI book must not lose its disability flow from the service
    result and revenue (the disability_cf-omission seam fix)."""
    basis = _di_assumptions(sojourn_tracking_months=12, recovery_monthly=0.02)
    mp = fcf.ModelPoints(
        issue_age=np.array([45], dtype=np.int64),
        benefits={"DEATH": np.array([0.0])}, premium=np.array([1000.0]),
        term_months=np.array([60], dtype=np.int64),
        disability_income=np.array([1_000_000.0]),
        calculation_methods=PATTERNS)            # seated active by default
    m = fcf.gmm.measure(mp, basis, full=True)
    cf = m.cashflows
    assert cf.disability_cf.sum() > 0.0          # the book genuinely pays DI
    rep = fcf.report(m)
    expected = (cf.mortality_cf + cf.morbidity_cf + cf.disability_cf + cf.expense_cf)
    np.testing.assert_allclose(rep.insurance_service_expense, expected, rtol=1e-12)
    # the fix is non-vacuous: dropping disability_cf would give a different line
    without = cf.mortality_cf + cf.morbidity_cf + cf.expense_cf
    assert not np.allclose(rep.insurance_service_expense, without)
