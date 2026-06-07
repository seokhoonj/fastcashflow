"""Optional mortality improvement curve layer.

A workbook may carry an `improvement_tables` sheet with rows
``(table_id, year, factor)`` -- the cumulative improvement multiplier at
each policy year -- plus an optional `mortality_improvement_table` column
on `segments`. The reader wraps the mortality callable so it multiplies by
the improvement factor at lookup time. Missing reference = no improvement
(factor 1.0).
"""
from pathlib import Path

import numpy as np
import openpyxl

from fastcashflow import read_basis


def _build(path: Path, *, improvement_curve=None):
    """One-segment workbook, flat mortality 0.001, optional improvement curve."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    seg = wb.create_sheet("segments")
    cols = ["product", "channel", "mortality_table", "lapse_table",
            "discount_table", "inflation_table",
            "ra_confidence", "mortality_cv",
            "morbidity_cv"]
    row = ["TERM_A", "GA", "MORT", "LAPSE", "DISC", "INFL",
           0.75, 0.10, 0.10]
    if improvement_curve is not None:
        cols.append("mortality_improvement_table")
        row.append("IMPR")
    seg.append(cols)
    seg.append(row)

    rd = wb.create_sheet("coverages")
    rd.append(["coverage", "rate_table"])
    # No rate-driven coverages registered -- this segment has no claim
    # benefits; mortality_table only drives the in-force decrement.

    mt = wb.create_sheet("mortality_tables")
    mt.append(["table_id", "sex", "age", "rate"])
    for sex in (0, 1):
        for age in range(20, 81):
            mt.append(["MORT", sex, age, 0.001])

    for sn, header, value_row in [
        ("lapse_tables", ["table_id", "duration", "rate"], ["LAPSE", 0, 0.05]),
        ("discount_tables", ["table_id", "year", "rate"], ["DISC", 0, 0.03]),
        ("inflation_tables", ["table_id", "year", "rate"], ["INFL", 0, 0.02]),
    ]:
        s = wb.create_sheet(sn)
        s.append(header)
        s.append(value_row)

    if improvement_curve is not None:
        imp = wb.create_sheet("improvement_tables")
        imp.append(["table_id", "year", "factor"])
        for year, factor in enumerate(improvement_curve):
            imp.append(["IMPR", year, factor])

    wb.save(path)


def _segment(path):
    return read_basis(path).resolve(("TERM_A", "GA"))


def test_no_improvement_sheet(tmp_path):
    """Without the sheet (and without the segments column), mortality is unchanged."""
    p = tmp_path / "a.xlsx"
    _build(p)
    basis = _segment(p)
    s = np.array([0, 0, 0]); a = np.array([30, 30, 30]); d = np.array([0, 5, 10])
    out = basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))
    assert np.allclose(out, [0.001, 0.001, 0.001])


def test_improvement_curve_applied(tmp_path):
    """A decreasing curve scales mortality down at later durations."""
    p = tmp_path / "a.xlsx"
    # 1.5% annual improvement -- cumulative factor at year t is 0.985 ^ t
    curve = [0.985 ** t for t in range(20)]
    _build(p, improvement_curve=curve)
    basis = _segment(p)
    s = np.array([0, 0, 0]); a = np.array([30, 30, 30]); d = np.array([0, 5, 10])
    out = basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))
    expected = [0.001 * curve[t] for t in (0, 5, 10)]
    assert np.allclose(out, expected)


def test_improvement_clips_past_end(tmp_path):
    """Lookups beyond the curve's last year hold the last factor flat."""
    p = tmp_path / "a.xlsx"
    curve = [1.0, 0.9, 0.8]                                # only 3 years defined
    _build(p, improvement_curve=curve)
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([10])
    # year 10 clips to year 2 -> factor 0.8
    assert np.isclose(basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0], 0.001 * 0.8)


def test_improvement_only_touches_mortality(tmp_path):
    """Coverages / lapse stay at their base values -- v1 improvement is mortality-only."""
    # We don't have rate-driven coverages in this minimal workbook; check lapse instead.
    p = tmp_path / "a.xlsx"
    curve = [0.5] * 5                                       # extreme factor for visibility
    _build(p, improvement_curve=curve)
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    # mortality scaled
    assert np.isclose(basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0], 0.001 * 0.5)
    # lapse unaffected (no improvement applied)
    assert np.isclose(basis.lapse_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0], 0.05)
