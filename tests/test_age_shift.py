"""Optional age_shift columns on segments.

A segment may carry integer ``mortality_age_shift``, ``morbidity_age_shift``,
``waiver_age_shift`` columns. The reader wraps the corresponding base-rate
callable so its ``issue_age`` argument is shifted before lookup, letting
one base table be reused across cohorts with different selection /
underwriting effects. Missing column or 0 = no shift.
"""
from pathlib import Path

import numpy as np
import openpyxl

from fastcashflow import CoverageRate, RISK_MORTALITY, read_basis
from fastcashflow.basis import Basis


def _build_workbook(path: Path, *, mortality_age_shift=None,
                    morbidity_age_shift=None):
    """Tiny self-contained basis workbook with one segment."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # segments
    seg = wb.create_sheet("segments")
    cols = ["product", "channel", "mortality_table", "lapse_table",
            "discount_table", "inflation_table",
            "ra_confidence", "mortality_cv",
            "morbidity_cv"]
    if mortality_age_shift is not None:
        cols.append("mortality_age_shift")
    if morbidity_age_shift is not None:
        cols.append("morbidity_age_shift")
    seg.append(cols)
    row = ["TERM_A", None, "MORT", "LAPSE", "DISC", "INFL",
           0.75, 0.10, 0.10]
    if mortality_age_shift is not None:
        row.append(mortality_age_shift)
    if morbidity_age_shift is not None:
        row.append(morbidity_age_shift)
    seg.append(row)

    # coverages
    rd = wb.create_sheet("coverages")
    rd.append(["coverage", "rate_table"])
    rd.append(["INPATIENT", "HOSP"])

    # mortality table -- linear in age so a shift is easy to verify
    mt = wb.create_sheet("mortality_tables")
    mt.append(["table_id", "sex", "age", "rate"])
    for sex in (0, 1):
        for age in range(20, 81):
            mt.append(["MORT", sex, age, 0.001 * (age - 20 + 1)])     # 0.001, 0.002, ...

    # coverage rate table for hosp -- linear in age too
    rr = wb.create_sheet("incidence_rate_tables")
    rr.append(["table_id", "sex", "age", "rate"])
    for sex in (0, 1):
        for age in range(20, 81):
            rr.append(["HOSP", sex, age, 0.01 * (age - 20 + 1)])

    # lapse (flat)
    lp = wb.create_sheet("lapse_tables")
    lp.append(["table_id", "duration", "rate"])
    lp.append(["LAPSE", 0, 0.05])

    # discount / inflation (flat)
    dt = wb.create_sheet("discount_tables")
    dt.append(["table_id", "year", "rate"])
    dt.append(["DISC", 0, 0.03])

    inf = wb.create_sheet("inflation_tables")
    inf.append(["table_id", "year", "rate"])
    inf.append(["INFL", 0, 0.02])

    wb.save(path)


def _segment(path):
    return next(iter(read_basis(path).values()))


def test_no_age_shift(tmp_path):
    """Missing age_shift columns leave rates untouched."""
    p = tmp_path / "a.xlsx"
    _build_workbook(p)
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.001 * 11      # age 30, base = 0.011
    assert basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.01 * 11         # hosp at age 30


def test_mortality_age_shift_positive(tmp_path):
    """A +5 shift treats every life as 5 years older in the mortality lookup."""
    p = tmp_path / "a.xlsx"
    _build_workbook(p, mortality_age_shift=5)
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    # mortality at apparent age 30 + 5 = 35 (sex 0)
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.001 * 16          # 0.016
    # morbidity (hosp) unaffected -- mortality_age_shift does not touch coverage rates
    assert basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.01 * 11             # 0.11


def test_mortality_age_shift_negative(tmp_path):
    """A -3 shift treats every life as 3 years younger."""
    p = tmp_path / "a.xlsx"
    _build_workbook(p, mortality_age_shift=-3)
    basis = _segment(p)
    s = np.array([0]); a = np.array([40]); d = np.array([0])
    # apparent age 40 - 3 = 37 -> rate 0.001 * 18 = 0.018
    assert np.isclose(basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0], 0.001 * 18)


def test_morbidity_age_shift_applies_to_all_coverages(tmp_path):
    """morbidity_age_shift shifts every rate-driven coverage, mortality untouched."""
    p = tmp_path / "a.xlsx"
    _build_workbook(p, morbidity_age_shift=2)
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    # mortality unaffected
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.001 * 11
    # hosp shifted -- apparent age 32 -> 0.01 * 13 = 0.13
    assert basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.01 * 13


def test_shifts_compose_independently(tmp_path):
    """mortality and morbidity shifts apply independently to their own tables."""
    p = tmp_path / "a.xlsx"
    _build_workbook(p, mortality_age_shift=5, morbidity_age_shift=2)
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.001 * 16        # 30 + 5 = 35
    assert basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.01 * 13            # 30 + 2 = 32


def test_with_age_shift_zero_is_identity(tmp_path):
    """shift = 0 returns the original callable object (no wrapping cost)."""
    p = tmp_path / "a.xlsx"
    _build_workbook(p, mortality_age_shift=0)
    basis = _segment(p)
    p2 = tmp_path / "b.xlsx"
    _build_workbook(p2)                                            # no column at all
    asmp2 = _segment(p2)
    # Both behave identically -- 0 shift is a no-op
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == asmp2.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0]
