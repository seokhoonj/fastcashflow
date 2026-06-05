"""Optional A/E factor layer.

A workbook may carry an ``ae_factors`` sheet with rows
``(product, channel, coverage, factor)`` plus optional axis columns
``{sex, age, issue_age, duration}``. The reader looks up the (product,
channel, coverage) factor and wraps the base rate so the engine
multiplies by the factor at lookup time. The layer applies to the
rate-driven coverages on the ``coverages`` sheet; the in-force decrement
``mortality_annual`` is not coverage-keyed, so its calibration goes
through the underlying mortality table directly. Missing sheet or
missing key = no adjustment (factor 1.0).
"""
from pathlib import Path

import numpy as np
import openpyxl

from fastcashflow import read_basis


def _build(path: Path, *, ae_rows=None):
    """A small basis workbook with one segment + an optional ae_factors sheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    seg = wb.create_sheet("segments")
    seg.append(["product", "channel", "mortality_table", "lapse_table",
                "discount_table", "inflation_table",
                "ra_confidence", "mortality_cv",
                "morbidity_cv"])
    seg.append(["TERM_A", "GA", "MORT", "LAPSE", "DISC", "INFL",
                0.75, 0.10, 0.10])

    rd = wb.create_sheet("coverages")
    rd.append(["coverage", "rate_table"])
    rd.append(["INPATIENT", "HOSP"])

    mt = wb.create_sheet("mortality_tables")
    mt.append(["table_id", "sex", "age", "rate"])
    for sex in (0, 1):
        for age in range(20, 81):
            mt.append(["MORT", sex, age, 0.001])

    rr = wb.create_sheet("incidence_rate_tables")
    rr.append(["table_id", "sex", "age", "rate"])
    for sex in (0, 1):
        for age in range(20, 81):
            rr.append(["HOSP", sex, age, 0.02])

    lp = wb.create_sheet("lapse_tables")
    lp.append(["table_id", "duration", "rate"])
    lp.append(["LAPSE", 0, 0.05])

    dt = wb.create_sheet("discount_tables")
    dt.append(["table_id", "year", "rate"])
    dt.append(["DISC", 0, 0.03])

    inf = wb.create_sheet("inflation_tables")
    inf.append(["table_id", "year", "rate"])
    inf.append(["INFL", 0, 0.02])

    if ae_rows is not None:
        ae = wb.create_sheet("ae_factors")
        ae.append(ae_rows[0])
        for r in ae_rows[1:]:
            ae.append(r)

    wb.save(path)


def _segment(path):
    return read_basis(path)[("TERM_A", "GA")]


def test_no_ae_factor_sheet(tmp_path):
    """Workbook with no ae_factors sheet leaves rates unchanged."""
    p = tmp_path / "a.xlsx"
    _build(p)
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.001
    assert basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.02


def test_scalar_ae_factor_per_coverage(tmp_path):
    """A single (product, channel, coverage, factor) row scales the coverage rate."""
    p = tmp_path / "a.xlsx"
    _build(p, ae_rows=[
        ["product", "channel", "coverage", "factor"],
        ["TERM_A", "GA", "INPATIENT", 1.5],          # 손해율 150%
    ])
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    # base 0.02 * factor 1.5 = 0.03
    assert np.isclose(basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0], 0.03)
    # mortality unchanged (no A/E row touches the mortality decrement)
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.001


def test_ae_factor_does_not_apply_to_mortality_decrement(tmp_path):
    """``mortality_annual`` is decrement-only; the A/E layer is coverage-keyed,
    so an A/E row whose ``coverage`` would match a registered death
    coverage does not adjust the in-force decrement rate (the death-claim
    payment rate is a separate ``coverages`` entry that does receive A/E)."""
    p = tmp_path / "a.xlsx"
    _build(p, ae_rows=[
        ["product", "channel", "coverage", "factor"],
        ["TERM_A", "GA", "INPATIENT", 1.5],
    ])
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    assert basis.mortality_annual(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.001


def test_ae_factor_varies_by_age(tmp_path):
    """A factor table with an age column captures age-dependent A/E (e.g., 20s anti-selection).

    The dense-grid rule applies to factor tables like to rate tables -- list
    every age, even when bands repeat. Three bands here: 25-29 = 3.0
    (heavy anti-selection), 30-49 = 1.5, 50-60 = 1.0 (ultimate).
    """
    rows = [["product", "channel", "coverage", "age", "factor"]]
    for age in range(25, 30):
        rows.append(["TERM_A", "GA", "INPATIENT", age, 3.0])
    for age in range(30, 50):
        rows.append(["TERM_A", "GA", "INPATIENT", age, 1.5])
    for age in range(50, 61):
        rows.append(["TERM_A", "GA", "INPATIENT", age, 1.0])
    p = tmp_path / "a.xlsx"
    _build(p, ae_rows=rows)
    basis = _segment(p)
    s = np.array([0, 0, 0]); a = np.array([25, 40, 60]); d = np.array([0, 0, 0])
    out = basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))
    assert np.allclose(out, [0.02 * 3.0, 0.02 * 1.5, 0.02 * 1.0])


def test_ae_factor_only_applies_to_matching_segment(tmp_path):
    """A factor for (TERM_A, FC, hosp) does NOT apply to (TERM_A, GA, hosp)."""
    p = tmp_path / "a.xlsx"
    _build(p, ae_rows=[
        ["product", "channel", "coverage", "factor"],
        ["TERM_A", "FC", "INPATIENT", 1.5],         # different channel
    ])
    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    # GA segment: no matching A/E -> base rate unchanged
    assert basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0] == 0.02


def test_ae_factor_composes_with_age_shift(tmp_path):
    """age_shift and A/E factor compose -- the shift moves the lookup, the factor scales it."""
    p = tmp_path / "a.xlsx"
    # Build a workbook with an age-varying rate table on a rate-driven
    # coverage (INPATIENT), apply morbidity_age_shift and an A/E factor for
    # the same code, and verify the two layers compose on that coverage's
    # rate. (The decrement ``mortality_annual`` is no longer coverage-keyed,
    # so the composition is exercised on the coverage rate instead.)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    seg = wb.create_sheet("segments")
    seg.append(["product", "channel", "mortality_table", "lapse_table",
                "discount_table", "inflation_table",
                "ra_confidence", "mortality_cv",
                "morbidity_cv", "morbidity_age_shift"])
    seg.append(["TERM_A", "GA", "MORT", "LAPSE", "DISC", "INFL",
                0.75, 0.10, 0.10, 5])
    rd = wb.create_sheet("coverages")
    rd.append(["coverage", "rate_table"])
    rd.append(["INPATIENT", "HOSP"])
    mt = wb.create_sheet("mortality_tables")
    mt.append(["table_id", "sex", "age", "rate"])
    for sex in (0, 1):
        for age in range(20, 81):
            mt.append(["MORT", sex, age, 0.001])
    rr = wb.create_sheet("incidence_rate_tables")
    rr.append(["table_id", "sex", "age", "rate"])
    for sex in (0, 1):
        for age in range(20, 81):
            rr.append(["HOSP", sex, age, 0.001 * (age - 20 + 1)])   # 0.001, 0.002, ...
    for sn, header, row in [
        ("lapse_tables", ["table_id", "duration", "rate"], ["LAPSE", 0, 0.05]),
        ("discount_tables", ["table_id", "year", "rate"], ["DISC", 0, 0.03]),
        ("inflation_tables", ["table_id", "year", "rate"], ["INFL", 0, 0.02]),
    ]:
        s_ws = wb.create_sheet(sn)
        s_ws.append(header)
        s_ws.append(row)
    ae = wb.create_sheet("ae_factors")
    ae.append(["product", "channel", "coverage", "factor"])
    ae.append(["TERM_A", "GA", "INPATIENT", 0.5])
    wb.save(p)

    basis = _segment(p)
    s = np.array([0]); a = np.array([30]); d = np.array([0])
    # apparent age 30 + 5 = 35 -> base rate 0.001 * 16 = 0.016; A/E 0.5 -> 0.008
    assert np.isclose(
        basis.coverages[0].rate(s, a, d, np.zeros_like(d), np.zeros_like(d))[0],
        0.001 * 16 * 0.5,
    )
