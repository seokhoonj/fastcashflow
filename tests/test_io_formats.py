"""Model-point I/O formats -- read_model_points reads xlsx and feather,
and picks up the optional per-coverage benefit-rule columns.

xlsx: a two-sheet (policies + coverages) workbook. feather: the Arrow IPC
format. Both round-trip to the same valuation as the bundled in-memory sample.
"""
import fastcashflow as fcf
import numpy as np
import openpyxl
import polars as pl
import pytest

from fastcashflow import read_model_points, write_measurement
from fastcashflow.gmm import measure
from conftest import mp_to_frames


def _write_sheets(path, sheets):
    """Write ``(name, polars-frame)`` pairs to an .xlsx file."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, frame in sheets:
        ws = wb.create_sheet(name)
        ws.append(list(frame.columns))
        for row in frame.iter_rows():
            ws.append(list(row))
    wb.save(path)


def test_read_xlsx(tmp_path):
    """A two-sheet .xlsx -- policies and coverages sheets in one workbook."""
    
    basis = next(iter(fcf.samples.basis().segments.values()))
    patterns = fcf.samples.calculation_methods()
    mps = fcf.samples.model_points()
    policies, coverages = mp_to_frames(mps, basis)
    path = tmp_path / "book.xlsx"
    _write_sheets(path, [("policies", policies), ("coverages", coverages)])

    back = read_model_points(path, calculation_methods=patterns)
    assert back.n_mp == mps.n_mp
    assert np.allclose(measure(back, basis, full=False).bel, measure(mps, basis, full=False).bel)


def test_read_feather(tmp_path):
    """A .feather (Arrow IPC) model-point file round-trips."""
    
    basis = next(iter(fcf.samples.basis().segments.values()))
    patterns = fcf.samples.calculation_methods()
    mps = fcf.samples.model_points()
    policies, coverages = mp_to_frames(mps, basis)
    pol_path = tmp_path / "policies.feather"
    cov_path = tmp_path / "coverages.feather"
    policies.write_ipc(pol_path)
    coverages.write_ipc(cov_path)

    back = read_model_points(pol_path, coverages=cov_path, calculation_methods=patterns)
    assert back.n_mp == mps.n_mp
    assert np.allclose(measure(back, basis, full=False).bel, measure(mps, basis, full=False).bel)


def test_write_measurement_feather(tmp_path):
    """write_measurement writes a .feather result file."""
    basis = next(iter(fcf.samples.basis().segments.values()))
    mps = fcf.samples.model_points()
    path = tmp_path / "results.feather"
    write_measurement(measure(mps, basis, full=False), path)
    assert path.exists()


def test_reads_coverage_benefit_rules(tmp_path):
    """The coverages frame reads the waiting / reduction columns."""
    
    basis = next(iter(fcf.samples.basis().segments.values()))
    patterns = fcf.samples.calculation_methods()
    mps = fcf.samples.model_points()
    policies, coverages = mp_to_frames(mps, basis)
    coverages = coverages.with_columns(
        pl.lit(6).alias("waiting"),
        pl.lit(24).alias("reduction_end"),
        pl.lit(0.5).alias("reduction_factor"),
    )
    pol_path = tmp_path / "policies.csv"
    cov_path = tmp_path / "coverages.csv"
    policies.write_csv(pol_path)
    coverages.write_csv(cov_path)

    back = read_model_points(pol_path, coverages=cov_path,
                             calculation_methods=patterns)
    assert np.all(back.coverage_waiting == 6)
    assert np.all(back.coverage_reduction_end == 24)
    assert np.allclose(back.coverage_reduction_factor, 0.5)


# ---------------------------------------------------------------------------
# elapsed_months -- inforce_state is the source of truth, the policies frame
# silently ignores the column. The warning surfaces a common misuse.
# ---------------------------------------------------------------------------

def test_policies_elapsed_months_emits_warning(tmp_path):
    """Same guard fires on the (policies + coverages) path."""
    import warnings
    
    basis = next(iter(fcf.samples.basis().segments.values()))
    patterns = fcf.samples.calculation_methods()
    mps = fcf.samples.model_points()
    policies, coverages = mp_to_frames(mps, basis)
    policies = policies.with_columns(pl.lit(12).alias("elapsed_months"))
    pol_path = tmp_path / "policies.csv"
    cov_path = tmp_path / "coverages.csv"
    policies.write_csv(pol_path)
    coverages.write_csv(cov_path)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        back = read_model_points(pol_path, coverages=cov_path,
                                 calculation_methods=patterns)
    msgs = [str(w.message) for w in caught if issubclass(w.category, UserWarning)]
    assert any("elapsed_months" in m for m in msgs), msgs
    assert np.all(back.elapsed_months == 0)


# ---------------------------------------------------------------------------
# read_vfa_model_points -- account-value contracts read from a single policies
# file (no coverages). The coverage-column guard keeps the wide footgun closed.
# ---------------------------------------------------------------------------

def test_read_vfa_model_points_round_trips(tmp_path):
    """A VFA policies file (account value + guarantee floors, no coverages)
    reads and measures the same as the bundled in-memory sample."""
    vmp = fcf.samples.model_points("vfa")
    vb = fcf.samples.basis("vfa")
    pl.DataFrame({
        "mp_id": [f"V{i + 1:03d}" for i in range(vmp.n_mp)],
        "issue_age": np.asarray(vmp.issue_age),
        "term_months": np.asarray(vmp.term_months),
        "account_value": np.asarray(vmp.account_value),
        "minimum_crediting_rate": np.asarray(vmp.minimum_crediting_rate),
        "minimum_death_benefit": np.asarray(vmp.minimum_death_benefit),
        "minimum_accumulation_benefit": np.asarray(vmp.minimum_accumulation_benefit),
    }).write_csv(tmp_path / "vfa.csv")

    back = fcf.read_vfa_model_points(tmp_path / "vfa.csv",
                                     calculation_methods=fcf.samples.calculation_methods())
    assert back.n_mp == vmp.n_mp
    assert np.allclose(fcf.vfa.measure(back, vb).bel, fcf.vfa.measure(vmp, vb).bel)
    assert np.allclose(fcf.vfa.measure(back, vb).csm, fcf.vfa.measure(vmp, vb).csm)


def test_read_vfa_model_points_rejects_coverage_columns(tmp_path):
    """A ``<code>_benefit`` column is a coverage encoded as a column -- VFA
    carries no coverages, so it is rejected rather than silently dropped."""
    pl.DataFrame({
        "issue_age": [40], "term_months": [120], "CANCER_benefit": [1e7],
    }).write_csv(tmp_path / "wide.csv")
    with pytest.raises(ValueError, match="coverage benefit column"):
        fcf.read_vfa_model_points(tmp_path / "wide.csv")
