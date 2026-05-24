"""Assumptions workbook reader.

A single ``assumptions.xlsx`` carries the segment mapping plus the named rate
tables. The ``segments`` sheet has a ``defaults`` row whose values blank
cells inherit, and one row per (product, channel) segment; the reader returns
one ``Assumptions`` per segment. See docs/assumptions-format.md.
"""
import numpy as np

from fastcashflow import (
    ModelPoints, load_sample_assumptions, measure, value,
)


def test_segments_resolve():
    """The sample workbook resolves to two segments -- TERM_A on GA and FC."""
    basis = load_sample_assumptions()
    assert set(basis) == {("TERM_A", "GA"), ("TERM_A", "FC")}


def test_defaults_inherited():
    """Blank cells in a segment row inherit from the ``defaults`` row."""
    basis = load_sample_assumptions()
    ga, fc = basis[("TERM_A", "GA")], basis[("TERM_A", "FC")]
    # ra_confidence / mortality_cv / morbidity_cv live only on the defaults row
    assert ga.ra_confidence == 0.75 and fc.ra_confidence == 0.75
    assert ga.mortality_cv == 0.10 and fc.mortality_cv == 0.10
    assert ga.morbidity_cv == 0.12 and fc.morbidity_cv == 0.12
    # shared economic / maintenance tables -- inherited identically
    assert ga.discount_annual == 0.03 and fc.discount_annual == 0.03
    assert ga.expense_inflation == 0.02 and fc.expense_inflation == 0.02
    assert ga.expense_maintenance_annual == 60_000.0
    assert fc.expense_maintenance_annual == 60_000.0


def test_channel_segmented_lapse():
    """GA and FC reference different lapse tables -- the per-segment table
    reference. GA persistency is worse than FC."""
    basis = load_sample_assumptions()
    dur = np.arange(6)
    zero = np.zeros_like(dur)
    ga_lapse = basis[("TERM_A", "GA")].lapse_annual(zero, zero, dur)
    fc_lapse = basis[("TERM_A", "FC")].lapse_annual(zero, zero, dur)
    assert np.all(ga_lapse > fc_lapse)


def test_per_segment_scalar():
    """``expense_acquisition`` is filled per segment row (GA vs FC commission)."""
    basis = load_sample_assumptions()
    assert basis[("TERM_A", "GA")].expense_acquisition == 150_000.0
    assert basis[("TERM_A", "FC")].expense_acquisition == 80_000.0


def test_riders_resolved():
    """Rate-driven riders resolve from ``rider_rate_tables``; non-rate-driven
    types stay in ``coverage_types`` only."""
    basis = load_sample_assumptions()
    asmp = basis[("TERM_A", "GA")]
    # adb is rate-driven (death-type), so it joins the riders tuple too.
    assert [r.code for r in asmp.riders] == ["hosp", "cancer", "adb"]
    assert asmp.coverage_types == {
        "dth_main": "death_main",
        "hosp": "morbidity",
        "cancer": "diagnosis",
        "adb": "death",
        "ann": "annuity",
        "mat": "maturity",
    }


def test_resolved_basis_values():
    """A resolved ``Assumptions`` runs through ``value`` and ``measure``; the
    GA and FC segments give different BEL because lapse differs (channel
    segmentation actually bites the valuation)."""
    basis = load_sample_assumptions()
    mp = ModelPoints.single(issue_age=40, death_benefit=100_000_000.0,
                            level_premium=50_000.0, term_months=120)
    ga = value(mp, basis[("TERM_A", "GA")]).bel[0]
    fc = value(mp, basis[("TERM_A", "FC")]).bel[0]
    assert np.isfinite(ga) and np.isfinite(fc)
    assert not np.isclose(ga, fc)
    # fused and detailed paths agree
    assert np.isclose(measure(mp, basis[("TERM_A", "GA")]).bel[0, 0], ga)


# ---------------------------------------------------------------------------
# state_model column + STATE_MODELS registry (U+W)
# ---------------------------------------------------------------------------


def test_state_model_column_resolves_to_registry_entry():
    """The sample workbook's ``defaults`` row carries
    ``state_model = WAIVER`` -- both segments inherit and resolve to
    ``STATE_MODELS['WAIVER']``.
    """
    from fastcashflow import STATE_MODELS
    basis = load_sample_assumptions()
    for asmp in basis.values():
        assert asmp.state_model is STATE_MODELS["WAIVER"]


def test_state_model_column_blank_keeps_none():
    """A segment row with a blank ``state_model`` cell falls back to the
    defaults row; an empty defaults cell leaves ``Assumptions.state_model``
    as ``None`` (matching the engine's pre-registry behaviour).
    """
    import openpyxl, tempfile, shutil
    import importlib.resources as resources
    from fastcashflow import read_assumptions
    # Copy the sample workbook and clear the defaults row's state_model.
    sample = resources.files("fastcashflow").joinpath(
        "sample_data/sample_assumptions.xlsx")
    with tempfile.TemporaryDirectory() as d:
        dst = f"{d}/blank_state.xlsx"
        shutil.copy(sample, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["segments"]
        col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "state_model":
                col = c; break
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).value = None
        wb.save(dst)
        basis = read_assumptions(dst)
        for asmp in basis.values():
            assert asmp.state_model is None


def test_state_model_unknown_key_raises():
    """An unrecognised state_model key is rejected at read time, with a
    hint listing the registry contents."""
    import openpyxl, tempfile, shutil, pytest
    import importlib.resources as resources
    from fastcashflow import read_assumptions
    sample = resources.files("fastcashflow").joinpath(
        "sample_data/sample_assumptions.xlsx")
    with tempfile.TemporaryDirectory() as d:
        dst = f"{d}/bad_state.xlsx"
        shutil.copy(sample, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["segments"]
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "state_model":
                ws.cell(row=2, column=c).value = "NOT_A_MODEL"
                break
        wb.save(dst)
        with pytest.raises(ValueError, match="NOT_A_MODEL"):
            read_assumptions(dst)
